#!/usr/bin/env python3
"""Ãƒâ€Ã‚Â°ddaaPro - Mackolik MS Oran Veri ÃƒÆ’Ã¢â‚¬Â¡ekici"""
from __future__ import annotations

import contextlib
import datetime as dt
import re
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import unquote, urljoin

import requests
from bs4 import BeautifulSoup

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from tkcalendar import DateEntry
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver import FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.support.ui import WebDriverWait

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Sabitler ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
BASE_URL  = "https://www.mackolik.com"
LIVE_URL  = f"{BASE_URL}/futbol/canli-sonuclar"

DATE_TOGGLE_CLASS   = "widget-dateslider__datepicker-toggle"
DATE_YEAR_SELECTOR  = "widget-datepicker__selector--year"
DATE_MONTH_SELECTOR = "widget-datepicker__selector--month"
DATE_VALUE_CLASS    = "widget-datepicker__value"
DATE_NAV_NEXT       = "widget-datepicker__nav--next"
DATE_NAV_PREV       = "widget-datepicker__nav--previous"
DATE_CALENDAR_BODY  = "widget-datepicker__calendar-body"

MATCH_ROW_CSS = "div[class*='match-row--']"
MATCH_CONTENT_XPATH = ".//div[contains(concat(' ', normalize-space(@class), ' '), ' match-row__match-content ')]"

POPUP_SELECTORS = [
    (By.XPATH,      "//button[@class='mobinterkapat']"),
    (By.CLASS_NAME, "rupclose"),
    (By.CLASS_NAME, "widget-gdpr-banner__accept"),
]

IDDAA_ICON_CSS = ".widget-stencilbar-livescore__icon--iddaa"
IDDAA_INPUT_CSS = "input[data-type='iddaa']"
SPORTS_CLASS    = "widget-stencilbar-livescore__sports"

MARKET_ITEM_CLASS = "widget-iddaa-markets__market-item"
IDDAA_CODE_CLASS  = "widget-iddaa-markets__iddaa-code"
OPTION_CLASS      = "widget-iddaa-markets__option"
MBS_CLASS         = "widget-iddaa-markets__mbc"
DATE_INFO_CLASS   = "p0c-soccer-match-details-header__info-date"
COMPETITION_CLASS = "p0c-soccer-match-details-header__competition-link"

HOME_CSS  = ".match-row__team-name--home"
AWAY_CSS  = ".match-row__team-name--away"
TIME_CSS  = ".match-row__start-time"
HT_CSS    = ".match-row__half-time-score"
FT_CSS    = ".match-row__score"

MONTH_ABBR = ["oca","sub","mar","nis","may","haz","tem","agu","eyl","eki","kas","ara"]

BASE_DIR = Path(__file__).resolve().parent
# -- HTTP Session (requests) -- Selenium yerine hizli market cekme ----------
_SESSION = requests.Session()
_SESSION.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'tr,en-US;q=0.7,en;q=0.3',
    'Connection': 'keep-alive',
})
# Connection pooling — ayni host'a baglantiyi yeniden kullan
from urllib3.util.retry import Retry as _Retry
from requests.adapters import HTTPAdapter as _HTTPAdapter
_adapter = _HTTPAdapter(pool_connections=5, pool_maxsize=10, max_retries=0)
_SESSION.mount('https://', _adapter)
_SESSION.mount('http://', _adapter)
_SESSION_READY = False

def _init_session_cookies(driver):
    global _SESSION_READY
    try:
        for c in driver.get_cookies():
            _SESSION.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
        _SESSION_READY = True
    except Exception:
        pass

# -- Mackolik JSON API ile mac listesi (Selenium'a gerek yok) --
_API_URL = "https://www.mackolik.com/perform/p0/ajax/components/competition/livescores/json?"

# API'den gelen competition verisini LEAGUE_LIST'e gore filtrele
# Sadece LEAGUE_LIST'teki ulke+lig kombinasyonlari kabul edilir
_BANNED_KEYWORDS = {'women', 'u23', 'u21', 'u19', 'u18', 'u17', 'reserve',
                    'youth', 'amateur', 'regional',
                    'non lig', 'non-league', 'non league',
                    'national league south', 'national league north',
                    'isthmian', 'southern league', 'northern league',
                    'a-lig kadin', 'a-league women',
                    'friendly', 'arkadas'}

# API ulke adlari (Ingilizce) → LEAGUE_LIST ulke adlari (Turkce) eslesmesi
_API_COUNTRY_TO_LEAGUE_LIST: dict[str, str] = {
    'turkey': 'TÜRKİYE', 'england': 'İNGİLTERE', 'spain': 'İSPANYA',
    'italy': 'İTALYA', 'germany': 'ALMANYA', 'france': 'FRANSA',
    'netherlands': 'HOLLANDA', 'portugal': 'PORTEKİZ', 'belgium': 'BELÇİKA',
    'austria': 'AVUSTURYA', 'czech republic': 'ÇEK CUMHURİYETİ',
    'czechia': 'ÇEK CUMHURİYETİ', 'denmark': 'DANİMARKA',
    'finland': 'FİNLANDİYA', 'croatia': 'HIRVATISTAN', 'scotland': 'İSKOÇYA',
    'sweden': 'İSVEÇ', 'switzerland': 'İSVİÇRE', 'hungary': 'MACARİSTAN',
    'norway': 'NORVEÇ', 'poland': 'POLONYA', 'romania': 'ROMANYA',
    'russia': 'RUSYA', 'serbia': 'SİRBİSTAN', 'greece': 'YUNANİSTAN',
    'usa': 'ABD', 'united states': 'ABD', 'brazil': 'BREZİLYA',
    'japan': 'JAPONYA', 'south korea': 'GÜNEY KORE',
    'korea republic': 'GÜNEY KORE', 'china': 'ÇİN',
    'australia': 'AVUSTRALYA',
    # Turkce API yanitlari icin birebir esleme
    'türkiye': 'TÜRKİYE', 'ingiltere': 'İNGİLTERE', 'ispanya': 'İSPANYA',
    'italya': 'İTALYA', 'almanya': 'ALMANYA', 'fransa': 'FRANSA',
    'hollanda': 'HOLLANDA', 'portekiz': 'PORTEKİZ', 'belçika': 'BELÇİKA',
    'avusturya': 'AVUSTURYA', 'çek cumhuriyeti': 'ÇEK CUMHURİYETİ',
    'danimarka': 'DANİMARKA', 'finlandiya': 'FİNLANDİYA',
    'hırvatistan': 'HIRVATISTAN', 'iskoçya': 'İSKOÇYA', 'isveç': 'İSVEÇ',
    'isviçre': 'İSVİÇRE', 'isvicre': 'İSVİÇRE', 'macaristan': 'MACARİSTAN',
    'norveç': 'NORVEÇ', 'polonya': 'POLONYA', 'romanya': 'ROMANYA',
    'rusya': 'RUSYA', 'sırbistan': 'SİRBİSTAN', 'yunanistan': 'YUNANİSTAN',
    'abd': 'ABD', 'brezilya': 'BREZİLYA', 'japonya': 'JAPONYA',
    'güney kore': 'GÜNEY KORE', 'çin': 'ÇİN', 'avustralya': 'AVUSTRALYA',
}

# LEAGUE_LIST'ten izin verilen ulke isimleri seti (lazy init)
_ALLOWED_COUNTRIES: set[str] | None = None

def _get_allowed_countries() -> set[str]:
    global _ALLOWED_COUNTRIES
    if _ALLOWED_COUNTRIES is None:
        # Hem orijinal key'leri hem normalize edilmiş halleri ekle
        raw = set(_API_COUNTRY_TO_LEAGUE_LIST.keys())
        normalized = {_normalize_country(k) for k in raw}
        _ALLOWED_COUNTRIES = raw | normalized
    return _ALLOWED_COUNTRIES

def _normalize_country(s: str) -> str:
    """Türkçe İ/ı sorununu çöz: 'İngiltere' → 'ingiltere' (Python İ.lower() = i̇ sorunu)."""
    s = s.replace("İ", "i").replace("ı", "i").replace("Ş", "s").replace("ş", "s")
    s = s.replace("Ğ", "g").replace("ğ", "g").replace("Ü", "u").replace("ü", "u")
    s = s.replace("Ö", "o").replace("ö", "o").replace("Ç", "c").replace("ç", "c")
    return s.lower().strip()

def _is_allowed_competition(comp: dict) -> bool:
    """API'den gelen competition LEAGUE_LIST'teki ulkelerden mi?
    Sadece ulke bazli on-filtre — asil lig filtresi sonra lig_filtreli_key ile yapilir."""
    if not comp:
        return False
    area_obj = comp.get("country") or comp.get("area") or {}
    area = _normalize_country((area_obj.get("name", "") or ""))
    league = (comp.get("name", "") or "").lower().strip()
    if any(b in league for b in _BANNED_KEYWORDS):
        return False
    if not area:
        return False
    return area in _get_allowed_countries()

def fetch_matches_api(target_date: dt.date, max_retries: int = 4, lig_filtre: set | None = None) -> list[dict]:
    """Mackolik JSON API ile bir gunun tum maclarini cek. Selenium gerektirmez.
    lig_filtre verilirse sadece o liglerdeki maçlar döner (erken filtreleme)."""
    params = {"sports[]": "Soccer", "matchDate": target_date.strftime("%Y-%m-%d")}
    data = {}
    for attempt in range(max_retries):
        try:
            resp = _SESSION.get(_API_URL, params=params, timeout=10)
            resp.raise_for_status()
            json_resp = resp.json()
            # API yaniti "data" altinda veya dogrudan kok seviyede olabilir
            data = json_resp.get("data", {})
            if not data or (not data.get("matches") and json_resp.get("matches")):
                data = json_resp
            break
        except Exception as e:
            if attempt < max_retries - 1:
                wait = (attempt + 1) * 2
                print(f'  [API] {target_date} deneme {attempt+1} hata: {e} — {wait}s bekle')
                time.sleep(wait)
            else:
                print(f'  [API] {target_date} tum denemeler basarisiz: {e}')
                return []

    matches_raw = data.get("matches", {})
    competitions = data.get("competitions", {})
    results = []

    # API bazen matches'i liste olarak donebilir
    if isinstance(matches_raw, list):
        matches_raw = {str(i): m for i, m in enumerate(matches_raw) if isinstance(m, dict)}
    if not isinstance(matches_raw, dict):
        print(f'  [API] {target_date} beklenmeyen matches tipi: {type(matches_raw).__name__}')
        return []

    # competitions da ayni sekilde
    if isinstance(competitions, list):
        competitions = {str(c.get("id", i)): c for i, c in enumerate(competitions) if isinstance(c, dict)}

    skipped_count = 0
    skipped_countries: set[str] = set()
    for mid, m in matches_raw.items():
        iddaa_code = str(m.get("iddaaCode", "") or "")
        if not iddaa_code or iddaa_code == "None":
            continue

        home = m.get("homeTeam", {}).get("name", "")
        away = m.get("awayTeam", {}).get("name", "")
        if not home or not away:
            continue

        comp_id = m.get("competitionId")
        comp = competitions.get(str(comp_id)) or competitions.get(comp_id) or {}

        # Tanimsiz ulke/lig filtreleme — sadece LEAGUE_LIST'teki kombinasyonlar gecerli
        if not _is_allowed_competition(comp):
            skipped_count += 1
            area_obj = comp.get("country") or comp.get("area") or {}
            area_name = (area_obj.get("name", "") or "").strip()
            if area_name:
                skipped_countries.add(area_name)
            continue

        country_obj = comp.get("country") or comp.get("area") or {}
        country = country_obj.get("name", "")
        league = comp.get("name", "")

        score_h = m.get("score", {}).get("home", "")
        score_a = m.get("score", {}).get("away", "")
        mac_skoru = f"{score_h}-{score_a}" if score_h != "" and score_a != "" else ""

        try:
            ht = m.get("score", {}).get("ht", {})
            home_ht = str(ht.get('home', '')).strip() if isinstance(ht, dict) else ''
            away_ht = str(ht.get('away', '')).strip() if isinstance(ht, dict) else ''
            iy_skor = f"{home_ht}-{away_ht}" if (home_ht or away_ht) else ""
        except Exception:
            iy_skor = ""

        # UTC timestamp → Istanbul saat
        mst_utc = m.get("mstUtc")
        mac_saati = ""
        mac_tarihi = target_date.strftime("%d.%m.%Y")
        if mst_utc:
            try:
                utc_dt = dt.datetime.fromtimestamp(mst_utc / 1000, tz=dt.timezone.utc)
                ist_dt = utc_dt + dt.timedelta(hours=3)
                mac_saati = ist_dt.strftime("%H:%M")
                mac_tarihi = ist_dt.strftime("%d.%m.%Y")
            except Exception:
                pass

        # Slug olustur — iddaa linki icin
        def _slug(name):
            s = name.lower()
            for old, new in [("\u0131","i"),("\u015f","s"),("\u011f","g"),("\u00e7","c"),("\u00fc","u"),("\u00f6","o"),
                             ("\u00e2","a"),("\u00ee","i"),("\u00e9","e"),("\u00e8","e"),("\u00e3","a"),
                             ("\u00ed","i"),("\u00f3","o"),("\u00fa","u"),("\u00e1","a"),("\u00f1","n"),
                             ("\u00e4","a"),("\u00eb","e"),("\u00ef","i"),
                             (" ","-"),(".","-"),("'",""),("\"","")]:
                s = s.replace(old, new)
            # Kalan unicode karakterleri temizle
            s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
            # Arka arkaya tire temizle
            s = re.sub(r'-{2,}', '-', s).strip('-')
            return s

        # API'den gelen slug varsa onu kullan, yoksa olustur
        api_slug = m.get("slug") or m.get("url") or ""
        if api_slug and "/mac/" in api_slug:
            iddaa_link = urljoin(BASE_URL, api_slug.rstrip("/") + "/iddaa")
            if "/iddaa/iddaa" in iddaa_link:
                iddaa_link = iddaa_link.replace("/iddaa/iddaa", "/iddaa")
        else:
            home_slug = _slug(home)
            away_slug = _slug(away)
            iddaa_link = f"https://www.mackolik.com/mac/{home_slug}-vs-{away_slug}/iddaa/{mid}"

        # Lig key olustur
        lig_key_val = _fold_lig(country) + " " + _fold_lig(league)

        # Erken lig filtresi — seçili liglerde değilse atla (detay HTTP request yapılmaz)
        if lig_filtre and not lig_filtreli_key(lig_key_val, lig_filtre):
            continue

        results.append({
            "ev_sahibi": home,
            "konuk_ekip": away,
            "mac_saati": mac_saati,
            "mac_tarihi": mac_tarihi,
            "ilk_yari_skor": iy_skor,
            "mac_skoru": mac_skoru,
            "iddaa_link": iddaa_link,
            "lig_key": lig_key_val,
            "lig": f"{country} {league}".strip(),
            "ms_kodu": iddaa_code,
        })

    if skipped_count:
        print(f'  [API] {target_date} → {len(results)} maç seçili liglerden, {skipped_count} maç atlandı (tanımsız lig)', flush=True)
        if skipped_countries:
            print(f'        Atlanan ülkeler: {", ".join(sorted(skipped_countries))}', flush=True)
    elif results:
        print(f'  [API] {target_date} → {len(results)} maç seçili liglerden', flush=True)

    return results


# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ YardÃƒâ€Ã‚Â±mcÃƒâ€Ã‚Â± fonksiyonlar ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def _norm(v: str | None) -> str:
    if not v:
        return ""
    return " ".join(v.replace("\xa0", " ").split()).strip()

def _norm_month(s: str) -> str:
    return (s.lower()
             .replace("\u015f", "s").replace("\u011f", "g").replace("\u00fc", "u")
             .replace("\u00f6", "o").replace("\u00e7", "c").replace("\u0131", "i"))[:3]

def _fold_text(s: str | None) -> str:
    if not s:
        return ""
    s = _norm(s)
    # ÖNEMLİ: Türkçe büyük harfleri .lower()'dan ÖNCE değiştir!
    # Python'da "İ".lower() = "i̇" (i + combining dot above, 2 karakter) → ASCII'ye çevrilince sorun çıkar.
    s = s.replace("İ", "i").replace("I", "i")
    s = s.lower()
    # ı (dotless-i) ASCII encode'da düşer → "yarı"→"yar" → eşleşmez. Türkçe karakterleri değiştir.
    s = (s.replace("ş", "s").replace("ğ", "g").replace("ü", "u")
          .replace("ö", "o").replace("ç", "c").replace("ı", "i"))
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return " ".join(s.split())

def _safe_text(el, css: str) -> str:
    try:
        child = el.find_element(By.CSS_SELECTOR, css)
        return _norm(child.text)
    except Exception:
        return ""

def _safe_href(block) -> str:
    try:
        for a in block.find_elements(By.TAG_NAME, "a"):
            try:
                href = a.get_attribute("href") or ""
                if "/mac/" in href:
                    return href
            except Exception:
                continue
    except Exception:
        pass
    return ""

def _js_click(driver, el) -> bool:
    if not el:
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", el)
        return True
    except Exception:
        return False

def _safe_find(parent, by, val):
    try:
        return parent.find_element(by, val)
    except (NoSuchElementException, StaleElementReferenceException):
        return None

def _safe_find_all(parent, by, val):
    try:
        return list(parent.find_elements(by, val))
    except (NoSuchElementException, StaleElementReferenceException):
        return []

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Driver ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def build_driver(headless: bool = True) -> webdriver.Firefox:
    opts = FirefoxOptions()
    opts.page_load_strategy = "eager"
    if headless:
        opts.add_argument("-headless")
    gd = BASE_DIR / "geckodriver.exe"
    svc = FirefoxService(executable_path=str(gd)) if gd.exists() else None
    drv = webdriver.Firefox(service=svc, options=opts) if svc else webdriver.Firefox(options=opts)
    drv.set_page_load_timeout(15)
    drv.implicitly_wait(0.2)
    return drv

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Popup ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def close_popups(driver) -> None:
    for by, val in POPUP_SELECTORS:
        el = _safe_find(driver, by, val)
        if el:
            _js_click(driver, el)
            time.sleep(0.1)

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Ãƒâ€Ã‚Â°ddaa / Futbol filtresi ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def enable_iddaa(driver) -> None:
    toggle = _safe_find(driver, By.CSS_SELECTOR, IDDAA_INPUT_CSS)
    if toggle and "active" in _norm(toggle.get_attribute("class")):
        return
    el = _safe_find(driver, By.CSS_SELECTOR, IDDAA_ICON_CSS)
    if el:
        _js_click(driver, el)

def enable_football(driver) -> None:
    sports = _safe_find(driver, By.CLASS_NAME, SPORTS_CLASS)
    if not sports:
        return
    items = _safe_find_all(sports, By.TAG_NAME, "li")
    if items and "active" not in _norm(items[0].get_attribute("class")):
        _js_click(driver, items[0])
        time.sleep(0.1)

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Sayfa aÃƒÆ’Ã‚Â§Ãƒâ€Ã‚Â±lÃƒâ€Ã‚Â±Ãƒâ€¦Ã…Â¸ ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def wait_rows(driver, timeout: int = 4) -> None:
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR, MATCH_ROW_CSS)) > 0
        )
    except TimeoutException:
        pass

def open_main_page(driver, max_retries: int = 3) -> None:
    for attempt in range(max_retries):
        try:
            driver.get(LIVE_URL)
            wait_rows(driver, timeout=5)
            close_popups(driver)
            enable_football(driver)
            return
        except Exception as e:
            print(f'  [open_main_page] deneme {attempt+1}/{max_retries} hata: {e}')
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                raise

# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Tarih Picker ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def _wait_picker(driver, timeout: int = 3) -> bool:
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: len(d.find_elements(By.CLASS_NAME, DATE_VALUE_CLASS)) > 0
        )
        return True
    except TimeoutException:
        return False

def _move_year(driver, target: int) -> None:
    for _ in range(30):
        sel = _safe_find(driver, By.CLASS_NAME, DATE_YEAR_SELECTOR)
        if not sel:
            raise RuntimeError("Yil selector bulunamadi")
        val_el = _safe_find(sel, By.CLASS_NAME, DATE_VALUE_CLASS)
        if not val_el:
            raise RuntimeError("Yil value bulunamadi")
        try:
            cur = int(_norm(val_el.text).strip())
        except ValueError:
            raise RuntimeError(f"Yil parse hatasi: {val_el.text!r}")
        if cur == target:
            return
        nav = _safe_find(sel, By.CLASS_NAME, DATE_NAV_NEXT if cur < target else DATE_NAV_PREV)
        if not nav:
            raise RuntimeError("Yil nav bulunamadi")
        _js_click(driver, nav)
        time.sleep(0.05)
    raise RuntimeError("Yil ayarlanamadi")

def _move_month(driver, target: int) -> None:
    for _ in range(20):
        sel = _safe_find(driver, By.CLASS_NAME, DATE_MONTH_SELECTOR)
        if not sel:
            raise RuntimeError("Ay selector bulunamadi")
        val_el = _safe_find(sel, By.CLASS_NAME, DATE_VALUE_CLASS)
        if not val_el:
            raise RuntimeError("Ay value bulunamadi")
        label = _norm_month(_norm(val_el.text))
        try:
            cur = MONTH_ABBR.index(label) + 1
        except ValueError:
            raise RuntimeError(f"Ay parse hatasi: {val_el.text!r}")
        if cur == target:
            return
        nav = _safe_find(sel, By.CLASS_NAME, DATE_NAV_NEXT if cur < target else DATE_NAV_PREV)
        if not nav:
            raise RuntimeError("Ay nav bulunamadi")
        _js_click(driver, nav)
        time.sleep(0.05)
    raise RuntimeError("Ay ayarlanamadi")

def _pick_day(driver, target: int) -> None:
    for attempt in range(3):
        body = _safe_find(driver, By.CLASS_NAME, DATE_CALENDAR_BODY)
        if not body:
            raise RuntimeError("Takvim bulunamadi")
        found = False
        for td in _safe_find_all(body, By.TAG_NAME, "td"):
            try:
                cls = _norm(td.get_attribute("class"))
                txt = _norm(td.text)
            except StaleElementReferenceException:
                break
            if "not-month-day" in cls:
                continue
            if txt == str(target):
                _js_click(driver, td)
                time.sleep(0.05)
                return
        else:
            raise RuntimeError("Gun secilemedi")
        time.sleep(0.1)
    raise RuntimeError("Gun secilemedi (3 deneme)")

def pick_date(driver, date: dt.date) -> None:
    el = _safe_find(driver, By.CLASS_NAME, DATE_TOGGLE_CLASS)
    if not el:
        raise RuntimeError("Toggle bulunamadi")
    _js_click(driver, el)
    time.sleep(0.1)
    if not _wait_picker(driver):
        raise RuntimeError("Picker acilmadi")
    _move_year(driver, date.year)
    _move_month(driver, date.month)
    _pick_day(driver, date.day)
    time.sleep(0.15)
    # Mackolik tarih sectikten sonra maclari JS ile yeniler
    # Yenilenmeyi bekle: dateslider uzerindeki aktif tarihin degismesini kontrol et
    target_str = f"{date.day:02d}/{date.month:02d}"
    for _ in range(15):
        try:
            active = driver.execute_script("""
                var dates = document.querySelectorAll('.widget-dateslider__date');
                for (var i = 0; i < dates.length; i++) {
                    if (dates[i].className.indexOf('active') > -1) {
                        var dd = dates[i].querySelector('.widget-dateslider__day-date');
                        return dd ? dd.textContent.trim() : '';
                    }
                }
                return '';
            """)
            if active == target_str:
                break
        except Exception:
            pass
        time.sleep(0.1)
    # Maclarin DOM'da render edilmesini bekle
    time.sleep(0.25)

# -- Mac listesi toplama (JavaScript ile HIZLI) --
def collect_summaries(driver) -> list[dict]:
    """Tum mac bilgilerini tek bir JS cagrisyla topla - 100x hizli."""
    JS_COLLECT = """
    var results = [];
    var seen = {};
    var currentLig = '';
    var rows = document.querySelectorAll("div[class*='match-row--']");
    for (var i = 0; i < rows.length; i++) {
        var row = rows[i];
        var links = row.querySelectorAll('a[href]');
        for (var li = 0; li < links.length; li++) {
            var href = links[li].getAttribute('href') || '';
            if (href.indexOf('/puan-durumu/') > -1) {
                try {
                    var p = href.split('mackolik.com/')[1].replace(/^\//, '');
                    var segs = p.split('/').filter(Boolean);
                    if (segs.length >= 2) currentLig = decodeURIComponent(segs[1]);
                } catch(e) {}
                break;
            } else if (href.indexOf('/futbol/') > -1 && href.indexOf('/mac/') === -1) {
                try {
                    var p2 = href.split('mackolik.com/')[1].replace(/^\//, '');
                    var segs2 = p2.split('/').filter(Boolean);
                    if (segs2.length >= 3) currentLig = decodeURIComponent(segs2[1]) + '/' + decodeURIComponent(segs2[2]);
                    else if (segs2.length >= 2) currentLig = decodeURIComponent(segs2[1]);
                } catch(e) {}
                break;
            }
        }
        var homeEl = row.querySelector('.match-row__team-name--home');
        var awayEl = row.querySelector('.match-row__team-name--away');
        if (!homeEl || !awayEl) continue;
        var home = (homeEl.textContent || '').trim();
        var away = (awayEl.textContent || '').trim();
        if (!home || !away) continue;
        var macHref = '';
        for (var ai = 0; ai < links.length; ai++) {
            var ah = links[ai].getAttribute('href') || '';
            if (ah.indexOf('/mac/') > -1) { macHref = ah; break; }
        }
        if (!macHref || seen[macHref]) continue;
        seen[macHref] = true;
        var timeEl = row.querySelector('.match-row__start-time');
        var htEl = row.querySelector('.match-row__half-time-score') || row.querySelector('[class*="half-time"]');
        var ftEl = row.querySelector('.match-row__score');
        results.push({
            ev_sahibi: home, konuk_ekip: away,
            mac_saati: timeEl ? timeEl.textContent.trim() : '',
            ilk_yari_skor: htEl ? htEl.textContent.replace(/IY|iy/g,'').trim() : '',
            mac_skoru: ftEl ? ftEl.textContent.trim() : '',
            mac_href: macHref, lig_raw: currentLig
        });
    }
    return results;
    """
    try:
        raw_list = driver.execute_script(JS_COLLECT)
    except Exception:
        return []

    summaries = []
    for item in (raw_list or []):
        href = item.get('mac_href', '')
        full = urljoin(BASE_URL, href)
        if '/iddaa/' not in full:
            idx = full.rfind('/')
            full = full[:idx] + '/iddaa' + full[idx:]
        full = full.replace('karsilastirma/', '')

        lig_raw = item.get('lig_raw', '')
        if '/' in lig_raw:
            parts = lig_raw.split('/')
            blok_lig_key = _fold_lig(parts[0]) + ' ' + _fold_lig(parts[1])
        else:
            blok_lig_key = _fold_lig(lig_raw)

        summaries.append({
            'ev_sahibi':     item.get('ev_sahibi', ''),
            'konuk_ekip':    item.get('konuk_ekip', ''),
            'mac_saati':     item.get('mac_saati', ''),
            'ilk_yari_skor': item.get('ilk_yari_skor', ''),
            'mac_skoru':     item.get('mac_skoru', ''),
            'iddaa_link':    full,
            'overview_link': full.replace('/iddaa', ''),
            'lig_key':       blok_lig_key,
        })
    return summaries

    for row in rows:
        try:
            blocks = row.find_elements(By.XPATH, MATCH_CONTENT_XPATH)
            # Mac bloku olmayan satirlar lig basligidir
            if not blocks:
                try:
                    for cl in row.find_elements(By.XPATH, ".//a[@href]"):
                        href = cl.get_attribute("href") or ""
                        if "/puan-durumu/" in href:
                            path = href.split("mackolik.com/", 1)[-1].strip("/")
                            parts = [unquote(x) for x in path.split("/") if x]
                            if len(parts) >= 2:
                                current_lig_key = _fold_lig(parts[1])
                                break
                        elif "/futbol/" in href:
                            path = href.split("mackolik.com/", 1)[-1].strip("/")
                            parts = [unquote(x) for x in path.split("/") if x]
                            if len(parts) >= 3:
                                current_lig_key = _fold_lig(parts[1]) + " " + _fold_lig(parts[2])
                                break
                except Exception:
                    pass
        except Exception:
            continue
        for block in blocks:
            try:
                home = _safe_text(block, HOME_CSS)
                away = _safe_text(block, AWAY_CSS)
                if not home or not away:
                    # Competition header block olabilir - lig_key guncelle
                    try:
                        for a in block.find_elements(By.XPATH, ".//a[@href]"):
                            h = a.get_attribute("href") or ""
                            if "/puan-durumu/" in h:
                                p = h.split("mackolik.com/", 1)[-1].strip("/")
                                pts = [unquote(x) for x in p.split("/") if x]
                                if len(pts) >= 2:
                                    current_lig_key = _fold_lig(pts[1])
                                    break
                            elif "/futbol/" in h:
                                p = h.split("mackolik.com/", 1)[-1].strip("/")
                                pts = [unquote(x) for x in p.split("/") if x]
                                if len(pts) >= 3:
                                    current_lig_key = _fold_lig(pts[1]) + " " + _fold_lig(pts[2])
                                    break
                    except Exception:
                        pass
                    continue
                href = _safe_href(block)
                if not href or href in seen:
                    continue
                seen.add(href)
                mac_saati = _safe_text(block, TIME_CSS)
                ht = _safe_text(block, HT_CSS).replace("\u0130Y","").replace("iy","").strip()
                ft = _safe_text(block, FT_CSS)
                raw_href = href
                full = urljoin(BASE_URL, raw_href)
                if "/iddaa/" not in full:
                    idx = full.rfind("/")
                    full = full[:idx] + "/iddaa" + full[idx:]
                full = full.replace("karsilastirma/", "")
                # Lig bilgisini: 1) blok icindeki /futbol/ linkinden, 2) maç URL'sinden, 3) current_lig_key'den al
                blok_lig_key = current_lig_key  # varsayılan: üstteki lig header'ından
                if not blok_lig_key:
                    print(f"[WARN] {home} - {away}: current_lig_key boş, blok linkleri taranacak")
                try:
                    # Blok içindeki tüm linkleri tara, puan-durumu veya futbol URL'si bul
                    for a in block.find_elements(By.XPATH, ".//a[@href]"):
                        h = a.get_attribute("href") or ""
                        if "/puan-durumu/" in h:
                            p = h.split("mackolik.com/", 1)[-1].strip("/")
                            pts = [unquote(x) for x in p.split("/") if x]
                            if len(pts) >= 2:
                                blok_lig_key = _fold_lig(pts[1])
                                break
                        elif "/futbol/" in h:
                            p = h.split("mackolik.com/", 1)[-1].strip("/")
                            pts = [unquote(x) for x in p.split("/") if x]
                            if len(pts) >= 3 and not pts[2].replace("-", "").isdigit():
                                blok_lig_key = _fold_lig(pts[1]) + " " + _fold_lig(pts[2])
                                break
                except Exception:
                    pass
                summaries.append({
                    "ev_sahibi":     home,
                    "konuk_ekip":    away,
                    "mac_saati":     mac_saati,
                    "ilk_yari_skor": ht,
                    "mac_skoru":     ft,
                    "iddaa_link":    full,
                    "overview_link": full.replace("/iddaa", ""),
                    "lig_key":       blok_lig_key,
                })
            except Exception:
                continue
    return summaries


# ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Detay sayfasÃƒâ€Ã‚Â± ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
def _extract_league(href: str) -> str:
    try:
        path = href.split("mackolik.com/", 1)[1].strip("/")
    except IndexError:
        return ""
    parts = [unquote(x) for x in path.split("/") if x]
    # parts[0] = "futbol", parts[1] = ülke, parts[2] = lig - sadece bunları al
    if len(parts) <= 1:
        return ""
    clean = [p for p in parts[1:3] if p and not p.replace("-", "").isdigit()]
    return " ".join(clean).replace("-", " ").title()

_COUNTRY_ALIASES = {
    "turkey": "turkiye", "england": "ingiltere", "spain": "ispanya",
    "italy": "italya", "germany": "almanya", "france": "fransa",
    "netherlands": "hollanda", "portugal": "portekiz", "belgium": "belcika",
    "austria": "avusturya", "czech republic": "cek cumhuriyeti", "czechia": "cek cumhuriyeti",
    "denmark": "danimarka", "finland": "finlandiya", "croatia": "hirvatistan",
    "scotland": "iskocya", "sweden": "isvec", "switzerland": "isvicre",
    "hungary": "macaristan", "norway": "norvec", "poland": "polonya",
    "romania": "romanya", "russia": "rusya", "serbia": "sirbistan",
    "greece": "yunanistan", "united states": "abd", "usa": "abd",
    "brazil": "brezilya", "japan": "japonya", "south korea": "guney kore",
    "china": "cin", "australia": "avustralya",
}

# Yaygın lig adı varyasyonları normalizasyonu (Ingilizce → Turkce / standart)
_LEAGUE_NAME_ALIASES: dict[str, str] = {
    'league': 'lig', 'liga': 'lig', 'primera': '1',
    'segunda': '2', 'division': 'lig',
    'first': '1', 'second': '2',
    'one': '1', 'two': '2', 'three': '3', 'four': '4',
    'national': 'ulusal',
}
# Sponsor/dönemsel kelimeler — lig karşılaştırmasında görmezden gelinir
# "Trendyol Süper Lig" ve "Süper Lig" aynı lig olarak eşleşmeli
_SPONSOR_WORDS = {'trendyol', 'spor', 'toto', 'stsl', 'misli', 'bilyoner',
                  'cemil', 'usta', 'turkcell', 'digiturk', 'bein',
                  'ptt', 'tff', 'nesine', 'bank', 'asya', 'sigorta',
                  # Ek sponsor / isim degisikligi kelimeleri (2020-2026)
                  'vodafone', 'ziraat', 'thy', 'betsson', 'macron',
                  'parions', 'uber', 'eats', 'ea', 'tim',
                  'scotiabank', 'apertura', 'clausura',
                  }

_COUNTRY_TOP_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'turkiye': ('super lig',),
    'ingiltere': ('premier lig',),
    'ispanya': ('laliga', 'la liga', 'primera division'),
    'italya': ('serie a',),
    'almanya': ('bundesliga',),
    'fransa': ('ligue 1',),
    'hollanda': ('eredivisie',),
    'portekiz': ('premier lig', 'primeira liga'),
    'belcika': ('pro lig', 'first division a'),
    'avusturya': ('bundesliga',),
    'cek cumhuriyeti': ('czech liga', 'first liga'),
    'danimarka': ('superliga',),
    'finlandiya': ('veikkausliiga',),
    'hirvatistan': ('1 hnl',),
    'iskocya': ('premiership',),
    'isvec': ('allsvenskan',),
    'isvicre': ('super lig', 'super league'),
    'macaristan': ('nb i',),
    'norvec': ('eliteserien',),
    'polonya': ('ekstraklasa',),
    'romanya': ('liga 1', 'liga i'),
    'rusya': ('premier lig',),
    'sirbistan': ('super lig',),
    'yunanistan': ('super lig',),
    'abd': ('mls', 'major league soccer'),
    'brezilya': ('serie a', 'brasileirao'),
    'japonya': ('j1 ligi', 'j1 lig'),
    'guney kore': ('k lig 1',),
    'cin': ('super lig',),
    'avustralya': ('a lig', 'aleague'),
}

_COUNTRY_SECOND_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'turkiye': ('1 lig',),
    'ingiltere': ('championship',),
    'ispanya': ('laliga 2', 'segunda division'),
    'italya': ('serie b',),
    'almanya': ('2 bundesliga',),
    'fransa': ('ligue 2',),
    'hollanda': ('eerste divisie',),
    'portekiz': ('2 lig', 'segunda liga'),
    'belcika': ('challenger pro lig', 'first division b'),
    'avusturya': ('1 lig', 'erste liga'),
    'danimarka': ('1 lig', '1 division'),
    'finlandiya': ('ykkosliiga', 'ykkonen'),
    'hirvatistan': ('2 hnl',),
    'iskocya': ('championship',),
    'isvicre': ('challenge lig', 'challenge league'),
    'macaristan': ('nb ii', '2 lig'),
    'norvec': ('1 lig', 'obos ligaen'),
    'polonya': ('1 lig', 'i liga'),
    'romanya': ('2 lig', 'liga 2', 'liga ii'),
    'rusya': ('fnl', 'first league'),
    'sirbistan': ('1 lig',),
    'yunanistan': ('2 lig', 'super league 2'),
}

_COUNTRY_THIRD_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'fransa': ('ulusal lig 1', 'national 1'),
    'ingiltere': ('league one', '1 lig', 'lig 1'),
    'polonya': ('2 lig', 'ii liga'),
    'iskocya': ('3 lig', 'league one'),
    'isvec': ('3 lig',),
}

_COUNTRY_FOURTH_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'ingiltere': ('league two', '2 lig', 'lig 2'),
    'iskocya': ('4 lig', 'league two'),
}

def _fold_lig(s: str) -> str:
    """Lig adı/slug normalizer — nokta ve tire → boşluk, Türkçe → ASCII, İngilizce ülke → Türkçe.
    Sponsor kelimelerini de temizler (yildan yila degisen lig isimleri icin)."""
    r = _fold_text(s.replace(".", " ").replace("-", " "))
    # Ingilizce ulke adini Turkce karsiligina cevir
    for eng, tr in _COUNTRY_ALIASES.items():
        if r == eng:
            return tr
        if r.startswith(eng + " "):
            r = tr + r[len(eng):]
            break
    # Lig adi varyasyonlarini normalize et (league→lig, liga→lig vb.)
    words = r.split()
    words = [_LEAGUE_NAME_ALIASES.get(w, w) for w in words]
    # Sponsor kelimelerini temizle (yildan yila degisen isimler icin)
    words = [w for w in words if w not in _SPONSOR_WORDS]
    return " ".join(words) if words else " ".join(r.split())

def lig_key(country_display: str, league_display: str) -> str:
    """LEAGUE_LIST'teki bir giriş için filtre anahtarı üret."""
    return _fold_lig(country_display) + " " + _fold_lig(league_display)

def _split_known_lig_key(value: str) -> tuple[str, str]:
    """Normalize key'i bilinen ülke prefix'ine göre (ülke, lig) ayır."""
    value = _fold_lig(value)
    for country, _ in LEAGUE_LIST:
        ckey = _fold_lig(country)
        prefix = ckey + " "
        if value == ckey:
            return ckey, ""
        if value.startswith(prefix):
            return ckey, value[len(prefix):].strip()
    return "", value

def _league_tier(country_key: str, league_text: str) -> int | None:
    league_norm = _fold_lig(league_text)
    league_core = " ".join(w for w in league_norm.split() if w not in _SPONSOR_WORDS)

    def has_any(*phrases: str) -> bool:
        return any(p in league_core for p in phrases)

    if has_any(*_COUNTRY_FOURTH_TIER_HINTS.get(country_key, ())):
        return 4
    if has_any(*_COUNTRY_THIRD_TIER_HINTS.get(country_key, ())):
        return 3
    if has_any(*_COUNTRY_SECOND_TIER_HINTS.get(country_key, ())):
        return 2
    if has_any(*_COUNTRY_TOP_TIER_HINTS.get(country_key, ())):
        return 1

    numeric_patterns = [
        (r'\b4\s*(?:lig|liga|division|divisie|hnl)\b', 4),
        (r'\b3\s*(?:lig|liga|division|divisie|hnl)\b', 3),
        (r'\b2\s*(?:lig|liga|division|divisie|bundesliga|hnl)\b', 2),
        (r'\b1\s*(?:lig|liga|division|divisie|hnl)\b', 1),
    ]
    for pattern, tier in numeric_patterns:
        if re.search(pattern, league_core):
            return tier
    return None

def _lig_components_match(found_key: str, selected_key: str) -> bool:
    found_country, found_league = _split_known_lig_key(found_key)
    sel_country, sel_league = _split_known_lig_key(selected_key)

    # Secili ulke bilinmiyorsa → tanimsiz kaynak, eslesme yok
    if not sel_country:
        return False
    # Bulunan ulke bilinmiyorsa → LEAGUE_LIST'te tanimsiz ulke, ESLESMEZ
    # (Fas "Premier Lig" ≠ Ingiltere "Premier Lig" — ulke sarti zorunlu)
    if not found_country:
        return False
    # Ülke farklıysa kesinlikle eşleşmez
    if found_country != sel_country:
        return False
    # Ayni ulke — lig karsilastir
    if not found_league or not sel_league:
        return found_country == sel_country
    found_tier = _league_tier(found_country, found_league)
    sel_tier = _league_tier(sel_country, sel_league)
    # Esas kural: ayarlarda seçilen lig, aynı ülkedeki aynı seviye ile eşleşir.
    if found_tier is not None and sel_tier is not None:
        return found_tier == sel_tier
    # Seviye çözülemediyse sadece birebir normalize isim eşleşmesi kabul edilir.
    return found_league == sel_league

def lig_filtreli(iddaa_link: str, sel_leagues: set[str] | None) -> bool:
    """URL'den ülke+lig key'i çıkar, seçili liglerdeyse True döndür."""
    if sel_leagues is None:
        return True
    if not sel_leagues:
        return False
    try:
        path = iddaa_link.split("mackolik.com/", 1)[1].strip("/")
    except IndexError:
        return False
    parts = [unquote(x) for x in path.split("/") if x]
    # parts[0]=futbol, parts[1]=ulke, parts[2]=lig-slug
    if len(parts) >= 3:
        url_country = _fold_lig(parts[1])
        url_lig     = _fold_lig(parts[2])
        key = url_country + " " + url_lig
    elif len(parts) >= 2:
        url_country = _fold_lig(parts[1])
        url_lig     = ""
        key = url_country
    else:
        return False
    return any(_lig_components_match(key, k) for k in sel_leagues)

def lig_filtreli_key(lig_key: str, sel_leagues: set[str] | None) -> bool:
    """Macin lig_key'ini seçili liglerle karşılaştırır."""
    if sel_leagues is None:
        return True
    if not sel_leagues:
        return False
    if not lig_key or not lig_key.strip():
        return False
    for k in sel_leagues:
        if _lig_components_match(lig_key, k):
            return True
    return False

def parse_header(driver) -> dict:
    data = {"mac_tarihi": "", "lig": ""}
    el = _safe_find(driver, By.CSS_SELECTOR, f"span[class*='{DATE_INFO_CLASS}']")
    if el:
        data["mac_tarihi"] = _norm(el.text)
    comp = _safe_find(driver, By.CSS_SELECTOR, f"a[class*='{COMPETITION_CLASS}']")
    if comp:
        # Elementin görünen metnini al; çok satırlıysa ilk satır yeter
        txt = _norm(comp.text.splitlines()[0]) if comp.text else ""
        if not txt:
            txt = _extract_league(_norm(comp.get_attribute("href")))
        data["lig"] = txt
    return data

def _split_option(opt_el) -> tuple[str, str]:
    lines = [_norm(x) for x in opt_el.text.splitlines() if _norm(x)]
    if not lines:
        return "", ""
    for i in range(len(lines)-1, -1, -1):
        v = lines[i].replace(",", ".")
        if v.replace(".", "", 1).isdigit():
            return " ".join(lines[:i]).strip(), v
    return _norm(opt_el.text), ""

def parse_all_markets(driver, market_keys: set[str] | None = None) -> dict:
    result = {
        "ms_kodu":"", "ms1":"", "ms0":"", "ms2":"",
        "cs_1x":"", "cs_12":"", "cs_x2":"",
        "iy1":"", "iy0":"", "iy2":"",
        "au_0_5_alt":"", "au_0_5_ust":"",
        "au_1_5_alt":"", "au_1_5_ust":"",
        "au_2_5_alt":"", "au_2_5_ust":"",
        "au_3_5_alt":"", "au_3_5_ust":"",
        "au_4_5_alt":"", "au_4_5_ust":"",
        "kg_var":"", "kg_yok":"",
        "hnd_1":"", "hnd_x":"", "hnd_2":"",
        "hnd2_1":"", "hnd2_x":"", "hnd2_2":"",
        "iy_au_05_alt":"", "iy_au_05_ust":"",
        "iy_au_15_alt":"", "iy_au_15_ust":"",
        "iy_ms_1_1":"", "iy_ms_1_x":"", "iy_ms_1_2":"",
        "iy_ms_x_1":"", "iy_ms_x_x":"", "iy_ms_x_2":"",
        "iy_ms_2_1":"", "iy_ms_2_x":"", "iy_ms_2_2":"",
        "tg_0_1":"", "tg_2_3":"", "tg_4_5":"", "tg_6p":"",
        "t1_1_5_ust":"", "t1_2_5_ust":"",
        "t2_1_5_ust":"", "t2_2_5_ust":"",
    }

    items = _safe_find_all(driver, By.CSS_SELECTOR, f"li.{MARKET_ITEM_CLASS}")
    if not items:
        items = _safe_find_all(driver, By.CSS_SELECTOR, f".{MARKET_ITEM_CLASS}")

    # Maç kodu: iddaa-code elementinden al
    try:
        code_els = driver.find_elements(By.CSS_SELECTOR, f".{IDDAA_CODE_CLASS}")
        for ce in code_els:
            ct = _norm(ce.text).strip()
            if ct and ct.isdigit():
                result["ms_kodu"] = ct
                break
    except Exception:
        pass

    def wants(*keys: str) -> bool:
        return market_keys is None or any(k in market_keys for k in keys)

    for item in items:
        try:
            raw = _norm(item.text)
            lines = raw.splitlines()
            first = _fold_text(lines[0] if lines else raw).replace(",", ".")
        except Exception:
            continue

        def opts_map():
            mapping = {}
            for opt in _safe_find_all(item, By.CSS_SELECTOR, f".{OPTION_CLASS}"):
                label, odd = _split_option(opt)
                if label:
                    mapping[_fold_text(label).replace(",", ".")] = odd
            return mapping

        try:
            if first.startswith("mac sonucu"):
                if not result["ms_kodu"]:
                    try:
                        code_el = item.find_element(By.CSS_SELECTOR, f".{IDDAA_CODE_CLASS}")
                        ct = _norm(code_el.text).strip()
                        if ct and ct.isdigit():
                            result["ms_kodu"] = ct
                    except Exception:
                        pass
                if not result["ms_kodu"]:
                    m = re.search(r"(\d+)", first)
                    result["ms_kodu"] = m.group(1) if m else ""
                if wants("ms1", "ms0", "ms2"):
                    mp = opts_map()
                    result["ms1"] = mp.get("1", "")
                    result["ms0"] = mp.get("x", "")
                    result["ms2"] = mp.get("2", "")
            elif ("cifte sans" in first or "çifte şans" in first) and wants("cs_1x", "cs_12", "cs_x2"):
                mp = opts_map()
                for k, v in mp.items():
                    kk = k.replace(" ", "")
                    if "1x" in kk or "1-x" in kk:
                        result["cs_1x"] = v
                    elif "12" in kk or "1-2" in kk:
                        result["cs_12"] = v
                    elif "x2" in kk or "x-2" in kk or "2x" in kk:
                        result["cs_x2"] = v
            elif ("ilk yari sonucu" in first or "1. yari sonucu" in first or "ilk yarı sonucu" in first or "1. yarı sonucu" in first) and "/" not in first and wants("iy1", "iy0", "iy2"):
                mp = opts_map()
                result["iy1"] = mp.get("1", "")
                result["iy0"] = mp.get("x", "")
                result["iy2"] = mp.get("2", "")
            elif (
                "alt/ust" in first and "ilk yari" not in first and "1. yari" not in first and "takim" not in first and "ev sahibi" not in first and "deplasman" not in first and "konuk" not in first
                and wants(
                    "au_0_5_alt", "au_0_5_ust",
                    "au_1_5_alt", "au_1_5_ust",
                    "au_2_5_alt", "au_2_5_ust",
                    "au_3_5_alt", "au_3_5_ust",
                    "au_4_5_alt", "au_4_5_ust",
                )
            ):
                for t, key in [("0.5", "0_5"), ("1.5", "1_5"), ("2.5", "2_5"), ("3.5", "3_5"), ("4.5", "4_5")]:
                    if t in first:
                        mp = opts_map()
                        for k, v in mp.items():
                            if "alt" in k:
                                result[f"au_{key}_alt"] = v
                            elif "ust" in k:
                                result[f"au_{key}_ust"] = v
                        break
            elif ("karsilikli" in first or ("iki takim" in first and "gol" in first)) and wants("kg_var", "kg_yok"):
                mp = opts_map()
                for k, v in mp.items():
                    if "var" in k or "evet" in k:
                        result["kg_var"] = v
                    elif "yok" in k or "hayir" in k:
                        result["kg_yok"] = v
            elif ("handikap" in first or "hnd" in first) and ("-1" in first or "(0:1)" in first or "0:1" in first) and wants("hnd_1", "hnd_x", "hnd_2"):
                mp = opts_map()
                for k, v in mp.items():
                    if k == "1": result["hnd_1"] = result["hnd_1"] or v
                    elif k == "x": result["hnd_x"] = result["hnd_x"] or v
                    elif k == "2" or "+" in k: result["hnd_2"] = result["hnd_2"] or v
            elif ("handikap" in first or "hnd" in first) and ("+1" in first or "(1:0)" in first or "1:0" in first) and wants("hnd2_1", "hnd2_x", "hnd2_2"):
                mp = opts_map()
                for k, v in mp.items():
                    if k == "1": result["hnd2_1"] = result["hnd2_1"] or v
                    elif k == "x": result["hnd2_x"] = result["hnd2_x"] or v
                    elif k == "2" or "-" in k: result["hnd2_2"] = result["hnd2_2"] or v
            elif (
                ("ilk yari" in first or "1. yari" in first) and ("alt" in first or "ust" in first) and "sonucu" not in first
                and wants("iy_au_05_alt", "iy_au_05_ust", "iy_au_15_alt", "iy_au_15_ust")
            ):
                for t, key in [("0.5", "05"), ("1.5", "15")]:
                    if t in first:
                        mp = opts_map()
                        for k, v in mp.items():
                            if "alt" in k:
                                result[f"iy_au_{key}_alt"] = v
                            elif "ust" in k:
                                result[f"iy_au_{key}_ust"] = v
                        break
            elif ("ilk yari" in first or "1. yari" in first) and "mac sonucu" in first and wants(
                "iy_ms_1_1", "iy_ms_1_x", "iy_ms_1_2",
                "iy_ms_x_1", "iy_ms_x_x", "iy_ms_x_2",
                "iy_ms_2_1", "iy_ms_2_x", "iy_ms_2_2",
            ):
                mp = opts_map()
                combos = {
                    "1/1": "iy_ms_1_1", "1/x": "iy_ms_1_x", "1/2": "iy_ms_1_2",
                    "x/1": "iy_ms_x_1", "x/x": "iy_ms_x_x", "x/2": "iy_ms_x_2",
                    "2/1": "iy_ms_2_1", "2/x": "iy_ms_2_x", "2/2": "iy_ms_2_2",
                }
                for k, v in mp.items():
                    norm_k = k.replace(" ", "")
                    if norm_k in combos:
                        result[combos[norm_k]] = v
            elif (
                "toplam gol" in first and ("0-1" in first or "2-3" in first or raw.count("\n") >= 3)
                and wants("tg_0_1", "tg_2_3", "tg_4_5", "tg_6p")
            ):
                mp = opts_map()
                for k, v in mp.items():
                    kk = k.replace(" ", "")
                    if "0-1" in kk:
                        result["tg_0_1"] = v
                    elif "2-3" in kk:
                        result["tg_2_3"] = v
                    elif "4-5" in kk:
                        result["tg_4_5"] = v
                    elif "6+" in kk or "6ve" in kk:
                        result["tg_6p"] = v
            elif ("ev sahibi" in first or "takim 1" in first) and "alt/ust" in first and wants("t1_1_5_ust", "t1_2_5_ust"):
                mp = opts_map()
                ust_val = mp.get("ust", "")
                if "1.5" in first:
                    result["t1_1_5_ust"] = ust_val
                elif "2.5" in first:
                    result["t1_2_5_ust"] = ust_val
            elif (
                ("konuk" in first or "deplasman" in first or "takim 2" in first)
                and "alt/ust" in first and wants("t2_1_5_ust", "t2_2_5_ust")
            ):
                mp = opts_map()
                ust_val = mp.get("ust", "")
                if "1.5" in first:
                    result["t2_1_5_ust"] = ust_val
                elif "2.5" in first:
                    result["t2_2_5_ust"] = ust_val
        except Exception:
            continue

    return result


def _parse_header_bs4(html: str) -> dict:
    data = {'mac_tarihi': '', 'lig': ''}
    soup = BeautifulSoup(html, 'html.parser')
    el = soup.select_one("span[class*='p0c-soccer-match-details-header__info-date']")
    if el:
        data['mac_tarihi'] = _norm(el.get_text())
    comp = soup.select_one("a[class*='p0c-soccer-match-details-header__competition-link']")
    if comp:
        txt = _norm(comp.get_text().splitlines()[0]) if comp.get_text() else ''
        if not txt:
            txt = _extract_league(comp.get('href', ''))
        data['lig'] = txt
    return data

def _bs4_opt_map(item_el):
    mp = {}
    for opt in item_el.select('.widget-iddaa-markets__option'):
        txt = opt.get_text(separator='|', strip=True)
        lines = [x.strip() for x in txt.split('|') if x.strip()]
        if not lines:
            continue
        for i in range(len(lines)-1, -1, -1):
            v = lines[i].replace(',', '.')
            if v.replace('.', '', 1).isdigit():
                label = _fold_text(' '.join(lines[:i]).strip())
                mp[label] = v
                break
    return mp

def _parse_markets_bs4(html: str, market_keys: set | None = None) -> dict:
    result = {
        'ms_kodu':'', 'ms1':'', 'ms0':'', 'ms2':'',
        'cs_1x':'', 'cs_12':'', 'cs_x2':'',
        'iy1':'', 'iy0':'', 'iy2':'',
        'au_0_5_alt':'', 'au_0_5_ust':'',
        'au_1_5_alt':'', 'au_1_5_ust':'',
        'au_2_5_alt':'', 'au_2_5_ust':'',
        'au_3_5_alt':'', 'au_3_5_ust':'',
        'au_4_5_alt':'', 'au_4_5_ust':'',
        'kg_var':'', 'kg_yok':'',
        'hnd_1':'', 'hnd_x':'', 'hnd_2':'',
        'hnd2_1':'', 'hnd2_x':'', 'hnd2_2':'',
        'iy_au_05_alt':'', 'iy_au_05_ust':'',
        'iy_au_15_alt':'', 'iy_au_15_ust':'',
        'iy_ms_1_1':'', 'iy_ms_1_x':'', 'iy_ms_1_2':'',
        'iy_ms_x_1':'', 'iy_ms_x_x':'', 'iy_ms_x_2':'',
        'iy_ms_2_1':'', 'iy_ms_2_x':'', 'iy_ms_2_2':'',
        'tg_0_1':'', 'tg_2_3':'', 'tg_4_5':'', 'tg_6p':'',
        't1_1_5_ust':'', 't1_2_5_ust':'',
        't2_1_5_ust':'', 't2_2_5_ust':'',
    }
    soup = BeautifulSoup(html, 'html.parser')
    # Maç kodu: iddaa-code elementlerinden al
    code_els = soup.select('.widget-iddaa-markets__iddaa-code')
    for ce in code_els:
        code_txt = _norm(ce.get_text()).strip()
        if code_txt and code_txt.isdigit():
            result['ms_kodu'] = code_txt
            break
    def wants(*keys):
        return market_keys is None or any(k in market_keys for k in keys)
    items = soup.select('li.widget-iddaa-markets__market-item')
    if not items:
        items = soup.select('.widget-iddaa-markets__market-item')
    for item in items:
        try:
            raw = item.get_text(separator='|', strip=True)
            parts = raw.split('|')
            first = _fold_text(parts[0] if parts else raw).replace(',', '.')
        except Exception:
            continue
        try:
            if first.startswith('mac sonucu'):
                # Maç kodu: önce iddaa-code elementinden dene, sonra header text'ten
                if not result['ms_kodu']:
                    code_el = item.select_one('.widget-iddaa-markets__iddaa-code')
                    if code_el:
                        ct = _norm(code_el.get_text()).strip()
                        if ct and ct.isdigit():
                            result['ms_kodu'] = ct
                if not result['ms_kodu']:
                    m = re.search(r'(\d+)', first)
                    result['ms_kodu'] = m.group(1) if m else ''
                if wants('ms1','ms0','ms2'):
                    mp = _bs4_opt_map(item)
                    result['ms1']=mp.get('1',''); result['ms0']=mp.get('x',''); result['ms2']=mp.get('2','')
            elif ('cifte sans' in first or 'çifte şans' in first) and wants('cs_1x','cs_12','cs_x2'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk=k.replace(' ','')
                    if '1x' in kk or '1-x' in kk: result['cs_1x']=v
                    elif '12' in kk or '1-2' in kk: result['cs_12']=v
                    elif 'x2' in kk or 'x-2' in kk or '2x' in kk: result['cs_x2']=v
            elif (('ilk yari' in first or '1. yari' in first or 'ilk yarı' in first or '1. yarı' in first) and 'sonucu' in first and 'mac' not in first and '/' not in first) and wants('iy1', 'iy0', 'iy2'):
                mp = _bs4_opt_map(item)
                result['iy1']=mp.get('1',''); result['iy0']=mp.get('x',''); result['iy2']=mp.get('2','')
            elif 'alt/ust' in first and 'ilk yari' not in first and '1. yari' not in first and 'takim' not in first and 'ev sahibi' not in first and 'deplasman' not in first and 'konuk' not in first:
                for th,key in [('0.5','0_5'),('1.5','1_5'),('2.5','2_5'),('3.5','3_5'),('4.5','4_5')]:
                    if th in first and wants(f'au_{key}_alt',f'au_{key}_ust'):
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if 'alt' in k: result[f'au_{key}_alt']=v
                            elif 'ust' in k: result[f'au_{key}_ust']=v
                        break
            elif ('karsilikli' in first or ('iki takim' in first and 'gol' in first)) and wants('kg_var','kg_yok'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    if 'var' in k or 'evet' in k: result['kg_var']=v
                    elif 'yok' in k or 'hayir' in k: result['kg_yok']=v
            elif ('handikap' in first or 'hnd' in first) and ('-1' in first or '0:1' in first) and wants('hnd_1','hnd_x','hnd_2'):
                mp = _bs4_opt_map(item)
                result['hnd_1']=mp.get('1',''); result['hnd_x']=mp.get('x',''); result['hnd_2']=mp.get('2','')
            elif ('handikap' in first or 'hnd' in first) and ('+1' in first or '1:0' in first) and wants('hnd2_1','hnd2_x','hnd2_2'):
                mp = _bs4_opt_map(item)
                result['hnd2_1']=mp.get('1',''); result['hnd2_x']=mp.get('x',''); result['hnd2_2']=mp.get('2','')
            elif ('ilk yari' in first or '1. yari' in first) and ('alt' in first or 'ust' in first) and 'sonucu' not in first:
                for th,key in [('0.5','05'),('1.5','15')]:
                    if th in first and wants(f'iy_au_{key}_alt',f'iy_au_{key}_ust'):
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if 'alt' in k: result[f'iy_au_{key}_alt']=v
                            elif 'ust' in k: result[f'iy_au_{key}_ust']=v
                        break
            elif ('ilk yari' in first or '1. yari' in first) and 'mac sonucu' in first:
                mp = _bs4_opt_map(item)
                combos = {'1/1':'iy_ms_1_1','1/x':'iy_ms_1_x','1/2':'iy_ms_1_2','x/1':'iy_ms_x_1','x/x':'iy_ms_x_x','x/2':'iy_ms_x_2','2/1':'iy_ms_2_1','2/x':'iy_ms_2_x','2/2':'iy_ms_2_2'}
                for k,v in mp.items():
                    nk=k.replace(' ','')
                    if nk in combos: result[combos[nk]]=v
            elif 'toplam gol' in first and ('0-1' in first or '2-3' in first or raw.count('|')>=3):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk=k.replace(' ','')
                    if '0-1' in kk: result['tg_0_1']=v
                    elif '2-3' in kk: result['tg_2_3']=v
                    elif '4-5' in kk: result['tg_4_5']=v
                    elif '6+' in kk or '6ve' in kk: result['tg_6p']=v
            elif ('ev sahibi' in first or 'takim 1' in first) and 'alt/ust' in first and wants('t1_1_5_ust','t1_2_5_ust'):
                mp = _bs4_opt_map(item)
                ust_val = mp.get('ust', '')
                alt_val = mp.get('alt', '')
                if '1.5' in first:
                    result['t1_1_5_ust'] = ust_val
                elif '2.5' in first:
                    result['t1_2_5_ust'] = ust_val
            elif ('konuk' in first or 'deplasman' in first or 'takim 2' in first) and 'alt/ust' in first and wants('t2_1_5_ust','t2_2_5_ust'):
                mp = _bs4_opt_map(item)
                ust_val = mp.get('ust', '')
                alt_val = mp.get('alt', '')
                if '1.5' in first:
                    result['t2_1_5_ust'] = ust_val
                elif '2.5' in first:
                    result['t2_2_5_ust'] = ust_val
        except Exception:
            continue
    return result

def _row_is_valid(r: dict) -> bool:
    """Maç satırı geçerli mi? Skor ve oran verisi zorunlu."""
    # IY skoru ve MS skoru mutlaka olmalı
    iy = (r.get('ilk_yari_skor') or '').strip()
    ms = (r.get('mac_skoru') or '').strip()
    if not iy or not ms:
        return False
    # En az temel oranlardan biri olmalı
    has_odds = any(r.get(k) for k in ('ms1', 'ms0', 'ms2'))
    return has_odds

# Adaptive throttle: 502/500 alınca tüm thread'ler yavaşlar, düzelince hızlanır
_throttle_delay = 0.02          # başlangıç: 20ms (agresif hız)
_throttle_lock = threading.Lock()

def _throttle_hit():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = min(_throttle_delay * 1.3, 0.5)

def _throttle_ok():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = max(_throttle_delay * 0.5, 0.02)

def scrape_match_fast(summary: dict, match_date=None, market_keys=None,
                      max_retries: int = 3) -> dict:
    row = dict(summary)
    if match_date:
        row['mac_tarihi'] = match_date.strftime('%d.%m.%Y')
    label = summary.get('ev_sahibi', '')

    urls_to_try = [summary['iddaa_link']]
    overview = summary.get('overview_link', '')
    if overview and overview != summary['iddaa_link']:
        urls_to_try.append(overview)

    for url in urls_to_try:
        for attempt in range(max_retries):
            try:
                time.sleep(_throttle_delay)
                resp = _SESSION.get(url, timeout=12, allow_redirects=True)
                if resp.status_code == 404:
                    break
                if resp.status_code in (500, 502, 503, 429):
                    _throttle_hit()
                    if attempt < max_retries - 1:
                        wait = (attempt + 1) * 2
                        time.sleep(wait)
                        continue
                    else:
                        break
                resp.raise_for_status()
                _throttle_ok()
                html = resp.text
                hdr = _parse_header_bs4(html)
                if not row.get('mac_tarihi') and hdr.get('mac_tarihi'):
                    row['mac_tarihi'] = hdr['mac_tarihi']
                if hdr.get('lig'):
                    row['lig'] = hdr['lig']
                api_kodu = row.get('ms_kodu', '')
                row.update(_parse_markets_bs4(html, market_keys=market_keys))
                if not row.get('ms_kodu') and api_kodu:
                    row['ms_kodu'] = api_kodu
                return row
            except requests.exceptions.HTTPError as e:
                status = getattr(getattr(e, 'response', None), 'status_code', 0)
                if status == 404:
                    break
                if status in (500, 502, 503, 429):
                    _throttle_hit()
                    wait = (attempt + 1) * 2
                else:
                    wait = (attempt + 1) * 1
                if attempt < max_retries - 1:
                    print(f'  [scrape_match_fast] {label} deneme {attempt+1} hata: {e} — {wait:.1f}s bekle')
                    time.sleep(wait)
                else:
                    print(f'  [scrape_match_fast] {label} hata (tum denemeler): {e}')
            except Exception as e:
                if attempt < max_retries - 1:
                    wait = (attempt + 1) * 1
                    print(f'  [scrape_match_fast] {label} deneme {attempt+1} hata: {e} — {wait:.1f}s bekle')
                    time.sleep(wait)
                else:
                    print(f'  [scrape_match_fast] {label} hata (tum denemeler): {e}')
    return row

def scrape_match(driver, summary: dict, stop_flag: threading.Event | None = None,
                 match_date: dt.date | None = None,
                 market_keys: set[str] | None = None) -> dict:
    row = dict(summary)
    # Tarih seçili günden gelir, overview sayfasına gitmeye gerek yok
    if match_date:
        row["mac_tarihi"] = match_date.strftime("%d.%m.%Y")
    if stop_flag and stop_flag.is_set():
        return row
    try:
        driver.get(summary["iddaa_link"])
        # Sayfa yüklenince market'ler hemen gelir; max 4s bekle
        WebDriverWait(driver, 4).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR, f"li.{MARKET_ITEM_CLASS}")) > 0
        )
    except TimeoutException:
        pass
    if stop_flag and stop_flag.is_set():
        return row
    # Lig bilgisini iddaa sayfasından al
    hdr = parse_header(driver)
    if not row.get("mac_tarihi") and hdr.get("mac_tarihi"):
        row["mac_tarihi"] = hdr["mac_tarihi"]
    row["lig"] = hdr.get("lig", "")
    api_kodu = row.get("ms_kodu", "")
    row.update(parse_all_markets(driver, market_keys=market_keys))
    if not row.get("ms_kodu") and api_kodu:
        row["ms_kodu"] = api_kodu
    return row

_BASE_KEYS    = ["ev_sahibi","konuk_ekip","mac_tarihi","mac_saati","lig",
                  "ms_kodu","ilk_yari_skor","mac_skoru"]
_BASE_HEADERS = ["Ev Sahibi","Konuk Ekip","Tarih","Saat","Lig",
                  "MS Kodu","IY Skor","MS Skor"]

def export_excel(rows: list[dict], path: Path,
                 market_keys: set[str] | None = None) -> None:
    mkt_keys    = KEYS[8:]
    mkt_headers = HEADERS[8:]
    if market_keys is not None:
        eff_keys  = _BASE_KEYS    + [k for k in mkt_keys    if k in market_keys]
        eff_hdrs  = _BASE_HEADERS + [h for k, h in zip(mkt_keys, mkt_headers) if k in market_keys]
    else:
        eff_keys, eff_hdrs = KEYS, HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = "MS Oranlari"
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(eff_hdrs)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(k, "") for k in eff_keys])
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
    wb.save(path)

# ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â
#  GUI
# ÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚ÂÃƒÂ¢Ã¢â‚¬Â¢Ã‚Â

class _Tooltip:
    """Buton ÃƒÆ’Ã‚Â¼zerine gelince aÃƒÆ’Ã‚Â§Ãƒâ€Ã‚Â±klama baloncuÃƒâ€Ã…Â¸u."""
    def __init__(self, widget, text: str):
        self._tip = None
        widget.bind("<Enter>", lambda e: self._show(widget, text))
        widget.bind("<Leave>", lambda e: self._hide())

    def _show(self, widget, text):
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + widget.winfo_height() + 2
        self._tip = tk.Toplevel(widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        tk.Label(self._tip, text=text, bg="#F7FBF5", fg="#224131",
                 font=("Segoe UI", 7), relief="solid", bd=1,
                 highlightbackground="#18A558", highlightthickness=1,
                 padx=6, pady=3).pack()

    def _hide(self):
        if self._tip:
            self._tip.destroy()
            self._tip = None
CLR_BG       = "#EEF4EC"
CLR_CARD     = "#FBFDF8"
CLR_HEADER   = "#FFFFFF"
CLR_ACCENT   = "#12B15B"
CLR_ACCENT_D = "#0E944A"
CLR_GOLD     = "#1F6A43"
CLR_BTN_G    = "#10B55D"
CLR_BTN_Y    = "#F4C542"
CLR_BTN_O    = "#F57C00"
CLR_BTN_R    = "#E53935"
CLR_BTN_GRY  = "#58717F"
CLR_BTN_TXT  = "#FFFFFF"
CLR_BORDER   = "#C8D8C7"
CLR_TXT_PRI  = "#25392D"
CLR_TXT_SEC  = "#5F7466"
CLR_GRID_H   = "#0F7A43"
CLR_HOVER    = "#E0EADF"
CLR_ROW_ODD  = "#F8FBF5"
CLR_ROW_EVEN = "#EEF5EC"
CLR_SELECTED = "#CDEFD9"

TREE_COLS = [
    ("Ma\u00e7",        220),
    ("Tarih",       90),
    ("Saat",        55),
    ("Lig",        150),
    ("MsKodu",      60),
    ("IY",          55),
    ("MS",          55),
    ("MS1",         45),
    ("MS0",         45),
    ("MS2",         45),
    ("CS1X",        48),
    ("CS12",        48),
    ("CSX2",        48),
    ("IY1",         45),
    ("IY0",         45),
    ("IY2",         45),
    ("AU05A",       52),
    ("AU05U",       52),
    ("AU15A",       52),
    ("AU15U",       52),
    ("AU25A",       52),
    ("AU25U",       52),
    ("AU35A",       52),
    ("AU35U",       52),
    ("AU45A",       52),
    ("AU45U",       52),
    ("KGVar",       50),
    ("KGYok",       50),
    ("HND1",        45),
    ("HNDX",        45),
    ("HND2",        45),
    ("H2-1",        45),
    ("H2-X",        45),
    ("H2-2",        45),
    ("IYAU05A",     55),
    ("IYAU05U",     55),
    ("IYAU15A",     55),
    ("IYAU15U",     55),
    ("IY/MS1/1",    55),
    ("IY/MS1/X",    55),
    ("IY/MS1/2",    55),
    ("IY/MSX/1",    55),
    ("IY/MSX/X",    55),
    ("IY/MSX/2",    55),
    ("IY/MS2/1",    55),
    ("IY/MS2/X",    55),
    ("IY/MS2/2",    55),
    ("TG0-1",       50),
    ("TG2-3",       50),
    ("TG4-5",       50),
    ("TG6+",        50),
    ("T1-1.5U",     52),
    ("T1-2.5U",     52),
    ("T2-1.5U",     52),
    ("T2-2.5U",     52),
]

KEYS = [
    "ev_sahibi", "konuk_ekip", "mac_tarihi", "mac_saati", "lig",
    "ms_kodu", "ilk_yari_skor", "mac_skoru",
    "ms1", "ms0", "ms2",
    "cs_1x", "cs_12", "cs_x2",
    "iy1", "iy0", "iy2",
    "au_0_5_alt", "au_0_5_ust",
    "au_1_5_alt", "au_1_5_ust",
    "au_2_5_alt", "au_2_5_ust",
    "au_3_5_alt", "au_3_5_ust",
    "au_4_5_alt", "au_4_5_ust",
    "kg_var", "kg_yok",
    "hnd_1", "hnd_x", "hnd_2",
    "hnd2_1", "hnd2_x", "hnd2_2",
    "iy_au_05_alt", "iy_au_05_ust",
    "iy_au_15_alt", "iy_au_15_ust",
    "iy_ms_1_1", "iy_ms_1_x", "iy_ms_1_2",
    "iy_ms_x_1", "iy_ms_x_x", "iy_ms_x_2",
    "iy_ms_2_1", "iy_ms_2_x", "iy_ms_2_2",
    "tg_0_1", "tg_2_3", "tg_4_5", "tg_6p",
    "t1_1_5_ust", "t1_2_5_ust",
    "t2_1_5_ust", "t2_2_5_ust",
]

HEADERS = [
    "Ev Sahibi", "Konuk Ekip", "Tarih", "Saat", "Lig",
    "MS Kodu", "IY Skor", "MS Skor",
    "MS1", "MS0", "MS2",
    "CS 1X", "CS 12", "CS X2",
    "IY1", "IY0", "IY2",
    "AU 0.5 Alt", "AU 0.5 Ust",
    "AU 1.5 Alt", "AU 1.5 Ust",
    "AU 2.5 Alt", "AU 2.5 Ust",
    "AU 3.5 Alt", "AU 3.5 Ust",
    "AU 4.5 Alt", "AU 4.5 Ust",
    "KG Var", "KG Yok",
    "HND1", "HNDX", "HND2",
    "HND2-1", "HND2-X", "HND2-2",
    "IY AU 0.5 Alt", "IY AU 0.5 Ust",
    "IY AU 1.5 Alt", "IY AU 1.5 Ust",
    "IY/MS 1/1", "IY/MS 1/X", "IY/MS 1/2",
    "IY/MS X/1", "IY/MS X/X", "IY/MS X/2",
    "IY/MS 2/1", "IY/MS 2/X", "IY/MS 2/2",
    "TG 0-1", "TG 2-3", "TG 4-5", "TG 6+",
    "T1 1.5 Ust", "T1 2.5 Ust",
    "T2 1.5 Ust", "T2 2.5 Ust",
]

# ── Lig listesi (country_display, [league_displays]) ─────────────────────────
LEAGUE_LIST: list[tuple[str, list[str]]] = [
    ("T\u00dcRK\u0130YE",      ["S\u00fcper Lig", "1. Lig"]),
    ("\u0130NG\u0130LTERE",     ["Premier Lig", "Championship", "1. Lig", "2. Lig"]),
    ("\u0130SPANYA",            ["LaLiga", "LaLiga 2"]),
    ("\u0130TALYA",             ["Serie A", "Serie B"]),
    ("ALMANYA",                 ["Bundesliga", "2. Bundesliga"]),
    ("FRANSA",                  ["Ligue 1", "Ligue 2", "Ulusal Lig 1"]),
    ("HOLLANDA",                ["Eredivisie", "Eerste Divisie"]),
    ("PORTEK\u0130Z",           ["Premier Lig", "2. Lig"]),
    ("BEL\u00c7\u0130KA",      ["Pro Lig", "Challenger Pro Lig"]),
    ("AVUSTURYA",               ["Bundesliga", "1. Lig"]),
    ("\u00c7EK CUMHUR\u0130YET\u0130", ["Czech Liga"]),
    ("DAN\u0130MARKA",          ["Superliga", "1. Lig"]),
    ("F\u0130NLAND\u0130YA",    ["Veikkausliiga", "Ykk\u00f6sliiga"]),
    ("HIRVATISTAN",             ["1. HNL", "2. HNL"]),
    ("\u0130SKO\u00c7YA",       ["Premiership", "Championship", "3. Lig", "4. Lig"]),
    ("\u0130SVE\u00c7",         ["Allsvenskan", "Superettan", "3. Lig"]),
    ("\u0130SV\u0130\u00c7RE",  ["S\u00fcper Lig", "Challenge Lig"]),
    ("MACAR\u0130STAN",         ["NB I", "2. Lig"]),
    ("NORVE\u00c7",             ["Eliteserien", "1. Lig"]),
    ("POLONYA",                 ["Ekstraklasa", "1. Lig", "2. Lig"]),
    ("ROMANYA",                 ["1. Lig", "2. Lig"]),
    ("RUSYA",                   ["Premier Lig", "FNL"]),
    ("S\u0130RB\u0130STAN",     ["S\u00fcper Lig", "1. Lig"]),
    ("YUNAN\u0130STAN",         ["S\u00fcper Lig", "2. Lig"]),
    ("ABD",                     ["MLS"]),
    ("BREZ\u0130LYA",           ["Serie A"]),
    ("JAPONYA",                 ["J1 Ligi"]),
    ("G\u00dcNEY KORE",         ["K Lig 1"]),
    ("\u00c7\u0130N",           ["S\u00fcper Lig"]),
    ("AVUSTRALYA",              ["A-Lig"]),
]

# Tum tanimli lig key'leri — "Tumunu Sec" durumunda bu set filtre olarak kullanilir
# (Cezayir, Fas, Misir vb. tanimsiz ligler otomatik atlanir)
_ALL_DEFINED_LEAGUE_KEYS: set[str] | None = None  # lazy init (lig_key fonksiyonu lazim)

def _get_all_league_keys() -> set[str]:
    global _ALL_DEFINED_LEAGUE_KEYS
    if _ALL_DEFINED_LEAGUE_KEYS is None:
        _ALL_DEFINED_LEAGUE_KEYS = {lig_key(c, l) for c, ls in LEAGUE_LIST for l in ls}
    return _ALL_DEFINED_LEAGUE_KEYS

# ── Market grupları (group_label, [(key, header_label), ...]) ─────────────────
MARKET_GROUPS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Ma\u00e7 Sonucu", [
        ("ms1","MS1"), ("ms0","MS0"), ("ms2","MS2"),
    ]),
    ("\u00c7ifte \u015eans", [
        ("cs_1x","CS 1X"), ("cs_12","CS 12"), ("cs_x2","CS X2"),
    ]),
    ("\u0130Y Sonucu", [
        ("iy1","IY1"), ("iy0","IY0"), ("iy2","IY2"),
    ]),
    ("Alt/\u00dcst 0.5", [
        ("au_0_5_alt","AU 0.5 Alt"), ("au_0_5_ust","AU 0.5 \u00dcst"),
    ]),
    ("Alt/\u00dcst 1.5", [
        ("au_1_5_alt","AU 1.5 Alt"), ("au_1_5_ust","AU 1.5 \u00dcst"),
    ]),
    ("Alt/\u00dcst 2.5", [
        ("au_2_5_alt","AU 2.5 Alt"), ("au_2_5_ust","AU 2.5 \u00dcst"),
    ]),
    ("Alt/\u00dcst 3.5", [
        ("au_3_5_alt","AU 3.5 Alt"), ("au_3_5_ust","AU 3.5 \u00dcst"),
    ]),
    ("Alt/\u00dcst 4.5", [
        ("au_4_5_alt","AU 4.5 Alt"), ("au_4_5_ust","AU 4.5 \u00dcst"),
    ]),
    ("Kar\u015f\u0131l\u0131kl\u0131 Gol", [
        ("kg_var","KG Var"), ("kg_yok","KG Yok"),
    ]),
    ("Handikap (-1)", [
        ("hnd_1","HND1"), ("hnd_x","HNDX"), ("hnd_2","HND2"),
    ]),
    ("Handikap (+1)", [
        ("hnd2_1","HND2-1"), ("hnd2_x","HND2-X"), ("hnd2_2","HND2-2"),
    ]),
    ("\u0130Y Alt/\u00dcst 0.5", [
        ("iy_au_05_alt","IY AU 0.5 Alt"), ("iy_au_05_ust","IY AU 0.5 \u00dcst"),
    ]),
    ("\u0130Y Alt/\u00dcst 1.5", [
        ("iy_au_15_alt","IY AU 1.5 Alt"), ("iy_au_15_ust","IY AU 1.5 \u00dcst"),
    ]),
    ("\u0130Y/MS Kombine", [
        ("iy_ms_1_1","IY/MS 1/1"), ("iy_ms_1_x","IY/MS 1/X"), ("iy_ms_1_2","IY/MS 1/2"),
        ("iy_ms_x_1","IY/MS X/1"), ("iy_ms_x_x","IY/MS X/X"), ("iy_ms_x_2","IY/MS X/2"),
        ("iy_ms_2_1","IY/MS 2/1"), ("iy_ms_2_x","IY/MS 2/X"), ("iy_ms_2_2","IY/MS 2/2"),
    ]),
    ("Toplam Gol", [
        ("tg_0_1","TG 0-1"), ("tg_2_3","TG 2-3"), ("tg_4_5","TG 4-5"), ("tg_6p","TG 6+"),
    ]),
    ("Tak\u0131m 1 Gol", [
        ("t1_1_5_ust","T1 1.5\u00dc"), ("t1_2_5_ust","T1 2.5\u00dc"),
    ]),
    ("Tak\u0131m 2 Gol", [
        ("t2_1_5_ust","T2 1.5\u00dc"), ("t2_2_5_ust","T2 2.5\u00dc"),
    ]),
]

class IddaaProApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("\u0130ddaaPro  \u2014  Ma\u00e7 Verisi & Oran Analizi")
        self.root.configure(bg=CLR_BG)
        self.root.geometry("1380x740")
        self.root.minsize(900, 580)
        self.root.resizable(True, True)

        self._set_icon()
        self._build_ui()

        self._driver: webdriver.Firefox | None = None
        self._thread: threading.Thread | None = None
        self._stop_flag = threading.Event()
        self._rows: list[dict] = []
        self._out_dir: Path = Path.home() / "Desktop"
        # Ayarlar: None=t\u00fcm\u00fc, set()=hi\u00e7biri, dolu set=sadece se\u00e7ilenler
        self._sel_leagues: set[str] | None = None
        self._sel_markets: set[str] | None = None

    # ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ Ãƒâ€Ã‚Â°kon ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
    def _set_icon(self):
        try:
            from PIL import Image, ImageTk
            jpg = BASE_DIR / "images.jpg"
            if jpg.exists():
                img = Image.open(jpg).resize((32, 32))
                self._icon_img = ImageTk.PhotoImage(img)
                self.root.iconphoto(True, self._icon_img)
        except Exception:
            pass

    # ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ UI ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬ÃƒÂ¢Ã¢â‚¬ÂÃ¢â€šÂ¬
    _HOVER_MAP = {
        CLR_BTN_G:   "#00E676",
        CLR_BTN_O:   "#FF9100",
        CLR_BTN_R:   "#FF5252",
        CLR_BTN_GRY: "#546E7A",
        CLR_BTN_Y:   "#FFE57F",
        "#1565C0":   "#1E88E5",
        "#1B5E20":   "#2E7D32",
    }

    def _btn(self, parent, text, cmd, bg=CLR_BTN_G, width=20, tooltip: str = ""):
        hover = self._HOVER_MAP.get(bg, bg)
        btn = tk.Button(parent, text=text, command=cmd,
                        bg=bg, fg="#0A1929" if bg in (CLR_GOLD, CLR_BTN_Y, CLR_ACCENT, "#00E676") else CLR_BTN_TXT,
                        font=("Segoe UI", 8, "bold"),
                        relief="flat", cursor="hand2",
                        activebackground=hover,
                        activeforeground="#0A1929" if bg in (CLR_GOLD, CLR_BTN_Y, CLR_ACCENT, "#00E676") else CLR_BTN_TXT,
                        width=width, pady=6, bd=0, highlightthickness=0)
        btn.bind("<Enter>", lambda e, b=btn, h=hover: b.configure(bg=h))
        btn.bind("<Leave>", lambda e, b=btn, n=bg:   b.configure(bg=n))
        if tooltip:
            _Tooltip(btn, tooltip)
        return btn

    def _lframe(self, parent, text):
        return tk.LabelFrame(parent, text=f"  {text}  ", bg=CLR_CARD,
                             font=("Segoe UI", 9, "bold"),
                             fg=CLR_ACCENT, padx=10, pady=8,
                             relief="solid", bd=1,
                             highlightbackground="#D7E6D5",
                             highlightthickness=0)

    def _build_ui(self):
        # ── Header ──────────────────────────────────────
        hdr = tk.Frame(self.root, bg=CLR_HEADER, height=64)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        try:
            from PIL import Image, ImageTk
            _logo = Image.open(BASE_DIR / "images.jpg").resize((42, 42))
            self._logo_img = ImageTk.PhotoImage(_logo)
            lbl_logo = tk.Label(hdr, image=self._logo_img, bg=CLR_HEADER,
                                bd=0, relief="flat", cursor="hand2")
            lbl_logo.pack(side="left", padx=(18, 10), pady=11)
        except Exception:
            tk.Label(hdr, text="\u26bd", bg=CLR_HEADER, fg=CLR_ACCENT,
                     font=("Segoe UI Symbol", 24)).pack(side="left", padx=(18, 10))

        title_blk = tk.Frame(hdr, bg=CLR_HEADER)
        title_blk.pack(side="left", pady=11)
        tk.Label(title_blk, text="\u0130ddaaPro", bg=CLR_HEADER, fg=CLR_ACCENT,
                 font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(title_blk, text="Ma\u00e7 Verisi & Oran Analizi",
                 bg=CLR_HEADER, fg=CLR_TXT_SEC,
                 font=("Segoe UI", 8)).pack(anchor="w")

        tk.Label(hdr, text=" v3.0 ", bg="#E6F6EA", fg=CLR_ACCENT_D,
                 font=("Segoe UI", 7, "bold"),
                 padx=6, pady=2).pack(side="right", padx=18)

        # Accent line
        tk.Frame(self.root, bg=CLR_ACCENT, height=2).pack(fill="x")

        # ── Control panel ───────────────────────────────
        ctrl = tk.Frame(self.root, bg=CLR_BG, pady=10)
        ctrl.pack(fill="x", padx=14)
        ctrl.columnconfigure(0, weight=3, minsize=280)
        ctrl.columnconfigure(1, weight=2, minsize=240)
        ctrl.columnconfigure(2, weight=2, minsize=200)

        # Card 1 – Veri Cekme
        frm_veri = self._lframe(ctrl, "Veri \u00c7ekme")
        frm_veri.grid(row=0, column=0, padx=(0, 6), pady=0, sticky="nsew")
        frm_veri.columnconfigure(0, weight=1)
        frm_veri.columnconfigure(1, weight=1)

        btn_baslat = tk.Button(frm_veri, text="\u25b6  GE\u00c7M\u0130\u015e MA\u00c7LARI \u00c7EK",
                               command=self._start_biten,
                               bg=CLR_ACCENT, fg="#0A1929",
                               font=("Segoe UI", 9, "bold"),
                               relief="flat", cursor="hand2",
                               activebackground=CLR_ACCENT_D,
                               activeforeground="#0A1929",
                               bd=0, highlightthickness=0, pady=8)
        btn_baslat.bind("<Enter>", lambda e: btn_baslat.configure(bg=CLR_ACCENT_D))
        btn_baslat.bind("<Leave>", lambda e: btn_baslat.configure(bg=CLR_ACCENT))
        _Tooltip(btn_baslat, "Se\u00e7ilen tarih aral\u0131\u011f\u0131ndaki ge\u00e7mi\u015f ma\u00e7lar\u0131 \u00e7ek")
        btn_baslat.grid(row=0, column=0, columnspan=2, pady=(2, 4), sticky="ew")

        btn_gunluk = tk.Button(frm_veri, text="\u26bd  G\u00dcNL\u00dcK MA\u00c7LARI \u00c7EK",
                               command=self._start_gunluk,
                               bg="#1565C0", fg="#FFFFFF",
                               font=("Segoe UI", 9, "bold"),
                               relief="flat", cursor="hand2",
                               activebackground="#1E88E5",
                               activeforeground="#FFFFFF",
                               bd=0, highlightthickness=0, pady=7)
        btn_gunluk.bind("<Enter>", lambda e: btn_gunluk.configure(bg="#1E88E5"))
        btn_gunluk.bind("<Leave>", lambda e: btn_gunluk.configure(bg="#1565C0"))
        _Tooltip(btn_gunluk, "Bug\u00fcn\u00fcn ma\u00e7lar\u0131n\u0131 \u00e7ek (g\u00fcnl\u00fck b\u00fclten)")
        btn_gunluk.grid(row=1, column=0, columnspan=2, pady=(0, 4), sticky="ew")

        self._btn(frm_veri, "\u23f9  \u0130PTAL", self._iptal, CLR_BTN_R, 12,
                  tooltip="Devam eden \u00e7ekimi iptal et").grid(
                  row=2, column=0, padx=(0, 3), pady=3, sticky="ew")
        self._btn(frm_veri, "\U0001f5d1  TEM\u0130ZLE", self._yenile, CLR_BTN_O, 12,
                  tooltip="Listeyi temizle \u2014 t\u00fcm \u00e7ekilen ma\u00e7lar\u0131 sil").grid(
                  row=2, column=1, padx=(3, 0), pady=3, sticky="ew")
        self._btn(frm_veri, "\u2699  AYARLAR", self._ayarlar, CLR_BTN_GRY, 24,
                  tooltip="Lig & oran se\u00e7imi, Excel klas\u00f6r\u00fc").grid(
                  row=3, column=0, columnspan=2, padx=0, pady=3, sticky="ew")

        # Card 2 – Tarih Aralığı
        frm_tarih = self._lframe(ctrl, "Tarih Aral\u0131\u011f\u0131")
        frm_tarih.grid(row=0, column=1, padx=6, pady=0, sticky="nsew")
        frm_tarih.columnconfigure(1, weight=1)

        cal_opts = dict(
            date_pattern="dd.MM.yyyy",
            mindate=dt.date(2015, 1, 1),
            maxdate=dt.date(2028, 1, 1),
            background="#FFFFFF",
            foreground=CLR_TXT_PRI,
            headersbackground="#E7F4EA",
            headersforeground=CLR_ACCENT_D,
            selectbackground=CLR_ACCENT,
            selectforeground="#FFFFFF",
            normalbackground="#FFFFFF",
            weekendbackground="#F5FAF4",
            normalforeground=CLR_TXT_PRI,
            weekendforeground=CLR_ACCENT_D,
            othermonthforeground="#8CA191",
            othermonthbackground="#FFFFFF",
            font=("Segoe UI", 8),
        )
        tk.Label(frm_tarih, text="Ba\u015flang\u0131\u00e7:", bg=CLR_CARD,
                 fg=CLR_TXT_SEC, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=6)
        self._start_cal = DateEntry(frm_tarih, width=12, **cal_opts)
        self._start_cal.set_date(dt.date(2021, 1, 1))
        self._start_cal.grid(row=0, column=1, padx=(10, 0), pady=6, sticky="ew")

        tk.Label(frm_tarih, text="Biti\u015f:", bg=CLR_CARD,
                 fg=CLR_TXT_SEC, font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", pady=6)
        self._end_cal = DateEntry(frm_tarih, width=12, **cal_opts)
        self._end_cal.set_date(dt.date.today())
        self._end_cal.grid(row=1, column=1, padx=(10, 0), pady=6, sticky="ew")

        # Card 3 – Excel
        frm_excel = self._lframe(ctrl, "Excel'e Aktar")
        frm_excel.grid(row=0, column=2, padx=(6, 0), pady=0, sticky="nsew")
        self._btn(frm_excel, "\u2b07  T\u00dcM\u00dcN\u00dc AKTAR", self._export_all, CLR_ACCENT, 16,
                  tooltip="T\u00fcm verileri Excel dosyas\u0131na aktar").pack(
                  pady=14, fill="x", ipady=8)

        # ── Table ───────────────────────────────────────
        frm_tree = tk.Frame(self.root, bg=CLR_CARD, bd=1, relief="solid",
                            highlightbackground="#D6E4D4", highlightthickness=1)
        frm_tree.pack(fill="both", expand=True, padx=14, pady=(6, 0))

        cols = [c[0] for c in TREE_COLS]
        self._tree = ttk.Treeview(frm_tree, columns=cols, show="headings",
                                   selectmode="extended")
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview.Heading",
                        background=CLR_GRID_H, foreground="#FFFFFF",
                        font=("Segoe UI", 8, "bold"), relief="flat", padding=(4, 6))
        style.map("Treeview.Heading",
                  background=[("active", "#129455")])
        style.configure("Treeview",
                        font=("Segoe UI", 8), rowheight=26,
                        background=CLR_ROW_ODD, fieldbackground=CLR_ROW_ODD,
                        foreground=CLR_TXT_PRI, borderwidth=0, relief="flat")
        style.map("Treeview",
                  background=[("selected", CLR_SELECTED)],
                  foreground=[("selected", "#153726")])
        self._tree.tag_configure("odd",  background=CLR_ROW_ODD,  foreground=CLR_TXT_PRI)
        self._tree.tag_configure("even", background=CLR_ROW_EVEN, foreground=CLR_TXT_PRI)

        style.configure("Vertical.TScrollbar",
                        background="#D6EAD9", troughcolor="#F3F8F2",
                        arrowcolor=CLR_ACCENT_D, borderwidth=0)
        style.map("Vertical.TScrollbar",
                  background=[("active", "#C3DEC8")])
        style.configure("Horizontal.TScrollbar",
                        background="#D6EAD9", troughcolor="#F3F8F2",
                        arrowcolor=CLR_ACCENT_D, borderwidth=0)
        style.map("Horizontal.TScrollbar",
                  background=[("active", "#C3DEC8")])

        for col, width in TREE_COLS:
            self._tree.heading(col, text=col)
            self._tree.column(col, width=width, minwidth=38, anchor="center")

        vsb = ttk.Scrollbar(frm_tree, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(frm_tree, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        # ── Status bar ──────────────────────────────────
        tk.Frame(self.root, bg="#D8E9D8", height=1).pack(fill="x")
        status_bar = tk.Frame(self.root, bg=CLR_HEADER, height=32)
        status_bar.pack(fill="x")
        status_bar.pack_propagate(False)

        style.configure("IP.Horizontal.TProgressbar",
                        troughcolor="#E4EFE3", background=CLR_ACCENT, thickness=8)
        self._progress = ttk.Progressbar(status_bar, length=100, mode="indeterminate",
                                          style="IP.Horizontal.TProgressbar")
        self._progress.pack(side="left", padx=10, pady=12)

        self._status_var = tk.StringVar(value="\u2713 Haz\u0131r")
        tk.Label(status_bar, textvariable=self._status_var,
                 bg=CLR_HEADER, fg=CLR_GOLD, font=("Segoe UI", 8)).pack(side="left", padx=6)

        self._count_var = tk.StringVar(value="Ma\u00e7lar (0)")
        tk.Label(status_bar, textvariable=self._count_var,
                 bg=CLR_HEADER, fg=CLR_ACCENT, font=("Segoe UI", 9, "bold")).pack(side="right", padx=14)

    def _set_status(self, msg: str, count: int | None = None):
        if "hata" in msg.lower() or "HATA" in msg:
            icon = "\u26a0 "
        elif "tamamland\u0131" in msg.lower():
            icon = "\u2713 "
        elif "haz\u0131r" in msg.lower():
            icon = "\u2713 "
        else:
            icon = "\u23f3 "
        self._status_var.set(icon + msg)
        if count is not None:
            self._count_var.set(f"Ma\u00e7lar ({count})")
        self.root.update_idletasks()

    def _add_row(self, r: dict):
        mac = f"{r.get('ev_sahibi','')} - {r.get('konuk_ekip','')}"
        tag = "even" if len(self._tree.get_children()) % 2 == 0 else "odd"
        sel = self._sel_markets
        mkt_keys = set(KEYS[8:])
        def mv(k):
            if sel is not None and k in mkt_keys and k not in sel:
                return ""
            return r.get(k, "")
        self._tree.insert("", "end", tags=(tag,), values=(
            mac,
            r.get("mac_tarihi",""),
            r.get("mac_saati",""),
            r.get("lig",""),
            r.get("ms_kodu",""),
            r.get("ilk_yari_skor",""),
            r.get("mac_skoru",""),
            mv("ms1"), mv("ms0"), mv("ms2"),
            mv("cs_1x"), mv("cs_12"), mv("cs_x2"),
            mv("iy1"), mv("iy0"), mv("iy2"),
            mv("au_0_5_alt"), mv("au_0_5_ust"),
            mv("au_1_5_alt"), mv("au_1_5_ust"),
            mv("au_2_5_alt"), mv("au_2_5_ust"),
            mv("au_3_5_alt"), mv("au_3_5_ust"),
            mv("au_4_5_alt"), mv("au_4_5_ust"),
            mv("kg_var"), mv("kg_yok"),
            mv("hnd_1"), mv("hnd_x"), mv("hnd_2"),
            mv("hnd2_1"), mv("hnd2_x"), mv("hnd2_2"),
            mv("iy_au_05_alt"), mv("iy_au_05_ust"),
            mv("iy_au_15_alt"), mv("iy_au_15_ust"),
            mv("iy_ms_1_1"), mv("iy_ms_1_x"), mv("iy_ms_1_2"),
            mv("iy_ms_x_1"), mv("iy_ms_x_x"), mv("iy_ms_x_2"),
            mv("iy_ms_2_1"), mv("iy_ms_2_x"), mv("iy_ms_2_2"),
            mv("tg_0_1"), mv("tg_2_3"), mv("tg_4_5"), mv("tg_6p"),
            mv("t1_1_5_ust"), mv("t1_2_5_ust"),
            mv("t2_1_5_ust"), mv("t2_2_5_ust"),
        ))
    def _parse_date(self, s: str) -> dt.date:
        return dt.datetime.strptime(s.strip(), "%d.%m.%Y").date()

    def _noop(self):
        messagebox.showinfo("Bilgi", "Bu \u00f6zellik mevcut de\u011fil.")

    def _iptal(self):
        self._stop_flag.set()
        with contextlib.suppress(Exception):
            if self._driver:
                self._driver.quit()
        self._set_status("\u0130ptal ediliyor...")

    def _yenile(self):
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._rows.clear()
        self._count_var.set("Ma\u00e7lar (0)")
        self._status_var.set("\u2713 Haz\u0131r")

    def _sec_klasor(self):
        folder = filedialog.askdirectory(title="Excel \u00e7\u0131kt\u0131 klas\u00f6r\u00fc",
                                         initialdir=str(self._out_dir))
        if folder:
            self._out_dir = Path(folder)
            messagebox.showinfo("Klas\u00f6r", f"\u00c7\u0131kt\u0131 klas\u00f6r\u00fc:\n{self._out_dir}")

    def _ayarlar(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("\u2699  Ayarlar \u2014 Lig & Oran Se\u00e7imi")
        dlg.geometry("980x680")
        dlg.resizable(True, True)
        dlg.configure(bg=CLR_BG)
        dlg.transient(self.root)
        dlg.grab_set()

        # ── Ana iki panel ──────────────────────────────────────────────────
        panels = tk.Frame(dlg, bg=CLR_BG)
        panels.pack(fill="both", expand=True, padx=10, pady=(8, 4))
        panels.columnconfigure(0, weight=55, minsize=460)
        panels.columnconfigure(1, weight=45, minsize=380)
        panels.rowconfigure(0, weight=1)

        # ────────── SOL: LİGLER ───────────────────────────────────────────
        frm_lig = self._lframe(panels, "\U0001f3c6  Ligler")
        frm_lig.grid(row=0, column=0, padx=(0, 5), sticky="nsew")
        frm_lig.rowconfigure(1, weight=1)
        frm_lig.columnconfigure(0, weight=1)

        lig_btn_row = tk.Frame(frm_lig, bg=CLR_CARD)
        lig_btn_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        lig_vars: dict[str, tk.BooleanVar] = {}

        def _lig_tumunu():
            for v in lig_vars.values(): v.set(True)
        def _lig_hicbiri():
            for v in lig_vars.values(): v.set(False)

        tk.Button(lig_btn_row, text="T\u00fcm\u00fcn\u00fc Se\u00e7", command=_lig_tumunu,
                  bg=CLR_BTN_GRY, fg="white", font=("Segoe UI", 7, "bold"),
                  relief="flat", pady=3, padx=10, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(lig_btn_row, text="Hi\u00e7birini Se\u00e7me", command=_lig_hicbiri,
                  bg=CLR_BTN_O, fg="white", font=("Segoe UI", 7, "bold"),
                  relief="flat", pady=3, padx=10, cursor="hand2").pack(side="left")

        cvs_lig = tk.Canvas(frm_lig, bg=CLR_CARD, highlightthickness=0)
        sb_lig  = ttk.Scrollbar(frm_lig, orient="vertical", command=cvs_lig.yview)
        cvs_lig.configure(yscrollcommand=sb_lig.set)
        sb_lig.grid(row=1, column=1, sticky="ns")
        cvs_lig.grid(row=1, column=0, sticky="nsew")
        inner_lig = tk.Frame(cvs_lig, bg=CLR_CARD)
        win_lig = cvs_lig.create_window((0, 0), window=inner_lig, anchor="nw")
        inner_lig.bind("<Configure>", lambda e: (
            cvs_lig.configure(scrollregion=cvs_lig.bbox("all")),
            cvs_lig.itemconfig(win_lig, width=cvs_lig.winfo_width()),
        ))
        cvs_lig.bind("<Configure>", lambda e: cvs_lig.itemconfig(win_lig, width=e.width))

        all_league_keys = {lig_key(c, l) for c, ls in LEAGUE_LIST for l in ls}
        for country, leagues in LEAGUE_LIST:
            # ── \u00dclke ba\u015fl\u0131\u011f\u0131 ──
            hdr_row = tk.Frame(inner_lig, bg=CLR_CARD)
            hdr_row.pack(fill="x", padx=4, pady=(6, 0))
            grp_vars: list[tk.BooleanVar] = []
            ctry_var = tk.BooleanVar()

            def _toggle_country(cv=None, gvs=None):
                state = cv.get()
                for gv in gvs: gv.set(state)
            def _update_ctry(cv=None, gvs=None, *a):
                cv.set(all(gv.get() for gv in gvs))

            tk.Checkbutton(
                hdr_row, text=country, variable=ctry_var,
                bg=CLR_CARD, fg=CLR_ACCENT, activebackground=CLR_CARD,
                selectcolor=CLR_CARD, activeforeground=CLR_ACCENT,
                font=("Segoe UI", 8, "bold"), anchor="w",
                command=lambda cv=ctry_var, gvs=grp_vars: _toggle_country(cv, gvs)
            ).pack(anchor="w")

            items_frame = tk.Frame(inner_lig, bg=CLR_CARD)
            items_frame.pack(fill="x")
            cols_per_row = 2
            for i, lig_name in enumerate(leagues):
                k = lig_key(country, lig_name)
                var = tk.BooleanVar(value=(self._sel_leagues is None or k in self._sel_leagues))
                lig_vars[k] = var
                grp_vars.append(var)
                col = i % cols_per_row
                row = i // cols_per_row
                lig_label = f"{country} - {lig_name}"
                tk.Checkbutton(
                    items_frame, text=lig_label, variable=var,
                    bg=CLR_CARD, fg=CLR_TXT_PRI, activebackground=CLR_CARD,
                    selectcolor=CLR_CARD, activeforeground=CLR_TXT_PRI,
                    font=("Segoe UI", 8), anchor="w"
                ).grid(row=row, column=col, sticky="w", padx=(20, 4))
                var.trace_add("write", lambda *a, cv=ctry_var, gvs=grp_vars: _update_ctry(cv, gvs))

            _update_ctry(ctry_var, grp_vars)

        # ────────── SAĞ: MARKETLER ────────────────────────────────────────
        frm_mkt = self._lframe(panels, "\U0001f4ca  Oranlar / Marketler")
        frm_mkt.grid(row=0, column=1, padx=(5, 0), sticky="nsew")
        frm_mkt.rowconfigure(1, weight=1)
        frm_mkt.columnconfigure(0, weight=1)

        mkt_btn_row = tk.Frame(frm_mkt, bg=CLR_CARD)
        mkt_btn_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        mkt_vars: dict[str, tk.BooleanVar] = {}

        def _mkt_tumunu():
            for v in mkt_vars.values(): v.set(True)
        def _mkt_hicbiri():
            for v in mkt_vars.values(): v.set(False)

        tk.Button(mkt_btn_row, text="T\u00fcm\u00fcn\u00fc Se\u00e7", command=_mkt_tumunu,
                  bg=CLR_BTN_GRY, fg="white", font=("Segoe UI", 7, "bold"),
                  relief="flat", pady=3, padx=10, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(mkt_btn_row, text="Hi\u00e7birini Se\u00e7me", command=_mkt_hicbiri,
                  bg=CLR_BTN_O, fg="white", font=("Segoe UI", 7, "bold"),
                  relief="flat", pady=3, padx=10, cursor="hand2").pack(side="left")

        cvs_mkt = tk.Canvas(frm_mkt, bg=CLR_CARD, highlightthickness=0)
        sb_mkt  = ttk.Scrollbar(frm_mkt, orient="vertical", command=cvs_mkt.yview)
        cvs_mkt.configure(yscrollcommand=sb_mkt.set)
        sb_mkt.grid(row=1, column=1, sticky="ns")
        cvs_mkt.grid(row=1, column=0, sticky="nsew")
        inner_mkt = tk.Frame(cvs_mkt, bg=CLR_CARD)
        win_mkt = cvs_mkt.create_window((0, 0), window=inner_mkt, anchor="nw")
        inner_mkt.bind("<Configure>", lambda e: (
            cvs_mkt.configure(scrollregion=cvs_mkt.bbox("all")),
            cvs_mkt.itemconfig(win_mkt, width=cvs_mkt.winfo_width()),
        ))
        cvs_mkt.bind("<Configure>", lambda e: cvs_mkt.itemconfig(win_mkt, width=e.width))

        all_mkt_keys = {k for _, items in MARKET_GROUPS for k, _ in items}
        for grp_name, grp_items in MARKET_GROUPS:
            item_vars: list[tk.BooleanVar] = []
            grp_var = tk.BooleanVar()

            def _toggle_grp(gv=None, ivs=None):
                state = gv.get()
                for iv in ivs: iv.set(state)
            def _update_grp(gv=None, ivs=None, *a):
                gv.set(all(iv.get() for iv in ivs))

            # Grup başlığı (tıklanabilir master checkbox)
            tk.Checkbutton(
                inner_mkt, text=grp_name, variable=grp_var,
                bg=CLR_CARD, fg=CLR_ACCENT, activebackground=CLR_CARD,
                selectcolor=CLR_CARD, activeforeground=CLR_ACCENT,
                font=("Segoe UI", 8, "bold"), anchor="w",
                command=lambda gv=grp_var, ivs=item_vars: _toggle_grp(gv, ivs)
            ).pack(anchor="w", padx=4, pady=(5, 0))

            sub_frame = tk.Frame(inner_mkt, bg=CLR_CARD)
            sub_frame.pack(fill="x")
            for col_i, (mkt_key, mkt_label) in enumerate(grp_items):
                iv = tk.BooleanVar(value=(self._sel_markets is None or mkt_key in self._sel_markets))
                mkt_vars[mkt_key] = iv
                item_vars.append(iv)
                tk.Checkbutton(
                    sub_frame, text=mkt_label, variable=iv,
                    bg=CLR_CARD, fg=CLR_TXT_PRI, activebackground=CLR_CARD,
                    selectcolor=CLR_CARD, activeforeground=CLR_TXT_PRI,
                    font=("Segoe UI", 7), anchor="w"
                ).grid(row=col_i // 3, column=col_i % 3, sticky="w", padx=(20, 2))
                iv.trace_add("write", lambda *a, gv=grp_var, ivs=item_vars: _update_grp(gv, ivs))

            _update_grp(grp_var, item_vars)

        # ────────── ALT BUTON ÇUBUĞU ─────────────────────────────────────
        btn_bar = tk.Frame(dlg, bg=CLR_BG, height=50)
        btn_bar.pack(fill="x", padx=10, pady=(4, 8))
        btn_bar.pack_propagate(False)

        def _kaydet():
            checked_lig = {k for k, v in lig_vars.items() if v.get()}
            self._sel_leagues = None if checked_lig == all_league_keys else checked_lig

            checked_mkt = {k for k, v in mkt_vars.items() if v.get()}
            self._sel_markets = None if checked_mkt == all_mkt_keys else checked_mkt

            n_lig = "T\u00fcm\u00fc" if self._sel_leagues is None else str(len(self._sel_leagues))
            n_mkt = "T\u00fcm\u00fc" if self._sel_markets is None else str(len(self._sel_markets))
            dlg.destroy()
            self._set_status(f"\u2699 Ayarlar kaydedildi \u2014 Lig: {n_lig} | Oran: {n_mkt}")

        tk.Button(btn_bar, text="\U0001f4be  KAYDET", command=_kaydet,
                  bg=CLR_ACCENT, fg="#0A1929", font=("Segoe UI", 9, "bold"),
                  relief="flat", pady=8, padx=24, cursor="hand2").pack(side="left")
        tk.Button(btn_bar, text="\u274c  \u0130PTAL", command=dlg.destroy,
                  bg=CLR_BTN_R, fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", pady=8, padx=18, cursor="hand2").pack(side="left", padx=(8, 0))
        tk.Button(btn_bar, text="\U0001f4c2  Excel Klas\u00f6r\u00fc",
                  command=self._sec_klasor,
                  bg=CLR_BTN_GRY, fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", pady=8, padx=18, cursor="hand2").pack(side="right")

    def _export_all(self):
        if not self._rows:
            messagebox.showwarning("Uyar\u0131", "D\u0131\u015fa aktar\u0131lacak veri yok.")
            return
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self._out_dir / f"iddaapro_{ts}.xlsx"
        try:
            export_excel(self._rows, path, market_keys=self._sel_markets)
            messagebox.showinfo("Ba\u015far\u0131l\u0131", f"Excel kaydedildi:\n{path}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def _start_biten(self):
        if self._thread and self._thread.is_alive():
            messagebox.showwarning("Uyar\u0131", "Zaten \u00e7al\u0131\u015f\u0131yor.")
            return
        start = self._start_cal.get_date()
        end = self._end_cal.get_date()
        if start > end:
            messagebox.showerror("Hata", "Ba\u015flang\u0131\u00e7 tarihi biti\u015f tarihinden b\u00fcy\u00fck olamaz.")
            return

        self._yenile()
        self._stop_flag.clear()
        self._thread = threading.Thread(
            target=self._scrape_worker,
            args=(start, end),
            daemon=True
        )
        self._thread.start()
        self._progress.start(10)

    def _start_gunluk(self):
        if self._thread and self._thread.is_alive():
            messagebox.showwarning("Uyar\u0131", "Zaten \u00e7al\u0131\u015f\u0131yor.")
            return
        self._yenile()
        self._stop_flag.clear()
        self._thread = threading.Thread(
            target=self._scrape_gunluk_worker,
            daemon=True
        )
        self._thread.start()
        self._progress.start(10)

    def _scrape_gunluk_worker(self):
        start = self._start_cal.get_date()
        end = self._end_cal.get_date()
        total_days = (end - start).days + 1
        print(f'Günlük maçlar çekiliyor: {start:%d.%m.%Y} - {end:%d.%m.%Y} ({total_days} gün)')
        self._set_status('Tarayıcı açılıyor...')
        try:
            driver = build_driver(headless=True)
            self._driver = driver
        except Exception as e:
            self._set_status(f'Driver hatası: {e}')
            self._progress.stop()
            return

        total = 0
        try:
            self._set_status('Ana sayfa yükleniyor...')
            open_main_page(driver)
            _init_session_cookies(driver)

            lig_filtre = self._sel_leagues if self._sel_leagues is not None else _get_all_league_keys()
            mkt_filtre = self._sel_markets

            for day_idx in range(total_days):
                if self._stop_flag.is_set():
                    break
                current_date = start + dt.timedelta(days=day_idx)

                self._set_status(f'[{day_idx+1}/{total_days}] {current_date:%d.%m.%Y} - maçlar toplanıyor...')
                pick_date(driver, current_date)
                wait_rows(driver, timeout=5)
                close_popups(driver)
                time.sleep(0.15)

                summaries = collect_summaries(driver)
                print(f'  [{current_date:%d.%m.%Y}] {len(summaries)} maç bulundu', flush=True)

                onceki = len(summaries)
                filtered = []
                no_lig_count = 0
                for s in summaries:
                    lk = (s.get('lig_key') or '').strip()
                    if lk:
                        if lig_filtreli_key(lk, lig_filtre):
                            filtered.append(s)
                    else:
                        if lig_filtreli(s.get('iddaa_link', ''), lig_filtre):
                            filtered.append(s)
                        else:
                            no_lig_count += 1
                summaries = filtered
                atlanan = onceki - len(summaries)
                if no_lig_count:
                    print(f'  [{current_date:%d.%m.%Y}] {no_lig_count} maçta lig bilgisi yok, atlandı', flush=True)
                print(f'  [{current_date:%d.%m.%Y}] Lig filtresi: {onceki} → {len(summaries)} maç ({atlanan} atlandı)', flush=True)
                self._set_status(f'[{day_idx+1}/{total_days}] {current_date:%d.%m.%Y} - {len(summaries)} maç | {atlanan} atlandı')

                if not summaries:
                    continue

                self._set_status(f'[{day_idx+1}/{total_days}] {current_date:%d.%m.%Y} - {len(summaries)} mac detayı çekiliyor...')
                retry_selenium = []
                with ThreadPoolExecutor(max_workers=5) as pool:
                    futures = {
                        pool.submit(scrape_match_fast, s, match_date=current_date, market_keys=mkt_filtre): s
                        for s in summaries
                    }
                    for fut in as_completed(futures):
                        if self._stop_flag.is_set():
                            break
                        try:
                            r = fut.result()
                            has_odds = any(r.get(k) for k in ('ms1','ms0','ms2','au_2_5_alt','au_2_5_ust','kg_var'))
                            if not has_odds:
                                retry_selenium.append(futures[fut])
                                continue
                            if not _row_is_valid(r):
                                continue
                            self._rows.append(r)
                            total += 1
                            self.root.after(0, self._add_row, r)
                            self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            self._set_status(f'[{day_idx+1}/{total_days}] [{total}] {r.get("ev_sahibi","")} - {r.get("konuk_ekip","")}')
                        except Exception as ex:
                            print(f'    HATA: {ex}')

                if retry_selenium and not self._stop_flag.is_set():
                    still_missing = []
                    self._set_status(f'[{day_idx+1}/{total_days}] {len(retry_selenium)} mac HTTP retry...')
                    with ThreadPoolExecutor(max_workers=5) as retry_pool:
                        retry_futures = {
                            retry_pool.submit(scrape_match_fast, s, match_date=current_date, market_keys=mkt_filtre): s
                            for s in retry_selenium
                        }
                        for fut in as_completed(retry_futures):
                            if self._stop_flag.is_set():
                                break
                            try:
                                r = fut.result()
                                if not _row_is_valid(r):
                                    continue
                                self._rows.append(r)
                                total += 1
                                self.root.after(0, self._add_row, r)
                                self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            except Exception:
                                pass

                    if still_missing and not self._stop_flag.is_set():
                        print(f'  Selenium fallback: {len(still_missing)} mac')
                        self._set_status(f'{len(still_missing)} mac Selenium ile deneniyor...')
                        for idx_s, s in enumerate(still_missing, 1):
                            if self._stop_flag.is_set():
                                break
                            label = f"{s['ev_sahibi']} - {s['konuk_ekip']}"
                            self._set_status(f'[Selenium {idx_s}/{len(still_missing)}] {label}')
                            try:
                                r = scrape_match(driver, s, self._stop_flag, match_date=current_date, market_keys=mkt_filtre)
                                if self._stop_flag.is_set():
                                    break
                                if not _row_is_valid(r):
                                    continue
                                self._rows.append(r)
                                total += 1
                                self.root.after(0, self._add_row, r)
                                self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            except Exception as ex:
                                print(f'    Selenium HATA: {ex}')

        except Exception as ex:
            print(f'  Gunluk hata: {ex}')
            self._set_status(f'Hata: {ex}')
        finally:
            with contextlib.suppress(Exception):
                driver.quit()
            self._driver = None
            self._progress.stop()
            if self._stop_flag.is_set():
                self._set_status(f"İptal edildi - {total} mac", total)
            else:
                self._set_status(f"Tamamlandı - {total} mac ({total_days} gün)", total)

    def _scrape_worker(self, start: dt.date, end: dt.date):
        print(f'Scraper başladı: {start:%d.%m.%Y} - {end:%d.%m.%Y}')
        total_days = (end - start).days + 1
        self._set_status('Tarayıcı başlatılıyor...')

        # Selenium ile maç listesi toplanacak, detaylar HTTP ile çekilecek
        try:
            driver = build_driver(headless=True)
            self._driver = driver
        except Exception as e:
            self._set_status(f'Driver hatası: {e}')
            self._progress.stop()
            return

        # Tumunu Sec = None → tanimli ligleri kullan (Cezayir/Fas/Misir vb. atlanir)
        lig_filtre = self._sel_leagues if self._sel_leagues is not None else _get_all_league_keys()
        mkt_filtre = self._sel_markets
        total = 0

        def _ensure_driver():
            nonlocal driver
            """Driver çökmüşse yeniden başlat."""
            try:
                driver.current_url  # driver yaşıyor mu test et
            except Exception:
                print('  Driver çökmüş, yeniden başlatılıyor...', flush=True)
                with contextlib.suppress(Exception):
                    driver.quit()
                driver = build_driver(headless=True)
                self._driver = driver
                open_main_page(driver)
                _init_session_cookies(driver)
            return driver

        def _collect_day(cur_date):
            """Bir gun icin Selenium ile mac listesi topla."""
            nonlocal driver
            summaries = []

            selenium_ok = False
            for retry in range(2):
                try:
                    driver = _ensure_driver()
                    pick_date(driver, cur_date)
                    wait_rows(driver, timeout=5)
                    close_popups(driver)
                    time.sleep(0.15)
                    summaries = collect_summaries(driver)
                    selenium_ok = True
                    break
                except Exception as e:
                    print(f'  [{cur_date:%d.%m.%Y}] Selenium deneme {retry+1} hata: {e}', flush=True)
                    if retry == 0:
                        try:
                            with contextlib.suppress(Exception):
                                driver.quit()
                            driver = build_driver(headless=True)
                            self._driver = driver
                            open_main_page(driver)
                            _init_session_cookies(driver)
                        except Exception as e2:
                            print(f'  [{cur_date:%d.%m.%Y}] Driver yeniden baslatma hatasi: {e2}', flush=True)

            if selenium_ok:
                print(f'  [{cur_date:%d.%m.%Y}] {len(summaries)} maç bulundu', flush=True)

            return summaries

        cur = start
        try:
            self._set_status('Ana sayfa yükleniyor...')
            open_main_page(driver)
            _init_session_cookies(driver)

            day_idx = 0
            while cur <= end:
                if self._stop_flag.is_set():
                    break

                day_idx += 1
                self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - maçlar toplanıyor...')
                print(f'  [{cur:%d.%m.%Y}] Maclar toplanıyor...', flush=True)

                summaries = _collect_day(cur)

                # -- LIG FILTRELEME (detay cekmeden ONCE) --
                onceki = len(summaries)
                if lig_filtre:
                    filtered = []
                    no_lig_count = 0
                    for s in summaries:
                        lk = (s.get('lig_key') or '').strip()
                        if lk:
                            if lig_filtreli_key(lk, lig_filtre):
                                filtered.append(s)
                            # else: lig_key var ama seçili ligde değil → atla
                        else:
                            # Lig bilgisi yok — URL'den dene, bulamazsa ATLA
                            if lig_filtreli(s.get('iddaa_link', ''), lig_filtre):
                                filtered.append(s)
                            else:
                                no_lig_count += 1
                    summaries = filtered
                    atlanan = onceki - len(summaries)
                    if no_lig_count:
                        print(f'  [{cur:%d.%m.%Y}] {no_lig_count} maçta lig bilgisi yok, atlandı', flush=True)
                    print(f'  [{cur:%d.%m.%Y}] Lig filtresi: {onceki} → {len(summaries)} maç ({atlanan} atlandı)', flush=True)
                    self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - {len(summaries)} maç (seçili liglerden) | {atlanan} atlandı')
                else:
                    self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - {len(summaries)} maç bulundu')

                if not summaries:
                    cur += dt.timedelta(days=1)
                    continue

                # -- MAC DETAYLARI CEKME (Paralel HTTP) --
                self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - {len(summaries)} maç detayı çekiliyor...')
                done_count = 0
                no_odds_count = 0
                invalid_count = 0
                retry_selenium = []
                with ThreadPoolExecutor(max_workers=5) as pool:
                    futures = {
                        pool.submit(scrape_match_fast, s, match_date=cur, market_keys=mkt_filtre): s
                        for s in summaries
                    }
                    for fut in as_completed(futures):
                        if self._stop_flag.is_set():
                            break
                        try:
                            r = fut.result()
                            done_count += 1
                            if done_count % 20 == 0 or done_count == len(summaries):
                                print(f'  [{cur:%d.%m.%Y}] Detay: {done_count}/{len(summaries)} tamamlandı', flush=True)
                            # Lig filtresi onceden uygulandi, tekrar kontrol gereksiz
                            # Oran verisi bos mu kontrol — bossa Selenium'a birak
                            has_odds = any(r.get(k) for k in ('ms1','ms0','ms2','au_2_5_alt','au_2_5_ust','kg_var'))
                            if not has_odds:
                                no_odds_count += 1
                                retry_selenium.append(futures[fut])
                                continue
                            if not _row_is_valid(r):
                                invalid_count += 1
                                continue
                            self._rows.append(r)
                            total += 1
                            self.root.after(0, self._add_row, r)
                            self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} [{done_count}/{len(summaries)}] {r.get("ev_sahibi","")} - {r.get("konuk_ekip","")}')
                        except Exception as ex:
                            done_count += 1
                            print(f'    HATA: {ex}')

                print(f'  [{cur:%d.%m.%Y}] Sonuç: {total} geçerli, {no_odds_count} oran yok, {invalid_count} geçersiz (skor eksik)', flush=True)

                # HTTP ile oran alinamayanlar icin: once HTTP retry (overview link ile), sonra Selenium
                if retry_selenium and not self._stop_flag.is_set():
                    still_missing = []
                    self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - {len(retry_selenium)} maç HTTP retry...')
                    with ThreadPoolExecutor(max_workers=5) as retry_pool:
                        retry_futures = {
                            retry_pool.submit(scrape_match_fast, s, match_date=cur, market_keys=mkt_filtre): s
                            for s in retry_selenium
                        }
                        for fut in as_completed(retry_futures):
                            if self._stop_flag.is_set():
                                break
                            try:
                                r = fut.result()
                                if not _row_is_valid(r):
                                    continue
                                self._rows.append(r)
                                total += 1
                                self.root.after(0, self._add_row, r)
                                self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            except Exception:
                                pass

                    # Hala alinamayanlar icin Selenium fallback (sıralı, ama sayı az olacak)
                    if still_missing and not self._stop_flag.is_set():
                        print(f'  [{cur:%d.%m.%Y}] Selenium fallback: {len(still_missing)} mac', flush=True)
                        self._set_status(f'[{day_idx}/{total_days}] {cur:%d.%m.%Y} - {len(still_missing)} maç Selenium ile deneniyor...')
                        for idx_s, s in enumerate(still_missing, 1):
                            if self._stop_flag.is_set():
                                break
                            label = f"{s['ev_sahibi']} - {s['konuk_ekip']}"
                            self._set_status(f'[{day_idx}/{total_days}] [Selenium {idx_s}/{len(still_missing)}] {label}')
                            try:
                                r = scrape_match(driver, s, self._stop_flag, match_date=cur, market_keys=mkt_filtre)
                                if self._stop_flag.is_set():
                                    break
                                if not _row_is_valid(r):
                                    continue
                                self._rows.append(r)
                                total += 1
                                self.root.after(0, self._add_row, r)
                                self.root.after(0, self._count_var.set, f'Maclar ({total})')
                            except Exception as ex:
                                print(f'    Selenium HATA: {ex}')

                cur += dt.timedelta(days=1)

        except Exception as ex:
            print(f'  Genel hata: {ex}')
            self._set_status(f'Hata: {ex}')
        finally:
            with contextlib.suppress(Exception):
                driver.quit()
            self._driver = None
            self._progress.stop()
            if self._stop_flag.is_set():
                self._set_status(f"\u0130ptal edildi - {total} ma\u00e7", total)
                print(f"\n\u0130ptal edildi. Toplam: {total} ma\u00e7")
            else:
                self._set_status(f"Tamamland\u0131 - {total} ma\u00e7", total)
                print(f"\nTamamland\u0131. Toplam: {total} ma\u00e7")

def main():
    root = tk.Tk()
    app = IddaaProApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
