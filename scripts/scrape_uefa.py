"""
scrape_uefa.py - Mackolik'ten UEFA maclarini cekip Excel'e ekler

Kullanim:
    python scripts/scrape_uefa.py --start 2020-01-01 --end 2021-06-30
    python scripts/scrape_uefa.py --start 2020-01-01 --end 2021-06-30 --resume
    python scripts/scrape_uefa.py --dry-run

GitHub Actions icin:
- --output  : cikti excel dosyasi (varsayilan: output/iddaa_uefa_START_END.xlsx)
- --resume  : progress.json'dan kaldigi yerden devam
- --max-hours N : N saat sonra kaydet ve cik (Actions timeout icin)
"""
from __future__ import annotations
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import argparse
import datetime as dt
import json
import time
import unicodedata
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# iddaapro'dan bagimsiz — hicbir dis import yok
import threading
from bs4 import BeautifulSoup

# ── Throttle ────────────────────────────────────────────────────────────────
_throttle_delay = 0.02
_throttle_lock  = threading.Lock()

def _throttle_hit():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = min(_throttle_delay * 1.3, 0.5)

def _throttle_ok():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = max(_throttle_delay * 0.5, 0.02)

# ── Yardimci ────────────────────────────────────────────────────────────────
def _norm(v):
    if not v: return ""
    return " ".join(v.replace("\xa0", " ").split()).strip()

def _fold_text(s):
    if not s: return ""
    s = _norm(s)
    s = s.replace("I", "i").replace("\u0130", "i")
    s = s.lower()
    for old, new in [("\u015f","s"),("\u011f","g"),("\u00fc","u"),("\u00f6","o"),
                     ("\u00e7","c"),("\u0131","i")]:
        s = s.replace(old, new)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.split())

def _extract_league(href):
    from urllib.parse import unquote
    try:
        path = href.split("mackolik.com/", 1)[-1].strip("/")
        parts = [unquote(x) for x in path.split("/") if x]
        if len(parts) >= 3:
            return parts[2].replace("-", " ").title()
    except Exception:
        pass
    return ""

def _parse_header_bs4(html):
    data = {"mac_tarihi": "", "lig": ""}
    soup = BeautifulSoup(html, "html.parser")
    el = soup.select_one("span[class*='p0c-soccer-match-details-header__info-date']")
    if el:
        data["mac_tarihi"] = _norm(el.get_text())
    comp = soup.select_one("a[class*='p0c-soccer-match-details-header__competition-link']")
    if comp:
        txt = _norm(comp.get_text().splitlines()[0]) if comp.get_text() else ""
        if not txt:
            txt = _extract_league(comp.get("href", ""))
        data["lig"] = txt
    return data

def _bs4_opt_map(item_el):
    mp = {}
    for opt in item_el.select(".widget-iddaa-markets__option"):
        txt = opt.get_text(separator="|", strip=True)
        lines = [x.strip() for x in txt.split("|") if x.strip()]
        if not lines: continue
        for i in range(len(lines)-1, -1, -1):
            v = lines[i].replace(",", ".")
            if v.replace(".", "", 1).isdigit():
                label = _fold_text(" ".join(lines[:i]).strip())
                mp[label] = v
                break
    return mp

def _parse_markets_bs4(html, market_keys=None):
    result = {k: "" for k in [
        "ms_kodu","ms1","ms0","ms2","cs_1x","cs_12","cs_x2",
        "iy1","iy0","iy2","au_0_5_alt","au_0_5_ust","au_1_5_alt","au_1_5_ust",
        "au_2_5_alt","au_2_5_ust","au_3_5_alt","au_3_5_ust","au_4_5_alt","au_4_5_ust",
        "kg_var","kg_yok","hnd_1","hnd_x","hnd_2","hnd2_1","hnd2_x","hnd2_2",
        "iy_au_05_alt","iy_au_05_ust","iy_au_15_alt","iy_au_15_ust",
        "iy_ms_1_1","iy_ms_1_x","iy_ms_1_2","iy_ms_x_1","iy_ms_x_x","iy_ms_x_2",
        "iy_ms_2_1","iy_ms_2_x","iy_ms_2_2","tg_0_1","tg_2_3","tg_4_5","tg_6p",
        "t1_1_5_ust","t1_2_5_ust","t2_1_5_ust","t2_2_5_ust",
    ]}
    soup = BeautifulSoup(html, "html.parser")
    for ce in soup.select(".widget-iddaa-markets__iddaa-code"):
        ct = _norm(ce.get_text()).strip()
        if ct and ct.isdigit():
            result["ms_kodu"] = ct
            break
    def wants(*keys):
        return market_keys is None or any(k in market_keys for k in keys)
    items = soup.select("li.widget-iddaa-markets__market-item") or soup.select(".widget-iddaa-markets__market-item")
    for item in items:
        try:
            raw = item.get_text(separator="|", strip=True)
            parts = raw.split("|")
            first = _fold_text(parts[0] if parts else raw).replace(",", ".")
        except Exception:
            continue
        try:
            if first.startswith("mac sonucu"):
                if not result["ms_kodu"]:
                    ce2 = item.select_one(".widget-iddaa-markets__iddaa-code")
                    if ce2:
                        ct = _norm(ce2.get_text()).strip()
                        if ct and ct.isdigit(): result["ms_kodu"] = ct
                if wants("ms1","ms0","ms2"):
                    mp = _bs4_opt_map(item)
                    result["ms1"]=mp.get("1",""); result["ms0"]=mp.get("x",""); result["ms2"]=mp.get("2","")
            elif ("cifte sans" in first or "cifte sans" in first) and wants("cs_1x","cs_12","cs_x2"):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk=k.replace(" ","")
                    if "1x" in kk or "1-x" in kk: result["cs_1x"]=v
                    elif "12" in kk or "1-2" in kk: result["cs_12"]=v
                    elif "x2" in kk or "x-2" in kk or "2x" in kk: result["cs_x2"]=v
            elif ("ilk yari" in first or "1. yari" in first) and "sonucu" in first and "mac" not in first and "/" not in first and wants("iy1","iy0","iy2"):
                mp = _bs4_opt_map(item)
                result["iy1"]=mp.get("1",""); result["iy0"]=mp.get("x",""); result["iy2"]=mp.get("2","")
            elif "alt/ust" in first and "ilk yari" not in first and "1. yari" not in first and "takim" not in first and "ev sahibi" not in first and "deplasman" not in first and "konuk" not in first:
                for th,key in [("0.5","0_5"),("1.5","1_5"),("2.5","2_5"),("3.5","3_5"),("4.5","4_5")]:
                    if th in first and wants(f"au_{key}_alt",f"au_{key}_ust"):
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if "alt" in k: result[f"au_{key}_alt"]=v
                            elif "ust" in k: result[f"au_{key}_ust"]=v
                        break
            elif ("karsilikli" in first or ("iki takim" in first and "gol" in first)) and wants("kg_var","kg_yok"):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    if "var" in k or "evet" in k: result["kg_var"]=v
                    elif "yok" in k or "hayir" in k: result["kg_yok"]=v
            elif ("handikap" in first or "hnd" in first) and ("-1" in first or "0:1" in first) and wants("hnd_1","hnd_x","hnd_2"):
                mp = _bs4_opt_map(item)
                result["hnd_1"]=mp.get("1",""); result["hnd_x"]=mp.get("x",""); result["hnd_2"]=mp.get("2","")
            elif ("handikap" in first or "hnd" in first) and ("+1" in first or "1:0" in first) and wants("hnd2_1","hnd2_x","hnd2_2"):
                mp = _bs4_opt_map(item)
                result["hnd2_1"]=mp.get("1",""); result["hnd2_x"]=mp.get("x",""); result["hnd2_2"]=mp.get("2","")
            elif ("ilk yari" in first or "1. yari" in first) and ("alt" in first or "ust" in first) and "sonucu" not in first:
                for th,key in [("0.5","05"),("1.5","15")]:
                    if th in first and wants(f"iy_au_{key}_alt",f"iy_au_{key}_ust"):
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if "alt" in k: result[f"iy_au_{key}_alt"]=v
                            elif "ust" in k: result[f"iy_au_{key}_ust"]=v
                        break
            elif ("ilk yari" in first or "1. yari" in first) and "mac sonucu" in first:
                mp = _bs4_opt_map(item)
                combos = {"1/1":"iy_ms_1_1","1/x":"iy_ms_1_x","1/2":"iy_ms_1_2","x/1":"iy_ms_x_1","x/x":"iy_ms_x_x","x/2":"iy_ms_x_2","2/1":"iy_ms_2_1","2/x":"iy_ms_2_x","2/2":"iy_ms_2_2"}
                for k,v in mp.items():
                    nk=k.replace(" ","")
                    if nk in combos: result[combos[nk]]=v
            elif "toplam gol" in first and ("0-1" in first or "2-3" in first or raw.count("|")>=3):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk=k.replace(" ","")
                    if "0-1" in kk: result["tg_0_1"]=v
                    elif "2-3" in kk: result["tg_2_3"]=v
                    elif "4-5" in kk: result["tg_4_5"]=v
                    elif "6+" in kk or "6ve" in kk: result["tg_6p"]=v
            elif ("ev sahibi" in first or "takim 1" in first) and "alt/ust" in first and wants("t1_1_5_ust","t1_2_5_ust"):
                mp = _bs4_opt_map(item)
                if "1.5" in first: result["t1_1_5_ust"]=mp.get("ust","")
                elif "2.5" in first: result["t1_2_5_ust"]=mp.get("ust","")
            elif ("konuk" in first or "deplasman" in first or "takim 2" in first) and "alt/ust" in first and wants("t2_1_5_ust","t2_2_5_ust"):
                mp = _bs4_opt_map(item)
                if "1.5" in first: result["t2_1_5_ust"]=mp.get("ust","")
                elif "2.5" in first: result["t2_2_5_ust"]=mp.get("ust","")
        except Exception:
            continue
    return result

def scrape_match_fast(summary: dict, match_date=None, market_keys=None, max_retries=3) -> dict:
    row = dict(summary)
    if match_date:
        row["mac_tarihi"] = match_date.strftime("%d.%m.%Y")
    label = summary.get("ev_sahibi", "")
    urls_to_try = [summary["iddaa_link"]]
    overview = summary.get("overview_link", "")
    if overview and overview != summary["iddaa_link"]:
        urls_to_try.append(overview)
    for url in urls_to_try:
        for attempt in range(max_retries):
            try:
                time.sleep(_throttle_delay)
                resp = _SESSION.get(url, timeout=12, allow_redirects=True)
                if resp.status_code == 404: break
                if resp.status_code in (500, 502, 503, 429):
                    _throttle_hit()
                    if attempt < max_retries - 1:
                        time.sleep((attempt + 1) * 2)
                        continue
                    else: break
                resp.raise_for_status()
                _throttle_ok()
                html = resp.text
                hdr = _parse_header_bs4(html)
                if not row.get("mac_tarihi") and hdr.get("mac_tarihi"):
                    row["mac_tarihi"] = hdr["mac_tarihi"]
                if hdr.get("lig"): row["lig"] = hdr["lig"]
                api_kodu = row.get("ms_kodu", "")
                row.update(_parse_markets_bs4(html, market_keys=market_keys))
                if not row.get("ms_kodu") and api_kodu:
                    row["ms_kodu"] = api_kodu
                return row
            except requests.exceptions.HTTPError as e:
                status = getattr(getattr(e, "response", None), "status_code", 0)
                if status == 404: break
                if status in (500, 502, 503, 429): _throttle_hit()
                wait = (attempt + 1) * 2
                if attempt < max_retries - 1:
                    print(f"  [{label}] deneme {attempt+1} hata: {e} — {wait}s")
                    time.sleep(wait)
                else:
                    print(f"  [{label}] tum denemeler basarisiz: {e}")
            except Exception as e:
                wait = (attempt + 1) * 1
                if attempt < max_retries - 1:
                    print(f"  [{label}] deneme {attempt+1} hata: {e} — {wait}s")
                    time.sleep(wait)
                else:
                    print(f"  [{label}] tum denemeler basarisiz: {e}")
    return row

# ── Sabitler ────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).resolve().parent.parent
BASE_URL  = "https://www.mackolik.com"
API_URL   = "https://www.mackolik.com/perform/p0/ajax/components/competition/livescores/json?"

_SESSION = requests.Session()
_SESSION.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'tr,en-US;q=0.7,en;q=0.3',
    'Connection': 'keep-alive',
})

UEFA_COMP_KEYWORDS = [
    "sampiyonlar ligi", "champions league",
    "avrupa ligi", "europa league",
    "konferans ligi", "conference league",
    "uluslar ligi", "nations league",
]
UEFA_EXCLUDE = [
    "caf", "afc", "concacaf", "kadin", "women", "arjantin",
    "afrika", "u21", "u23", "u19", "u18", "u17",
    "kupasi", "cup",
]

EXCEL_HEADERS = [
    "Ev Sahibi", "Deplasman", "Tarih", "Saat", "Lig",
    "MS Kodu", "IY Skor", "MS Skor",
    "MS1", "MS0", "MS2",
    "CS 1X", "CS 12", "CS X2",
    "IY1", "IY0", "IY2",
    "AU 0.5 Alt", "AU 0.5 Ust",
    "AU 1.5 Alt", "AU 1.5 Ust",
    "AU 2.5 Alt", "AU 2.5 Ust",
    "AU 3.5 Alt", "AU 3.5 Ust",
    "AU 4.5 Alt", "AU 4.5 Ust",
    "KG Var", "KG Yok",
    "HND1", "HNDX", "HND2",
    "HND2-1", "HND2-X", "HND2-2",
    "IY AU 0.5 Alt", "IY AU 0.5 Ust",
    "IY AU 1.5 Alt", "IY AU 1.5 Ust",
    "IY/MS 1/1", "IY/MS 1/X", "IY/MS 1/2",
    "IY/MS X/1", "IY/MS X/X", "IY/MS X/2",
    "IY/MS 2/1", "IY/MS 2/X", "IY/MS 2/2",
    "TG 0-1", "TG 2-3", "TG 4-5", "TG 6+",
    "T1 1.5 Ust", "T1 2.5 Ust",
    "T2 1.5 Ust", "T2 2.5 Ust",
]

KEY_TO_COL = {
    "ev_sahibi": "Ev Sahibi", "konuk_ekip": "Deplasman",
    "mac_tarihi": "Tarih", "mac_saati": "Saat", "lig": "Lig",
    "ms_kodu": "MS Kodu", "ilk_yari_skor": "IY Skor", "mac_skoru": "MS Skor",
    "ms1": "MS1", "ms0": "MS0", "ms2": "MS2",
    "cs_1x": "CS 1X", "cs_12": "CS 12", "cs_x2": "CS X2",
    "iy1": "IY1", "iy0": "IY0", "iy2": "IY2",
    "au_0_5_alt": "AU 0.5 Alt", "au_0_5_ust": "AU 0.5 Ust",
    "au_1_5_alt": "AU 1.5 Alt", "au_1_5_ust": "AU 1.5 Ust",
    "au_2_5_alt": "AU 2.5 Alt", "au_2_5_ust": "AU 2.5 Ust",
    "au_3_5_alt": "AU 3.5 Alt", "au_3_5_ust": "AU 3.5 Ust",
    "au_4_5_alt": "AU 4.5 Alt", "au_4_5_ust": "AU 4.5 Ust",
    "kg_var": "KG Var", "kg_yok": "KG Yok",
    "hnd_1": "HND1", "hnd_x": "HNDX", "hnd_2": "HND2",
    "hnd2_1": "HND2-1", "hnd2_x": "HND2-X", "hnd2_2": "HND2-2",
    "iy_au_05_alt": "IY AU 0.5 Alt", "iy_au_05_ust": "IY AU 0.5 Ust",
    "iy_au_15_alt": "IY AU 1.5 Alt", "iy_au_15_ust": "IY AU 1.5 Ust",
    "iy_ms_1_1": "IY/MS 1/1", "iy_ms_1_x": "IY/MS 1/X", "iy_ms_1_2": "IY/MS 1/2",
    "iy_ms_x_1": "IY/MS X/1", "iy_ms_x_x": "IY/MS X/X", "iy_ms_x_2": "IY/MS X/2",
    "iy_ms_2_1": "IY/MS 2/1", "iy_ms_2_x": "IY/MS 2/X", "iy_ms_2_2": "IY/MS 2/2",
    "tg_0_1": "TG 0-1", "tg_2_3": "TG 2-3", "tg_4_5": "TG 4-5", "tg_6p": "TG 6+",
    "t1_1_5_ust": "T1 1.5 Ust", "t1_2_5_ust": "T1 2.5 Ust",
    "t2_1_5_ust": "T2 1.5 Ust", "t2_2_5_ust": "T2 2.5 Ust",
}
COL_TO_KEY = {v: k for k, v in KEY_TO_COL.items()}


def _fold(s: str) -> str:
    if not s:
        return ""
    s = s.replace("I", "i").replace("\u0130", "i")
    s = s.lower()
    for old, new in [("\u015f","s"),("\u011f","g"),("\u00fc","u"),("\u00f6","o"),
                     ("\u00e7","c"),("\u0131","i"),("\u015e","s"),("\u011e","g"),
                     ("\u00dc","u"),("\u00d6","o"),("\u00c7","c")]:
        s = s.replace(old, new)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.split())


def is_uefa(comp_name: str, area_name: str) -> bool:
    cn = _fold(comp_name)
    an = _fold(area_name)
    if an not in ("avrupa", "europe", "uefa"):
        return False
    if any(ex in cn for ex in UEFA_EXCLUDE):
        return False
    return any(kw in cn for kw in UEFA_COMP_KEYWORDS)


def fetch_day(target_date: dt.date, max_retries: int = 4) -> list[dict]:
    params = {"sports[]": "Soccer", "matchDate": target_date.strftime("%Y-%m-%d")}
    data = {}
    for attempt in range(max_retries):
        try:
            resp = _SESSION.get(API_URL, params=params, timeout=15)
            resp.raise_for_status()
            jr = resp.json()
            data = jr.get("data", {})
            if not data or (not data.get("matches") and jr.get("matches")):
                data = jr
            break
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep((attempt + 1) * 2)
            else:
                print(f"  [API] {target_date} basarisiz: {e}")
                return []

    matches_raw  = data.get("matches", {})
    competitions = data.get("competitions", {})
    if isinstance(matches_raw, list):
        matches_raw = {str(i): m for i, m in enumerate(matches_raw) if isinstance(m, dict)}
    if isinstance(competitions, list):
        competitions = {str(c.get("id", i)): c for i, c in enumerate(competitions) if isinstance(c, dict)}

    def _slug(name):
        s = name.lower()
        for old, new in [("\u0131","i"),("\u015f","s"),("\u011f","g"),("\u00e7","c"),
                         ("\u00fc","u"),("\u00f6","o"),(" ","-"),(".","-"),("'",""),("\"","")]:
            s = s.replace(old, new)
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        return re.sub(r'-{2,}', '-', s).strip('-')

    results = []
    for mid, m in matches_raw.items():
        iddaa_code = str(m.get("iddaaCode", "") or "")
        if not iddaa_code or iddaa_code == "None":
            continue
        home = m.get("homeTeam", {}).get("name", "")
        away = m.get("awayTeam", {}).get("name", "")
        if not home or not away:
            continue
        comp_id = m.get("competitionId")
        comp    = competitions.get(str(comp_id)) or competitions.get(comp_id) or {}
        area_obj  = comp.get("country") or comp.get("area") or {}
        area_name = area_obj.get("name", "")
        comp_name = comp.get("name", "")
        if not is_uefa(comp_name, area_name):
            continue

        score_h = m.get("score", {}).get("home", "")
        score_a = m.get("score", {}).get("away", "")
        mac_skoru = f"{score_h}-{score_a}" if score_h != "" and score_a != "" else ""
        try:
            ht = m.get("score", {}).get("ht", {})
            home_ht = str(ht.get('home', '')).strip() if isinstance(ht, dict) else ''
            away_ht = str(ht.get('away', '')).strip() if isinstance(ht, dict) else ''
            iy_skor = f"{home_ht}-{away_ht}" if (home_ht or away_ht) else ""
        except Exception:
            iy_skor = ""

        mst_utc    = m.get("mstUtc")
        mac_saati  = ""
        mac_tarihi = target_date.strftime("%d.%m.%Y")
        if mst_utc:
            try:
                utc_dt = dt.datetime.fromtimestamp(mst_utc / 1000, tz=dt.timezone.utc)
                ist_dt = utc_dt + dt.timedelta(hours=3)
                mac_saati  = ist_dt.strftime("%H:%M")
                mac_tarihi = ist_dt.strftime("%d.%m.%Y")
            except Exception:
                pass

        api_slug = m.get("slug") or m.get("url") or ""
        if api_slug and "/mac/" in api_slug:
            iddaa_link = urljoin(BASE_URL, api_slug.rstrip("/") + "/iddaa")
            if "/iddaa/iddaa" in iddaa_link:
                iddaa_link = iddaa_link.replace("/iddaa/iddaa", "/iddaa")
        else:
            iddaa_link = f"{BASE_URL}/mac/{_slug(home)}-vs-{_slug(away)}/iddaa/{mid}"

        results.append({
            "ev_sahibi": home, "konuk_ekip": away,
            "mac_saati": mac_saati, "mac_tarihi": mac_tarihi,
            "ilk_yari_skor": iy_skor, "mac_skoru": mac_skoru,
            "iddaa_link": iddaa_link, "lig": comp_name,
            "lig_key": _fold(area_name) + " " + _fold(comp_name),
            "ms_kodu": iddaa_code,
        })
    return results


def make_key(ev: str, dep: str, tarih: str) -> str:
    return f"{str(ev).strip().lower()}|{str(dep).strip().lower()}|{str(tarih).strip()}"


def save_excel(rows: list[dict], path: Path) -> None:
    """Yeni Excel olustur veya mevcuta ekle."""
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "UEFA Oranlari"
        header = EXCEL_HEADERS
        ws.append(header)
        hfill = PatternFill("solid", fgColor="1A5276")
        hfont = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

    for row_dict in rows:
        excel_row = []
        for col_name in header:
            key = COL_TO_KEY.get(col_name)
            val = row_dict.get(key, "") if key else ""
            excel_row.append(val if val != "" else None)
        ws.append(excel_row)

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 35)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_progress(progress_file: Path) -> dict:
    if progress_file.exists():
        try:
            return json.loads(progress_file.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_progress(progress_file: Path, data: dict) -> None:
    progress_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def date_range(start: dt.date, end: dt.date):
    cur = start
    while cur <= end:
        yield cur
        cur += dt.timedelta(days=1)


def main():
    parser = argparse.ArgumentParser(description="UEFA maclarini cek")
    parser.add_argument("--start",     default="2017-08-01")
    parser.add_argument("--end",       default="2026-03-31")
    parser.add_argument("--output",    default=None,  help="Cikti xlsx (varsayilan: output/iddaa_uefa_START_END.xlsx)")
    parser.add_argument("--progress",  default="progress_uefa.json")
    parser.add_argument("--workers",   type=int, default=4)
    parser.add_argument("--max-hours", type=float, default=5.5, help="Maksimum calisma suresi (saat)")
    parser.add_argument("--resume",    action="store_true")
    parser.add_argument("--dry-run",   action="store_true")
    args = parser.parse_args()

    start_date = dt.date.fromisoformat(args.start)
    end_date   = dt.date.fromisoformat(args.end)
    total_days = (end_date - start_date).days + 1

    # Cikti dosyasi
    if args.output:
        output_path = Path(args.output)
    else:
        slug = f"{args.start.replace('-','')}_{args.end.replace('-','')}"
        output_path = BASE_DIR / "output" / f"iddaa_uefa_{slug}.xlsx"

    progress_file = BASE_DIR / args.progress
    deadline = dt.datetime.now() + dt.timedelta(hours=args.max_hours)

    print("=" * 60)
    print("UEFA Mac Scraper")
    print(f"Aralik  : {start_date} - {end_date} ({total_days} gun)")
    print(f"Cikti   : {output_path}")
    print(f"Max sure: {args.max_hours} saat")
    print("=" * 60)

    # Progress yukle
    progress = load_progress(progress_file) if args.resume else {}
    last_done_date = progress.get("last_done_date")
    existing_keys  = set(progress.get("existing_keys", []))

    # Mevcut Excel'deki key'leri yukle (resume degilse veya bos ise)
    if not existing_keys and output_path.exists():
        print(f"[1] Mevcut Excel okunuyor: {output_path}")
        df = pd.read_excel(output_path)
        print(f"    {len(df):,} satir")
        for _, row in df.iterrows():
            existing_keys.add(make_key(
                row.get("Ev Sahibi", ""),
                row.get("Deplasman", ""),
                row.get("Tarih", ""),
            ))
        print(f"    {len(existing_keys):,} dedup key yuklendi")

    # Kaldigi yerden devam
    effective_start = start_date
    if last_done_date and args.resume:
        effective_start = dt.date.fromisoformat(last_done_date) + dt.timedelta(days=1)
        print(f"[RESUME] {last_done_date} tarihinden devam ediliyor")

    # Gunleri tara
    print(f"\n[2] Gunler taranıyor...")
    all_summaries: list[tuple[dt.date, dict]] = []
    found = 0
    skipped = 0
    batch_rows: list[dict] = []
    BATCH_SIZE = 200  # Her 200 maci kaydet

    start_time = dt.datetime.now()

    for i, day in enumerate(date_range(effective_start, end_date), 1):
        # Zaman kontrolu
        if dt.datetime.now() >= deadline:
            print(f"\n[!] Zaman doldu ({args.max_hours} saat). Kaydedilip cikiliyor...")
            break

        matches = fetch_day(day)
        new_for_day = []
        for s in matches:
            k = make_key(s.get("ev_sahibi",""), s.get("konuk_ekip",""), s.get("mac_tarihi", day.strftime("%d.%m.%Y")))
            if k in existing_keys:
                skipped += 1
                continue
            new_for_day.append((day, s))
            existing_keys.add(k)
            found += 1

        all_summaries.extend(new_for_day)

        if i % 30 == 0:
            elapsed = (dt.datetime.now() - start_time).seconds // 60
            print(f"  {i} gun -- {found} yeni mac, {skipped} var, {elapsed}dk gecti")

        time.sleep(0.15)

        # Batch olarak oran cek ve kaydet
        if len(all_summaries) >= BATCH_SIZE or (i == total_days and all_summaries):
            if not args.dry_run:
                _process_batch(all_summaries, output_path, args.workers)
            all_summaries = []
            # Progress kaydet
            progress["last_done_date"] = day.isoformat()
            progress["existing_keys"] = list(existing_keys)
            save_progress(progress_file, progress)

    # Kalan batch
    if all_summaries and not args.dry_run:
        _process_batch(all_summaries, output_path, args.workers)

    # Son progress kaydet
    progress["last_done_date"] = end_date.isoformat()
    progress["existing_keys"] = list(existing_keys)
    save_progress(progress_file, progress)

    print(f"\n{'='*60}")
    print(f"OZET:")
    print(f"  Bulunan UEFA maci : {found}")
    print(f"  Zaten var         : {skipped}")
    print(f"  Cikti             : {output_path}")
    print("=" * 60)


def _process_batch(summaries: list[tuple[dt.date, dict]], output_path: Path, workers: int) -> None:
    """Bir batch macin oranlarini cek ve Excel'e ekle."""
    print(f"  --> {len(summaries)} mac orani cekiliyor...")
    rows = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(scrape_match_fast, s, day): (day, s)
            for day, s in summaries
        }
        for future in as_completed(future_map):
            day, s = future_map[future]
            try:
                row = future.result()
                if row:
                    if not row.get("lig"):
                        row["lig"] = s.get("lig", "")
                    rows.append(row)
                else:
                    rows.append(dict(s))
            except Exception:
                rows.append(dict(s))

    save_excel(rows, output_path)
    print(f"  --> {len(rows)} satir kaydedildi: {output_path}")


if __name__ == "__main__":
    main()
