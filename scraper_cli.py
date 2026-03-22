#!/usr/bin/env python3
"""IddaaPro CLI - GitHub Actions icin headless scraper.

Kullanim:
    python scraper_cli.py --start 2024-08-01 --end 2025-03-15
    python scraper_cli.py --start 2024-08-01 --end 2025-03-15 --output data/2024.xlsx
"""
from __future__ import annotations

import argparse
import contextlib
import datetime as dt
import json
import os
import re
import sys
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import unquote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver import FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.support.ui import WebDriverWait

from urllib3.util.retry import Retry as _Retry
from requests.adapters import HTTPAdapter as _HTTPAdapter

# ── Sabitler ──────────────────────────────────────────────────────────────────
BASE_URL  = "https://www.mackolik.com"
LIVE_URL  = f"{BASE_URL}/futbol/canli-sonuclar"

DATE_TOGGLE_CLASS   = "widget-dateslider__datepicker-toggle"
DATE_YEAR_SELECTOR  = "widget-datepicker__selector--year"
DATE_MONTH_SELECTOR = "widget-datepicker__selector--month"
DATE_VALUE_CLASS    = "widget-datepicker__value"
DATE_NAV_NEXT       = "widget-datepicker__nav--next"
DATE_NAV_PREV       = "widget-datepicker__nav--previous"
DATE_CALENDAR_BODY  = "widget-datepicker__calendar-body"

MATCH_ROW_CSS = "div[class*='match-row--']"

POPUP_SELECTORS = [
    (By.XPATH,      "//button[@class='mobinterkapat']"),
    (By.CLASS_NAME, "rupclose"),
    (By.CLASS_NAME, "widget-gdpr-banner__accept"),
]

MONTH_ABBR = ["oca","sub","mar","nis","may","haz","tem","agu","eyl","eki","kas","ara"]

BASE_DIR = Path(__file__).resolve().parent

# ── HTTP Session ──────────────────────────────────────────────────────────────
_SESSION = requests.Session()
_SESSION.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'tr,en-US;q=0.7,en;q=0.3',
    'Connection': 'keep-alive',
})
_adapter = _HTTPAdapter(pool_connections=5, pool_maxsize=10, max_retries=0)
_SESSION.mount('https://', _adapter)
_SESSION.mount('http://', _adapter)
_SESSION_READY = False

def _init_session_cookies(driver):
    global _SESSION_READY
    try:
        for c in driver.get_cookies():
            _SESSION.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
        _SESSION_READY = True
    except Exception:
        pass

# ── API ───────────────────────────────────────────────────────────────────────
_API_URL = "https://www.mackolik.com/perform/p0/ajax/components/competition/livescores/json?"

_BANNED_KEYWORDS = {'women', 'u23', 'u21', 'u19', 'u18', 'u17', 'u20',
                    'reserve', 'youth', 'amateur', 'regional',
                    'non lig', 'non-league', 'non league',
                    'national league south', 'national league north',
                    'isthmian', 'southern league', 'northern league',
                    'a-lig kadin', 'a-league women', 'kadin',
                    'friendly', 'arkadas', 'hazirlik mac'}

_SPONSOR_WORDS = {'trendyol', 'spor', 'toto', 'stsl', 'misli', 'bilyoner',
                  'cemil', 'usta', 'turkcell', 'digiturk', 'bein',
                  'ptt', 'tff', 'nesine', 'bank', 'asya', 'sigorta',
                  # Ek sponsor / isim degisikligi kelimeleri (2020-2026)
                  'vodafone', 'ziraat', 'thy', 'betsson', 'macron',
                  'parions', 'uber', 'eats', 'ea', 'tim',
                  'scotiabank', 'apertura', 'clausura',
                  }

_COUNTRY_TOP_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'turkiye': ('super lig',),
    'ingiltere': ('premier lig',),
    'ispanya': ('laliga', 'la liga', 'primera division'),
    'italya': ('serie a',),
    'almanya': ('bundesliga',),
    'fransa': ('ligue 1',),
    'hollanda': ('eredivisie',),
    'portekiz': ('premier lig', 'primeira liga'),
    'belcika': ('pro lig', 'first division a'),
    'avusturya': ('bundesliga',),
    'cek cumhuriyeti': ('czech liga', 'first liga'),
    'danimarka': ('superliga',),
    'finlandiya': ('veikkausliiga',),
    'hirvatistan': ('1 hnl',),
    'iskocya': ('premiership',),
    'isvec': ('allsvenskan',),
    'isvicre': ('super lig', 'super league'),
    'macaristan': ('nb i',),
    'norvec': ('eliteserien',),
    'polonya': ('ekstraklasa',),
    'romanya': ('liga 1', 'liga i'),
    'rusya': ('premier lig',),
    'sirbistan': ('super lig',),
    'yunanistan': ('super lig',),
    'abd': ('mls', 'major league soccer'),
    'brezilya': ('serie a', 'brasileirao'),
    'japonya': ('j1 ligi', 'j1 lig'),
    'guney kore': ('k lig 1',),
    'cin': ('super lig',),
    'avustralya': ('a lig', 'aleague'),
}

_COUNTRY_SECOND_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'turkiye': ('1 lig',),
    'ingiltere': ('championship',),
    'ispanya': ('laliga 2', 'segunda division'),
    'italya': ('serie b',),
    'almanya': ('2 bundesliga',),
    'fransa': ('ligue 2',),
    'hollanda': ('eerste divisie',),
    'portekiz': ('2 lig', 'segunda liga'),
    'belcika': ('challenger pro lig', 'first division b'),
    'avusturya': ('1 lig', 'erste liga'),
    'danimarka': ('1 lig', '1 division'),
    'finlandiya': ('ykkosliiga', 'ykkonen'),
    'hirvatistan': ('2 hnl',),
    'iskocya': ('championship',),
    'isvicre': ('challenge lig', 'challenge league'),
    'macaristan': ('nb ii', '2 lig'),
    'norvec': ('1 lig', 'obos ligaen'),
    'polonya': ('1 lig', 'i liga'),
    'romanya': ('2 lig', 'liga 2', 'liga ii'),
    'rusya': ('fnl', 'first league'),
    'sirbistan': ('1 lig',),
    'yunanistan': ('2 lig', 'super league 2'),
}

_COUNTRY_THIRD_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'fransa': ('ulusal lig 1', 'national 1'),
    'ingiltere': ('league one', '1 lig', 'lig 1'),
    'polonya': ('2 lig', 'ii liga'),
    'iskocya': ('3 lig', 'league one'),
    'isvec': ('3 lig',),
}

_COUNTRY_FOURTH_TIER_HINTS: dict[str, tuple[str, ...]] = {
    'ingiltere': ('league two', '2 lig', 'lig 2'),
    'iskocya': ('4 lig', 'league two'),
}

_API_COUNTRY_TO_LEAGUE_LIST: dict[str, str] = {
    "turkey": "TÜRKİYE", "türkiye": "TÜRKİYE", "turkiye": "TÜRKİYE",
    "england": "İNGİLTERE", "ingiltere": "İNGİLTERE",
    "spain": "İSPANYA", "ispanya": "İSPANYA",
    "italy": "İTALYA", "italya": "İTALYA",
    "germany": "ALMANYA", "almanya": "ALMANYA",
    "france": "FRANSA", "fransa": "FRANSA",
    "netherlands": "HOLLANDA", "hollanda": "HOLLANDA",
    "portugal": "PORTEKİZ", "portekiz": "PORTEKİZ",
    "belgium": "BELÇİKA", "belcika": "BELÇİKA",
    "austria": "AVUSTURYA", "avusturya": "AVUSTURYA",
    "czech republic": "ÇEK CUMHURİYETİ", "czechia": "ÇEK CUMHURİYETİ",
    "denmark": "DANİMARKA", "danimarka": "DANİMARKA",
    "finland": "FİNLANDİYA", "finlandiya": "FİNLANDİYA",
    "croatia": "HIRVATİSTAN", "hirvatistan": "HIRVATİSTAN",
    "scotland": "İSKOÇYA", "iskocya": "İSKOÇYA",
    "sweden": "İSVEÇ", "isvec": "İSVEÇ",
    "switzerland": "İSVİÇRE", "isvicre": "İSVİÇRE",
    "hungary": "MACARİSTAN", "macaristan": "MACARİSTAN",
    "norway": "NORVEÇ", "norvec": "NORVEÇ",
    "poland": "POLONYA", "polonya": "POLONYA",
    "romania": "ROMANYA", "romanya": "ROMANYA",
    "russia": "RUSYA", "rusya": "RUSYA",
    "serbia": "SIRBİSTAN", "sirbistan": "SIRBİSTAN",
    "greece": "YUNANİSTAN", "yunanistan": "YUNANİSTAN",
    "usa": "ABD", "abd": "ABD", "united states": "ABD",
    "brazil": "BREZİLYA", "brezilya": "BREZİLYA",
    "japan": "JAPONYA", "japonya": "JAPONYA",
    "south korea": "GÜNEY KORE", "guney kore": "GÜNEY KORE",
    "china": "ÇİN", "cin": "ÇİN",
    "australia": "AVUSTRALYA", "avustralya": "AVUSTRALYA",
    # Uluslararası bölgeler
    "avrupa": "AVRUPA", "europe": "AVRUPA",
    "dunya": "DÜNYA", "dünya": "DÜNYA", "world": "DÜNYA",
    "guney amerika": "GÜNEY AMERİKA", "south america": "GÜNEY AMERİKA",
}

_ALLOWED_COUNTRIES: set[str] | None = None

def _normalize_country(s: str) -> str:
    s = s.replace("İ", "i").replace("ı", "i").replace("Ş", "s").replace("ş", "s")
    s = s.replace("Ğ", "g").replace("ğ", "g").replace("Ü", "u").replace("ü", "u")
    s = s.replace("Ö", "o").replace("ö", "o").replace("Ç", "c").replace("ç", "c")
    return s.lower().strip()

def _get_allowed_countries() -> set[str]:
    global _ALLOWED_COUNTRIES
    if _ALLOWED_COUNTRIES is None:
        raw = set(_API_COUNTRY_TO_LEAGUE_LIST.keys())
        normalized = {_normalize_country(k) for k in raw}
        _ALLOWED_COUNTRIES = raw | normalized
    return _ALLOWED_COUNTRIES

# Uluslararası turnuvalar — country Avrupa/Dünya olsa bile kabul et
_INTERNATIONAL_TOURNAMENTS = {
    'sampiyonlar ligi', 'champions league', 'uefa champions',
    'avrupa ligi', 'europa league', 'uefa europa',
    'konferans ligi', 'conference league', 'uefa conference',
    'uluslar ligi', 'nations league', 'uefa nations',
    'dunya kupasi', 'world cup', 'dünya kupası',
    'avrupa sampiyonasi', 'euro 20', 'uefa euro',
    'avrupa sampiyonasi elemeler', 'euro qualif',
    'dunya kupasi elemeler', 'world cup qualif',
    'kulüpler dünya kupasi', 'club world cup', 'klubler dunya kupasi',
    'super kupa', 'super cup', 'uefa super',
    'copa america', 'copa libertadores', 'libertadores',
    'copa sudamericana', 'sudamericana',
    'concacaf', 'gold cup', 'altin kupa',
    'afrika kupasi', 'africa cup', 'afcon',
    'asya kupasi', 'asian cup', 'afc champions',
}

def _is_international_tournament(comp_name: str) -> bool:
    name_low = _normalize_country(comp_name)
    return any(t in name_low for t in _INTERNATIONAL_TOURNAMENTS)

def _is_allowed_competition(comp: dict) -> bool:
    country_raw = comp.get('countryName') or ''
    if not country_raw and isinstance(comp.get('country'), dict):
        country_raw = comp['country'].get('name', '')
    elif not country_raw:
        country_raw = str(comp.get('country', ''))
    country = country_raw.strip()
    if not country:
        return False

    comp_name_raw = (comp.get('competitionName') or comp.get('competition') or
                     comp.get('name') or '')
    comp_name = _normalize_country(comp_name_raw)  # Türkçe karakter normalize + lower

    # Banned keywords kontrolü
    if any(bw in comp_name for bw in _BANNED_KEYWORDS):
        return False

    # Uluslararası turnuva ise her zaman kabul et
    if _is_international_tournament(comp_name):
        return True

    # Ülke bazlı kontrol
    c_low = _normalize_country(country)
    if c_low in _get_allowed_countries():
        return True

    return False

# ── Text normalization ────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return unicodedata.normalize("NFC", s).strip() if s else ""

def _fold_text(s: str | None) -> str:
    if not s:
        return ""
    s = _norm(s)
    s = s.replace("İ", "i").replace("I", "i")
    s = s.lower()
    s = (s.replace("ş", "s").replace("ğ", "g").replace("ü", "u")
          .replace("ö", "o").replace("ç", "c").replace("ı", "i"))
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    return " ".join(s.split())

_COUNTRY_ALIASES = {
    'turkey': 'turkiye', 'england': 'ingiltere', 'spain': 'ispanya',
    'italy': 'italya', 'germany': 'almanya', 'france': 'fransa',
    'netherlands': 'hollanda', 'portugal': 'portekiz', 'belgium': 'belcika',
    'austria': 'avusturya', 'czech republic': 'cek cumhuriyeti',
    'czechia': 'cek cumhuriyeti', 'denmark': 'danimarka',
    'finland': 'finlandiya', 'croatia': 'hirvatistan',
    'scotland': 'iskocya', 'sweden': 'isvec', 'switzerland': 'isvicre',
    'hungary': 'macaristan', 'norway': 'norvec', 'poland': 'polonya',
    'romania': 'romanya', 'russia': 'rusya', 'serbia': 'sirbistan',
    'greece': 'yunanistan', 'usa': 'abd', 'united states': 'abd',
    'brazil': 'brezilya', 'japan': 'japonya',
    'south korea': 'guney kore', 'china': 'cin',
    'australia': 'avustralya',
    # Uluslararası bölgeler
    'europe': 'avrupa', 'world': 'dunya',
    'south america': 'guney amerika',
    'north / central america': 'kuzey / orta amerika',
    'asia': 'asya', 'africa': 'afrika',
}

_LIG_ALIASES = {
    'league': 'lig', 'liga': 'lig', 'ligue': 'lig',
    'division': 'lig', 'divisie': 'lig',
    'primera': '1', 'segunda': '2',
    'first': '1', 'second': '2',
    'one': '1', 'two': '2', 'three': '3', 'four': '4',
    'national': 'ulusal',
}

def _fold_lig(s: str) -> str:
    r = _fold_text(s.replace(".", " ").replace("-", " "))
    for eng, tr in _COUNTRY_ALIASES.items():
        if r == eng:
            return tr
        if r.startswith(eng + " "):
            r = tr + r[len(eng):]
            break
    # Kelime bazinda alias ve sponsor temizligi (substring replace yerine)
    words = r.split()
    cleaned = []
    for w in words:
        w = _LIG_ALIASES.get(w, w)   # tam kelime eslesirse alias uygula
        if w not in _SPONSOR_WORDS:   # sponsor kelimesi degilse tut
            cleaned.append(w)
    return " ".join(cleaned) if cleaned else r

LEAGUE_LIST: list[tuple[str, list[str]]] = [
    ("TÜRKİYE",      ["Süper Lig", "1. Lig"]),
    ("İNGİLTERE",     ["Premier Lig", "Championship", "1. Lig", "2. Lig"]),
    ("İSPANYA",       ["LaLiga", "LaLiga 2"]),
    ("İTALYA",        ["Serie A", "Serie B"]),
    ("ALMANYA",       ["Bundesliga", "2. Bundesliga"]),
    ("FRANSA",        ["Ligue 1", "Ligue 2", "Ulusal Lig 1"]),
    ("HOLLANDA",      ["Eredivisie", "Eerste Divisie"]),
    ("PORTEKİZ",      ["Premier Lig", "2. Lig"]),
    ("BELÇİKA",       ["Pro Lig", "Challenger Pro Lig"]),
    ("AVUSTURYA",     ["Bundesliga", "1. Lig"]),
    ("ÇEK CUMHURİYETİ", ["Czech Liga"]),
    ("DANİMARKA",     ["Superliga", "1. Lig"]),
    ("FİNLANDİYA",    ["Veikkausliiga", "Ykkösliiga"]),
    ("HIRVATİSTAN",   ["1. HNL", "2. HNL"]),
    ("İSKOÇYA",       ["Premiership", "Championship", "3. Lig", "4. Lig"]),
    ("İSVEÇ",         ["Allsvenskan", "Superettan", "3. Lig"]),
    ("İSVİÇRE",       ["Süper Lig", "Challenge Lig"]),
    ("MACARİSTAN",    ["NB I", "2. Lig"]),
    ("NORVEÇ",        ["Eliteserien", "1. Lig"]),
    ("POLONYA",       ["Ekstraklasa", "1. Lig", "2. Lig"]),
    ("ROMANYA",       ["1. Lig", "2. Lig"]),
    ("RUSYA",         ["Premier Lig", "FNL"]),
    ("SIRBİSTAN",     ["Süper Lig", "1. Lig"]),
    ("YUNANİSTAN",    ["Süper Lig", "2. Lig"]),
    ("ABD",           ["MLS"]),
    ("BREZİLYA",      ["Serie A"]),
    ("JAPONYA",       ["J1 Ligi"]),
    ("GÜNEY KORE",    ["K Lig 1"]),
    ("ÇİN",           ["Süper Lig"]),
    ("AVUSTRALYA",    ["A-Lig"]),
    # Uluslararası turnuvalar
    ("AVRUPA",        ["Şampiyonlar Ligi", "Avrupa Ligi", "Konferans Ligi",
                       "Uluslar Ligi", "Avrupa Şampiyonası", "Avrupa Şampiyonası Elemeler",
                       "Süper Kupa", "UEFA Gençlik Ligi"]),
    ("DÜNYA",         ["Dünya Kupası", "Dünya Kupası Elemeler", "Kulüpler Dünya Kupası"]),
    ("GÜNEY AMERİKA", ["Libertadores Kupası", "Copa Sudamericana"]),
]

def lig_key(country: str, league: str) -> str:
    return _fold_lig(country) + " " + _fold_lig(league)

_ALL_DEFINED_LEAGUE_KEYS: set[str] | None = None
def _get_all_league_keys() -> set[str]:
    global _ALL_DEFINED_LEAGUE_KEYS
    if _ALL_DEFINED_LEAGUE_KEYS is None:
        _ALL_DEFINED_LEAGUE_KEYS = {lig_key(c, l) for c, ls in LEAGUE_LIST for l in ls}
    return _ALL_DEFINED_LEAGUE_KEYS

def _split_known_lig_key(value: str) -> tuple[str, str]:
    value = _fold_lig(value)
    for country, _ in LEAGUE_LIST:
        ckey = _fold_lig(country)
        prefix = ckey + " "
        if value == ckey:
            return ckey, ""
        if value.startswith(prefix):
            return ckey, value[len(prefix):].strip()
    return "", value

def _league_tier(country_key: str, league_text: str) -> int | None:
    league_norm = _fold_lig(league_text)
    league_core = " ".join(w for w in league_norm.split() if w not in _SPONSOR_WORDS)

    def has_any(*phrases: str) -> bool:
        return any(p in league_core for p in phrases)

    if has_any(*_COUNTRY_FOURTH_TIER_HINTS.get(country_key, ())):
        return 4
    if has_any(*_COUNTRY_THIRD_TIER_HINTS.get(country_key, ())):
        return 3
    if has_any(*_COUNTRY_SECOND_TIER_HINTS.get(country_key, ())):
        return 2
    if has_any(*_COUNTRY_TOP_TIER_HINTS.get(country_key, ())):
        return 1

    numeric_patterns = [
        (r'\b4\s*(?:lig|liga|division|divisie|hnl)\b', 4),
        (r'\b3\s*(?:lig|liga|division|divisie|hnl)\b', 3),
        (r'\b2\s*(?:lig|liga|division|divisie|bundesliga|hnl)\b', 2),
        (r'\b1\s*(?:lig|liga|division|divisie|hnl)\b', 1),
    ]
    for pattern, tier in numeric_patterns:
        if re.search(pattern, league_core):
            return tier
    return None

def _lig_components_match(found_key: str, selected_key: str) -> bool:
    found_country, found_league = _split_known_lig_key(found_key)
    sel_country, sel_league = _split_known_lig_key(selected_key)
    if not sel_country or not found_country:
        return False
    if found_country != sel_country:
        return False
    if not found_league or not sel_league:
        return found_country == sel_country
    found_tier = _league_tier(found_country, found_league)
    sel_tier = _league_tier(sel_country, sel_league)
    if found_tier is not None and sel_tier is not None:
        return found_tier == sel_tier
    return found_league == sel_league

def lig_filtreli_key(lig_key_val: str, sel_leagues: set[str] | None) -> bool:
    if sel_leagues is None:
        return True
    if not sel_leagues:
        return False
    # Uluslararası turnuvalar her zaman geçer
    if _is_international_tournament(lig_key_val):
        return True
    return any(_lig_components_match(lig_key_val, k) for k in sel_leagues)

# ── API fetch ─────────────────────────────────────────────────────────────────
def fetch_matches_api(target_date: dt.date, max_retries: int = 4,
                      lig_filtre: set | None = None) -> list[dict]:
    """Mackolik API: data.competitions + data.matches yapisi."""
    params = {'sports[]': 'Soccer', 'matchDate': target_date.strftime('%Y-%m-%d')}
    data = None
    for attempt in range(max_retries):
        try:
            if not _SESSION_READY:
                _SESSION.get("https://www.mackolik.com", timeout=8)
            resp = _SESSION.get(_API_URL, params=params, timeout=10)
            if resp.status_code in (502, 500, 503):
                time.sleep((attempt + 1) * 2)
                continue
            resp.raise_for_status()
            data = resp.json()
            break
        except Exception:
            if attempt < max_retries - 1:
                time.sleep((attempt + 1) * 2)
    if not data:
        return []

    api_data = data.get('data', data)
    if not isinstance(api_data, dict):
        return []

    # ── Competition dict ──
    comps_raw = api_data.get('competitions', {})
    if isinstance(comps_raw, list):
        comps_raw = {str(i): c for i, c in enumerate(comps_raw)}

    allowed_comps = {}  # comp_id -> (country, league, lig_key)
    for cid, comp in comps_raw.items():
        if not isinstance(comp, dict):
            continue
        if not _is_allowed_competition(comp):
            continue
        country_val = comp.get('country', '')
        if isinstance(country_val, dict):
            country = country_val.get('name', '')
        else:
            country = str(country_val)
        league_name = comp.get('name', '')
        lk = _fold_lig(country) + ' ' + _fold_lig(league_name)
        if lig_filtre and not lig_filtreli_key(lk, lig_filtre):
            continue
        allowed_comps[cid] = (country, league_name, lk)

    # ── Matches dict ──
    matches_raw = api_data.get('matches', {})
    if isinstance(matches_raw, list):
        match_list = matches_raw
    elif isinstance(matches_raw, dict):
        match_list = list(matches_raw.values())
    else:
        return []

    summaries = []
    for m in match_list:
        if not isinstance(m, dict):
            continue
        comp_id = str(m.get('competitionId', ''))
        if comp_id not in allowed_comps:
            continue

        iddaa_code = m.get('iddaaCode') or ''
        if not iddaa_code:
            continue

        # Takim isimleri — {name, slug} dict veya string
        home_obj = m.get('homeTeam', {})
        away_obj = m.get('awayTeam', {})
        home = home_obj.get('name', '') if isinstance(home_obj, dict) else str(home_obj)
        away = away_obj.get('name', '') if isinstance(away_obj, dict) else str(away_obj)
        if not home or not away:
            continue

        # URL olustur
        match_slug = m.get('matchSlug', '')
        comp_data = comps_raw.get(comp_id, {})
        comp_slug = comp_data.get('competitionSlug', '')
        country_slug = comp_data.get('countrySlug', '')
        season_slug = comp_data.get('seasonSlug', '')
        if match_slug and comp_slug and country_slug:
            full = f"{BASE_URL}/futbol/{country_slug}/{comp_slug}/{season_slug}/mac/{match_slug}/iddaa"
            overview = full.replace('/iddaa', '')
        else:
            full = ''
            overview = ''

        # Skor
        score = m.get('score', {}) or {}
        ft_home = score.get('home', '')
        ft_away = score.get('away', '')
        ft_score = f"{ft_home}-{ft_away}" if ft_home != '' and ft_away != '' else ''
        ht = score.get('ht', {}) or {}
        ht_home = ht.get('home', '')
        ht_away = ht.get('away', '')
        ht_score = f"{ht_home}-{ht_away}" if ht_home != '' and ht_away != '' else ''

        # Mac saati (mstUtc ms -> TR saati HH:MM)
        mac_saati = ''
        mst_utc = m.get('mstUtc', 0)
        if mst_utc:
            try:
                match_dt = dt.datetime.utcfromtimestamp(mst_utc / 1000) + dt.timedelta(hours=3)
                mac_saati = match_dt.strftime('%H:%M')
            except Exception:
                pass

        _, _, lig_key_val = allowed_comps[comp_id]

        summaries.append({
            'ev_sahibi': home, 'konuk_ekip': away,
            'ms_kodu': str(iddaa_code),
            'mac_saati': mac_saati,
            'ilk_yari_skor': ht_score,
            'mac_skoru': ft_score,
            'iddaa_link': full,
            'overview_link': overview,
            'lig_key': lig_key_val,
        })
    return summaries

# ── Selenium helpers ──────────────────────────────────────────────────────────
def build_driver():
    opts = FirefoxOptions()
    opts.add_argument("-headless")
    opts.page_load_strategy = "eager"
    # GitHub Actions: geckodriver sistem PATH'inde
    gd = BASE_DIR / "geckodriver.exe"
    if gd.exists():
        svc = FirefoxService(executable_path=str(gd))
        drv = webdriver.Firefox(service=svc, options=opts)
    else:
        drv = webdriver.Firefox(options=opts)
    drv.set_page_load_timeout(15)
    drv.implicitly_wait(0.2)
    return drv

def _js_click(driver, by, val):
    try:
        el = driver.find_element(by, val)
        driver.execute_script("arguments[0].click();", el)
    except Exception:
        pass

def close_popups(driver):
    for by, val in POPUP_SELECTORS:
        _js_click(driver, by, val)
        time.sleep(0.1)

def enable_football(driver):
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "input[data-type='iddaa']")
        if not btn.is_selected():
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.3)
    except Exception:
        pass

def wait_rows(driver, timeout=4):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, MATCH_ROW_CSS))
    except Exception:
        pass

def open_main_page(driver, max_retries=3):
    for i in range(max_retries):
        try:
            driver.get(LIVE_URL)
            wait_rows(driver, timeout=5)
            close_popups(driver)
            enable_football(driver)
            return
        except Exception:
            if i < max_retries - 1:
                time.sleep(1)

def _move_year(driver, target_year: int):
    for _ in range(25):
        try:
            el = driver.find_element(By.CLASS_NAME, DATE_YEAR_SELECTOR)
            txt = el.text.strip()
            m = re.search(r'(\d{4})', txt)
            if m and int(m.group(1)) == target_year:
                return
            nav = DATE_NAV_PREV if int(m.group(1)) > target_year else DATE_NAV_NEXT
            btn = el.find_element(By.XPATH, f".//*[contains(@class,'{nav}')]")
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.15)
        except Exception:
            time.sleep(0.1)

def _move_month(driver, target_month: int):
    target_abbr = MONTH_ABBR[target_month - 1]
    for _ in range(15):
        try:
            el = driver.find_element(By.CLASS_NAME, DATE_MONTH_SELECTOR)
            txt = _fold_text(el.text)
            if target_abbr in txt:
                return
            current_idx = next((i for i, a in enumerate(MONTH_ABBR) if a in txt), -1)
            nav = DATE_NAV_PREV if current_idx > target_month - 1 else DATE_NAV_NEXT
            btn = el.find_element(By.XPATH, f".//*[contains(@class,'{nav}')]")
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.15)
        except Exception:
            time.sleep(0.1)

def _pick_day(driver, target_day: int):
    try:
        body = driver.find_element(By.CLASS_NAME, DATE_CALENDAR_BODY)
        cells = body.find_elements(By.TAG_NAME, "td")
        for c in cells:
            cls = c.get_attribute("class") or ""
            if "not-month-day" in cls:
                continue
            if c.text.strip() == str(target_day):
                driver.execute_script("arguments[0].click();", c)
                return
    except Exception:
        pass

def pick_date(driver, date: dt.date):
    try:
        toggle = WebDriverWait(driver, 5).until(
            lambda d: d.find_element(By.CLASS_NAME, DATE_TOGGLE_CLASS))
        driver.execute_script("arguments[0].click();", toggle)
        time.sleep(0.3)
    except Exception:
        pass
    WebDriverWait(driver, 5).until(
        lambda d: d.find_element(By.CLASS_NAME, DATE_VALUE_CLASS))
    _move_year(driver, date.year)
    _move_month(driver, date.month)
    _pick_day(driver, date.day)
    for _ in range(15):
        try:
            slider = driver.find_element(By.CLASS_NAME, "widget-dateslider__date--active")
            stxt = slider.text.strip()
            if str(date.day) in stxt:
                break
        except Exception:
            pass
        time.sleep(0.1)
    time.sleep(0.25)

def collect_summaries(driver) -> list[dict]:
    JS_COLLECT = """
    var results = [];
    var seen = {};
    var currentLig = '';
    var rows = document.querySelectorAll("div[class*='match-row--']");
    for (var i = 0; i < rows.length; i++) {
        var row = rows[i];
        var links = row.querySelectorAll('a[href]');
        for (var li = 0; li < links.length; li++) {
            var href = links[li].getAttribute('href') || '';
            if (href.indexOf('/puan-durumu/') > -1) {
                try {
                    var p = href.split('mackolik.com/')[1].replace(/^\\//, '');
                    var segs = p.split('/').filter(Boolean);
                    if (segs.length >= 2) currentLig = decodeURIComponent(segs[1]);
                } catch(e) {}
                break;
            } else if (href.indexOf('/futbol/') > -1 && href.indexOf('/mac/') === -1) {
                try {
                    var p2 = href.split('mackolik.com/')[1].replace(/^\\//, '');
                    var segs2 = p2.split('/').filter(Boolean);
                    if (segs2.length >= 3) currentLig = decodeURIComponent(segs2[1]) + '/' + decodeURIComponent(segs2[2]);
                    else if (segs2.length >= 2) currentLig = decodeURIComponent(segs2[1]);
                } catch(e) {}
                break;
            }
        }
        var homeEl = row.querySelector('.match-row__team-name--home');
        var awayEl = row.querySelector('.match-row__team-name--away');
        if (!homeEl || !awayEl) continue;
        var home = (homeEl.textContent || '').trim();
        var away = (awayEl.textContent || '').trim();
        if (!home || !away) continue;
        var macHref = '';
        for (var ai = 0; ai < links.length; ai++) {
            var ah = links[ai].getAttribute('href') || '';
            if (ah.indexOf('/mac/') > -1) { macHref = ah; break; }
        }
        if (!macHref || seen[macHref]) continue;
        seen[macHref] = true;
        var timeEl = row.querySelector('.match-row__start-time');
        var htEl = row.querySelector('.match-row__half-time-score') || row.querySelector('[class*="half-time"]');
        var ftEl = row.querySelector('.match-row__score');
        results.push({
            ev_sahibi: home, konuk_ekip: away,
            mac_saati: timeEl ? timeEl.textContent.trim() : '',
            ilk_yari_skor: htEl ? htEl.textContent.replace(/IY|iy/g,'').trim() : '',
            mac_skoru: ftEl ? ftEl.textContent.trim() : '',
            mac_href: macHref, lig_raw: currentLig
        });
    }
    return results;
    """
    try:
        raw_list = driver.execute_script(JS_COLLECT)
    except Exception:
        return []

    summaries = []
    for item in (raw_list or []):
        href = item.get('mac_href', '')
        full = urljoin(BASE_URL, href)
        if '/iddaa/' not in full:
            idx = full.rfind('/')
            full = full[:idx] + '/iddaa' + full[idx:]
        full = full.replace('karsilastirma/', '')

        lig_raw = item.get('lig_raw', '')
        if '/' in lig_raw:
            parts = lig_raw.split('/')
            blok_lig_key = _fold_lig(parts[0]) + ' ' + _fold_lig(parts[1])
        else:
            blok_lig_key = _fold_lig(lig_raw)

        summaries.append({
            'ev_sahibi':     item.get('ev_sahibi', ''),
            'konuk_ekip':    item.get('konuk_ekip', ''),
            'mac_saati':     item.get('mac_saati', ''),
            'ilk_yari_skor': item.get('ilk_yari_skor', ''),
            'mac_skoru':     item.get('mac_skoru', ''),
            'iddaa_link':    full,
            'overview_link': full.replace('/iddaa', ''),
            'lig_key':       blok_lig_key,
        })
    return summaries

# ── Market parsing ────────────────────────────────────────────────────────────
def _bs4_opt_map(item_el) -> dict:
    result = {}
    opts = item_el.select('.widget-iddaa-markets__option')
    for opt in opts:
        txt = _norm(opt.get_text(separator='|', strip=True))
        parts = txt.split('|')
        if len(parts) >= 2:
            label = _fold_text(parts[0])
            val = parts[-1].strip().replace(',', '.')
            result[label] = val
    return result

def _parse_header_bs4(html: str) -> dict:
    result = {}
    soup = BeautifulSoup(html, 'html.parser')
    date_el = soup.select_one('span[class*="p0c-soccer-match-details-header__info-date"]')
    if date_el:
        result['mac_tarihi'] = _norm(date_el.get_text()).strip()
    comp_el = soup.select_one('a[class*="p0c-soccer-match-details-header__competition-link"]')
    if comp_el:
        txt = _norm(comp_el.get_text()).strip()
        if txt:
            result['lig'] = txt
    return result

def _parse_markets_bs4(html: str, market_keys: set | None = None) -> dict:
    result = {
        'ms_kodu':'', 'ms1':'', 'ms0':'', 'ms2':'',
        'cs_1x':'', 'cs_12':'', 'cs_x2':'',
        'iy1':'', 'iy0':'', 'iy2':'',
        'au_0_5_alt':'', 'au_0_5_ust':'',
        'au_1_5_alt':'', 'au_1_5_ust':'',
        'au_2_5_alt':'', 'au_2_5_ust':'',
        'au_3_5_alt':'', 'au_3_5_ust':'',
        'au_4_5_alt':'', 'au_4_5_ust':'',
        'kg_var':'', 'kg_yok':'',
        'hnd_1':'', 'hnd_x':'', 'hnd_2':'',
        'hnd2_1':'', 'hnd2_x':'', 'hnd2_2':'',
        'iy_au_05_alt':'', 'iy_au_05_ust':'',
        'iy_au_15_alt':'', 'iy_au_15_ust':'',
        'iy_ms_1_1':'', 'iy_ms_1_x':'', 'iy_ms_1_2':'',
        'iy_ms_x_1':'', 'iy_ms_x_x':'', 'iy_ms_x_2':'',
        'iy_ms_2_1':'', 'iy_ms_2_x':'', 'iy_ms_2_2':'',
        'tg_0_1':'', 'tg_2_3':'', 'tg_4_5':'', 'tg_6p':'',
        't1_1_5_ust':'', 't1_2_5_ust':'',
        't2_1_5_ust':'', 't2_2_5_ust':'',
    }
    soup = BeautifulSoup(html, 'html.parser')
    code_els = soup.select('.widget-iddaa-markets__iddaa-code')
    for ce in code_els:
        code_txt = _norm(ce.get_text()).strip()
        if code_txt and code_txt.isdigit():
            result['ms_kodu'] = code_txt
            break
    def wants(*keys):
        return market_keys is None or any(k in market_keys for k in keys)
    items = soup.select('li.widget-iddaa-markets__market-item')
    if not items:
        items = soup.select('.widget-iddaa-markets__market-item')
    for item in items:
        try:
            raw = item.get_text(separator='|', strip=True)
            parts = raw.split('|')
            first = _fold_text(parts[0] if parts else raw).replace(',', '.')
        except Exception:
            continue
        try:
            if first.startswith('mac sonucu'):
                if not result['ms_kodu']:
                    code_el = item.select_one('.widget-iddaa-markets__iddaa-code')
                    if code_el:
                        ct = _norm(code_el.get_text()).strip()
                        if ct and ct.isdigit():
                            result['ms_kodu'] = ct
                if not result['ms_kodu']:
                    m = re.search(r'(\d+)', first)
                    result['ms_kodu'] = m.group(1) if m else ''
                if wants('ms1','ms0','ms2'):
                    mp = _bs4_opt_map(item)
                    result['ms1']=mp.get('1',''); result['ms0']=mp.get('x',''); result['ms2']=mp.get('2','')
            elif ('cifte sans' in first or 'cifte sans' in first) and wants('cs_1x','cs_12','cs_x2'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk=k.replace(' ','')
                    if '1x' in kk or '1-x' in kk: result['cs_1x']=v
                    elif '12' in kk or '1-2' in kk: result['cs_12']=v
                    elif 'x2' in kk or 'x-2' in kk or '2x' in kk: result['cs_x2']=v
            elif (('ilk yari' in first or '1. yari' in first) and 'sonucu' in first and 'mac' not in first) and wants('iy1','iy0','iy2'):
                mp = _bs4_opt_map(item)
                result['iy1']=mp.get('1',''); result['iy0']=mp.get('x',''); result['iy2']=mp.get('2','')
            elif 'alt' in first and 'ust' in first and 'ilk yari' not in first and '1. yari' not in first and 'takim' not in first and 'ev sahibi' not in first and 'deplasman' not in first and 'konuk' not in first:
                for th,key in [('0.5','0_5'),('1.5','1_5'),('2.5','2_5'),('3.5','3_5'),('4.5','4_5')]:
                    if th in first and wants(f'au_{key}_alt',f'au_{key}_ust'):
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if 'alt' in k: result[f'au_{key}_alt']=v
                            elif 'ust' in k: result[f'au_{key}_ust']=v
                        break
            elif ('karsilikli' in first or ('iki takim' in first and 'gol' in first)) and wants('kg_var','kg_yok'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    if 'var' in k or 'evet' in k: result['kg_var']=v
                    elif 'yok' in k or 'hayir' in k: result['kg_yok']=v
            elif ('handikap' in first or 'hnd' in first) and ('-1' in first or '0:1' in first) and wants('hnd_1','hnd_x','hnd_2'):
                mp = _bs4_opt_map(item)
                result['hnd_1']=mp.get('1',''); result['hnd_x']=mp.get('x',''); result['hnd_2']=mp.get('2','')
            # ── 2. Handikap ──
            elif ('handikap' in first or 'hnd' in first) and ('0:2' in first or '-2' in first) and wants('hnd2_1','hnd2_x','hnd2_2'):
                mp = _bs4_opt_map(item)
                result['hnd2_1']=mp.get('1',''); result['hnd2_x']=mp.get('x',''); result['hnd2_2']=mp.get('2','')
            # ── IY Alt/Üst ──
            elif ('ilk yari' in first or '1. yari' in first) and ('alt' in first or 'ust' in first) and wants('iy_au_05_alt','iy_au_05_ust','iy_au_15_alt','iy_au_15_ust'):
                for th,key in [('0.5','05'),('1.5','15')]:
                    if th in first:
                        mp = _bs4_opt_map(item)
                        for k,v in mp.items():
                            if 'alt' in k: result[f'iy_au_{key}_alt']=v
                            elif 'ust' in k: result[f'iy_au_{key}_ust']=v
                        break
            # ── IY/MS ──
            elif ('ilk yari' in first or '1. yari' in first) and ('mac sonucu' in first or 'iy/ms' in first or 'iy ms' in first) and wants('iy_ms_1_1','iy_ms_1_x','iy_ms_1_2','iy_ms_x_1','iy_ms_x_x','iy_ms_x_2','iy_ms_2_1','iy_ms_2_x','iy_ms_2_2'):
                combos = {'1/1':'iy_ms_1_1','1/x':'iy_ms_1_x','1/2':'iy_ms_1_2',
                           'x/1':'iy_ms_x_1','x/x':'iy_ms_x_x','x/2':'iy_ms_x_2',
                           '2/1':'iy_ms_2_1','2/x':'iy_ms_2_x','2/2':'iy_ms_2_2'}
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kn = k.replace(' ','').replace('-','/')
                    if kn in combos: result[combos[kn]]=v
            # ── Toplam Gol ──
            elif 'toplam gol' in first and ('0-1' in first or '2-3' in first or raw.count('|')>=3) and wants('tg_0_1','tg_2_3','tg_4_5','tg_6p'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    kk = k.replace(' ','')
                    if '0-1' in kk: result['tg_0_1']=v
                    elif '2-3' in kk: result['tg_2_3']=v
                    elif '4-5' in kk: result['tg_4_5']=v
                    elif '6' in kk: result['tg_6p']=v
            # ── Ev Sahibi Toplam Gol ──
            elif ('ev sahibi' in first or 'takim 1' in first) and ('1.5' in first or '2.5' in first) and wants('t1_1_5_ust','t1_2_5_ust'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    if 'ust' in k:
                        if '1.5' in first: result['t1_1_5_ust']=v
                        elif '2.5' in first: result['t1_2_5_ust']=v
            # ── Deplasman Toplam Gol ──
            elif ('deplasman' in first or 'konuk' in first or 'takim 2' in first) and ('1.5' in first or '2.5' in first) and wants('t2_1_5_ust','t2_2_5_ust'):
                mp = _bs4_opt_map(item)
                for k,v in mp.items():
                    if 'ust' in k:
                        if '1.5' in first: result['t2_1_5_ust']=v
                        elif '2.5' in first: result['t2_2_5_ust']=v
        except Exception:
            continue
    return result

# ── Throttle ──────────────────────────────────────────────────────────────────
_throttle_delay = 0.02
_throttle_lock = threading.Lock()

def _throttle_hit():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = min(_throttle_delay * 1.3, 0.5)

def _throttle_ok():
    global _throttle_delay
    with _throttle_lock:
        _throttle_delay = max(_throttle_delay * 0.5, 0.02)

# ── Row validation ────────────────────────────────────────────────────────────
def _row_is_valid(r: dict) -> bool:
    iy = (r.get('ilk_yari_skor') or '').strip()
    ms = (r.get('mac_skoru') or '').strip()
    if not iy or not ms:
        return False
    has_odds = any(r.get(k) for k in ('ms1', 'ms0', 'ms2'))
    return has_odds

# ── Match detail scraper ─────────────────────────────────────────────────────
def scrape_match_fast(summary: dict, match_date=None, market_keys=None,
                      max_retries: int = 3) -> dict:
    row = dict(summary)
    if match_date:
        row['mac_tarihi'] = match_date.strftime('%d.%m.%Y')
    label = summary.get('ev_sahibi', '')
    row['_http_failed'] = False

    urls_to_try = [summary['iddaa_link']]
    overview = summary.get('overview_link', '')
    if overview and overview != summary['iddaa_link']:
        urls_to_try.append(overview)

    for url in urls_to_try:
        if not url:
            continue
        for attempt in range(max_retries):
            try:
                time.sleep(_throttle_delay)
                resp = _SESSION.get(url, timeout=12, allow_redirects=False)
                # Redirect varsa /iddaa/ dusebilir, manual takip et
                if resp.status_code in (301, 302, 303, 307, 308):
                    loc = resp.headers.get('Location', '')
                    if loc and '/iddaa/' not in loc and '/iddaa/' in url:
                        # Redirect iddaa'yi kaldiriyor, bu URL oran vermez
                        row['_http_failed'] = True
                        return row
                    resp = _SESSION.get(loc if loc.startswith('http') else f'https://www.mackolik.com{loc}',
                                       timeout=12, allow_redirects=True)
                if resp.status_code == 404:
                    break
                if resp.status_code in (500, 502, 503, 429):
                    _throttle_hit()
                    if attempt < max_retries - 1:
                        time.sleep((attempt + 1) * 2)
                        continue
                    else:
                        break
                resp.raise_for_status()
                _throttle_ok()
                html = resp.text
                # Iddaa widget yoksa bu sayfa oran vermiyor
                if 'widget-iddaa-markets' not in html:
                    row['_http_failed'] = True
                    return row
                hdr = _parse_header_bs4(html)
                if not row.get('mac_tarihi') and hdr.get('mac_tarihi'):
                    row['mac_tarihi'] = hdr['mac_tarihi']
                if hdr.get('lig'):
                    row['lig'] = hdr['lig']
                api_kodu = row.get('ms_kodu', '')
                row.update(_parse_markets_bs4(html, market_keys=market_keys))
                if not row.get('ms_kodu') and api_kodu:
                    row['ms_kodu'] = api_kodu
                return row
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep((attempt + 1) * 1)
                else:
                    pass
    row['_http_failed'] = True
    return row


def scrape_match_selenium(driver, url: str, market_keys=None) -> dict:
    """Selenium ile mac detay sayfasina gidip oranlari cek."""
    result = {}
    try:
        driver.get(url)
        time.sleep(2)
        # Iddaa tab disabled mi kontrol et
        html = driver.page_source
        if 'widget-iddaa-markets' not in html:
            return result
        hdr = _parse_header_bs4(html)
        result.update(hdr)
        result.update(_parse_markets_bs4(html, market_keys=market_keys))
    except Exception as e:
        pass
    return result

# ── Excel export ──────────────────────────────────────────────────────────────
KEYS = [
    "ev_sahibi", "konuk_ekip", "mac_tarihi", "mac_saati", "lig",
    "ms_kodu", "ilk_yari_skor", "mac_skoru",
    "ms1", "ms0", "ms2",
    "cs_1x", "cs_12", "cs_x2",
    "iy1", "iy0", "iy2",
    "au_0_5_alt", "au_0_5_ust",
    "au_1_5_alt", "au_1_5_ust",
    "au_2_5_alt", "au_2_5_ust",
    "au_3_5_alt", "au_3_5_ust",
    "au_4_5_alt", "au_4_5_ust",
    "kg_var", "kg_yok",
    "hnd_1", "hnd_x", "hnd_2",
    "hnd2_1", "hnd2_x", "hnd2_2",
    "iy_au_05_alt", "iy_au_05_ust",
    "iy_au_15_alt", "iy_au_15_ust",
    "iy_ms_1_1", "iy_ms_1_x", "iy_ms_1_2",
    "iy_ms_x_1", "iy_ms_x_x", "iy_ms_x_2",
    "iy_ms_2_1", "iy_ms_2_x", "iy_ms_2_2",
    "tg_0_1", "tg_2_3", "tg_4_5", "tg_6p",
    "t1_1_5_ust", "t1_2_5_ust",
    "t2_1_5_ust", "t2_2_5_ust",
]

HEADERS = [
    "Ev Sahibi", "Konuk Ekip", "Tarih", "Saat", "Lig",
    "MS Kodu", "IY Skor", "MS Skor",
    "MS1", "MS0", "MS2",
    "CS 1X", "CS 12", "CS X2",
    "IY1", "IY0", "IY2",
    "AU 0.5 Alt", "AU 0.5 Ust",
    "AU 1.5 Alt", "AU 1.5 Ust",
    "AU 2.5 Alt", "AU 2.5 Ust",
    "AU 3.5 Alt", "AU 3.5 Ust",
    "AU 4.5 Alt", "AU 4.5 Ust",
    "KG Var", "KG Yok",
    "HND1", "HNDX", "HND2",
    "HND2-1", "HND2-X", "HND2-2",
    "IY AU 0.5 Alt", "IY AU 0.5 Ust",
    "IY AU 1.5 Alt", "IY AU 1.5 Ust",
    "IY/MS 1/1", "IY/MS 1/X", "IY/MS 1/2",
    "IY/MS X/1", "IY/MS X/X", "IY/MS X/2",
    "IY/MS 2/1", "IY/MS 2/X", "IY/MS 2/2",
    "TG 0-1", "TG 2-3", "TG 4-5", "TG 6+",
    "T1 1.5 Ust", "T1 2.5 Ust",
    "T2 1.5 Ust", "T2 2.5 Ust",
]

_BASE_KEYS    = ["ev_sahibi","konuk_ekip","mac_tarihi","mac_saati","lig",
                  "ms_kodu","ilk_yari_skor","mac_skoru"]
_BASE_HEADERS = ["Ev Sahibi","Konuk Ekip","Tarih","Saat","Lig",
                  "MS Kodu","IY Skor","MS Skor"]

def export_excel(rows: list[dict], path: Path, market_keys: set[str] | None = None) -> None:
    mkt_keys    = KEYS[8:]
    mkt_headers = HEADERS[8:]
    if market_keys is not None:
        eff_keys  = _BASE_KEYS    + [k for k in mkt_keys    if k in market_keys]
        eff_hdrs  = _BASE_HEADERS + [h for k, h in zip(mkt_keys, mkt_headers) if k in market_keys]
    else:
        eff_keys, eff_hdrs = KEYS, HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = "MS Oranlari"
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(eff_hdrs)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(k, "") for k in eff_keys])
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
    wb.save(path)

# ── Progress save/resume ─────────────────────────────────────────────────────
PROGRESS_FILE = BASE_DIR / "progress.json"

def save_progress(last_date: dt.date, end_date: dt.date, total: int, output_file: str):
    data = {
        'last_date': last_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_matches': total,
        'output_file': output_file,
        'timestamp': dt.datetime.now().isoformat(),
    }
    PROGRESS_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')

def load_progress() -> dict | None:
    if PROGRESS_FILE.exists():
        try:
            return json.loads(PROGRESS_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return None

# ── Main scraper ──────────────────────────────────────────────────────────────
def run_scraper(start: dt.date, end: dt.date, output_path: str,
                max_hours: float = 5.5):
    """Ana scraper loop — GitHub Actions icin 5.5 saat limiti var."""
    print(f"Scraper basliyor: {start:%d.%m.%Y} - {end:%d.%m.%Y}", flush=True)
    print(f"Cikti: {output_path}", flush=True)

    deadline = time.time() + max_hours * 3600
    lig_filtre = _get_all_league_keys()
    rows = []
    total = 0
    driver = None

    # Onceki partial Excel varsa yukle
    output = Path(output_path)
    if output.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output, read_only=True)
            ws = wb.active
            headers_row = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = {KEYS[i]: (row[i] if i < len(row) else '') for i in range(min(len(KEYS), len(row)))}
                rows.append(r)
            wb.close()
            total = len(rows)
            print(f"Onceki dosyadan {total} mac yuklendi", flush=True)
        except Exception as e:
            print(f"Onceki dosya okunamadi: {e}", flush=True)

    def _ensure_driver():
        nonlocal driver
        if driver is None:
            driver = build_driver()
            open_main_page(driver)
            _init_session_cookies(driver)
        return driver

    def _collect_day(cur_date):
        nonlocal driver
        summaries = []
        api_ok = False
        try:
            api_summaries = fetch_matches_api(cur_date, lig_filtre=lig_filtre)
            if api_summaries:
                summaries = api_summaries
                api_ok = True
                print(f'  [{cur_date:%d.%m.%Y}] {len(summaries)} mac (API)', flush=True)
        except Exception as e:
            print(f'  [{cur_date:%d.%m.%Y}] API hatasi: {e}', flush=True)

        if not api_ok:
            for retry in range(2):
                try:
                    d = _ensure_driver()
                    pick_date(d, cur_date)
                    wait_rows(d, timeout=5)
                    close_popups(d)
                    time.sleep(0.15)
                    summaries = collect_summaries(d)
                    print(f'  [{cur_date:%d.%m.%Y}] {len(summaries)} mac (Selenium)', flush=True)
                    break
                except Exception as e:
                    print(f'  [{cur_date:%d.%m.%Y}] Selenium hata {retry+1}: {e}', flush=True)
                    if retry == 0:
                        with contextlib.suppress(Exception):
                            if driver:
                                driver.quit()
                        driver = None
        return summaries

    cur = start
    day_count = 0
    total_days = (end - start).days + 1

    try:
        while cur <= end:
            # Zaman limiti kontrolu
            if time.time() > deadline:
                print(f"\n⏰ Zaman limiti ({max_hours}h) doldu. Son tarih: {cur:%d.%m.%Y}", flush=True)
                break

            day_count += 1
            print(f'[{day_count}/{total_days}] {cur:%d.%m.%Y}', flush=True)

            summaries = _collect_day(cur)

            # Lig filtreleme
            if lig_filtre:
                onceki = len(summaries)
                filtered = []
                for s in summaries:
                    lk = (s.get('lig_key') or '').strip()
                    if lk and lig_filtreli_key(lk, lig_filtre):
                        filtered.append(s)
                summaries = filtered
                if onceki != len(summaries):
                    print(f'  Lig filtre: {onceki} -> {len(summaries)}', flush=True)

            if not summaries:
                cur += dt.timedelta(days=1)
                continue

            # Paralel detay cekme (HTTP)
            done_count = 0
            day_matches = 0
            http_failed = []
            with ThreadPoolExecutor(max_workers=5) as pool:
                futures = {
                    pool.submit(scrape_match_fast, s, match_date=cur): s
                    for s in summaries
                }
                for fut in as_completed(futures):
                    try:
                        r = fut.result()
                        done_count += 1
                        if r.pop('_http_failed', False):
                            http_failed.append(r)
                        elif _row_is_valid(r):
                            rows.append(r)
                            total += 1
                            day_matches += 1
                    except Exception:
                        done_count += 1

            # HTTP basarisiz olanlari Selenium ile tek tek cek
            if http_failed:
                # Ilk macta Selenium dene, o da basarisizsa geri kalan butun
                # maclarda da iddaa tab disabled demektir, hepsini atla
                first = http_failed[0]
                d = _ensure_driver()
                url0 = first.get('iddaa_link', '')
                sel0 = scrape_match_selenium(d, url0) if url0 else {}
                if sel0 and any(sel0.get(k) for k in ('ms1','ms0','ms2')):
                    # Selenium calisiyor, geri kalanlari da cek
                    first.update(sel0)
                    if _row_is_valid(first):
                        rows.append(first)
                        total += 1
                        day_matches += 1
                    print(f'  [{cur:%d.%m.%Y}] {len(http_failed)} mac Selenium fallback...', flush=True)
                    for r in http_failed[1:]:
                        url = r.get('iddaa_link', '')
                        if not url:
                            continue
                        sel_data = scrape_match_selenium(d, url)
                        if sel_data:
                            r.update(sel_data)
                        if _row_is_valid(r):
                            rows.append(r)
                            total += 1
                            day_matches += 1
                    # Ana sayfaya geri don
                    try:
                        pick_date(d, cur)
                        time.sleep(0.3)
                    except Exception:
                        pass
                else:
                    print(f'  [{cur:%d.%m.%Y}] {len(http_failed)} mac iddaa verisi yok (disabled)', flush=True)

            print(f'  -> {day_matches} gecerli mac (toplam: {total})', flush=True)

            # Her 7 gunde bir ara kayit
            if day_count % 7 == 0:
                export_excel(rows, output)
                save_progress(cur, end, total, output_path)
                print(f'  [KAYIT] {total} mac kaydedildi', flush=True)

            cur += dt.timedelta(days=1)

    except KeyboardInterrupt:
        print("\nKullanici durdurdu.", flush=True)
    finally:
        # Son kayit
        if rows:
            export_excel(rows, output)
            save_progress(cur, end, total, output_path)
            print(f'\nToplam: {total} mac -> {output_path}', flush=True)
        if driver:
            with contextlib.suppress(Exception):
                driver.quit()

    return cur  # Son islenmiş tarih


def main():
    parser = argparse.ArgumentParser(description='IddaaPro CLI Scraper')
    parser.add_argument('--start', required=True, help='Baslangic tarihi (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='Bitis tarihi (YYYY-MM-DD)')
    parser.add_argument('--output', default='output/iddaa_data.xlsx', help='Cikti dosyasi')
    parser.add_argument('--max-hours', type=float, default=5.5, help='Max calisma suresi (saat)')
    parser.add_argument('--resume', action='store_true', help='Kaldigi yerden devam et')
    args = parser.parse_args()

    start = dt.datetime.strptime(args.start, '%Y-%m-%d').date()
    end = dt.datetime.strptime(args.end, '%Y-%m-%d').date()

    # Resume: onceki calismayi yukle
    if args.resume:
        prog = load_progress()
        if prog:
            last = dt.datetime.strptime(prog['last_date'], '%Y-%m-%d').date()
            start = last + dt.timedelta(days=1)
            print(f"Devam ediliyor: {start:%d.%m.%Y}'den itibaren ({prog['total_matches']} mac mevcut)", flush=True)

    # Output klasorunu olustur
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)

    run_scraper(start, end, args.output, max_hours=args.max_hours)


if __name__ == '__main__':
    main()
